"""사람 중심 Agent Workflow 관리시트 v3를 생성한다."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

AGENT_DIR = Path(__file__).resolve().parents[1]
OUTPUT_PATH = AGENT_DIR / "docs" / "agent-management-sheet-v3.xlsx"

SHEET_COLUMNS = {
    "Workflow Catalog": [
        "workflow_id",
        "workflow_name",
        "workflow_type",
        "description",
        "example_utterances",
        "entry_step_id",
        "max_risk_level",
        "approval_policy",
        "auth_policy",
        "workflow_version",
        "status",
        "notes",
    ],
    "Workflow Steps": [
        "workflow_id",
        "step_order",
        "step_id",
        "step_name",
        "step_purpose",
        "interaction_mode",
        "tool_id",
        "contract_id",
        "step_risk_level",
        "status",
        "notes",
        "external_action",
        "input_state_keys",
        "output_state_keys",
        "route_summary",
        "validation_result",
    ],
    "Workflow Routes": [
        "workflow_id",
        "from_step_id",
        "route_name",
        "condition_description",
        "to_step_id",
        "status",
        "notes",
    ],
    "Workflow Data Schema": [
        "schema_scope",
        "workflow_id",
        "state_key",
        "data_type",
        "nullable",
        "default_value",
        "description",
        "retention_scope",
        "clear_when",
        "sensitive",
        "log_policy",
        "notes",
    ],
    "Step Data Mapping": [
        "workflow_id",
        "step_id",
        "direction",
        "state_key",
        "contract_field_path",
        "required_at_step",
        "mapping_description",
        "notes",
        "validation_result",
    ],
    "Contract Registry": [
        "contract_id",
        "contract_type",
        "contract_name",
        "transport_target",
        "contract_summary",
        "source_document",
        "source_section",
        "contract_version",
        "status",
    ],
    "Contract Mapping": [
        "workflow_id",
        "step_id",
        "interaction_mode",
        "contract_id",
        "transport_target",
        "contract_version",
    ],
    "Enum Registry": [
        "enum_group",
        "enum_value",
        "display_name",
        "description",
        "source_type",
        "source_document",
        "status",
        "sort_order",
    ],
}

WORKFLOW_ROWS = [
    {
        "workflow_id": "wf_global_agent_entry",
        "workflow_name": "글로벌 Agent 진입",
        "workflow_type": "global",
        "description": "사용자 요청을 안전하게 분류하고 실행 가능한 업무 Workflow로 연결한다.",
        "example_utterances": "모든 사용자 최초 발화",
        "entry_step_id": "run_global_guardrail",
        "max_risk_level": "R5",
        "approval_policy": "none",
        "auth_policy": "none",
        "workflow_version": "0.9.0",
        "status": "review",
        "notes": "업무 Workflow의 공통 진입점",
    },
    {
        "workflow_id": "wf_account_list",
        "workflow_name": "계좌 목록 조회",
        "workflow_type": "inquiry",
        "description": "사용자가 보유한 계좌 목록을 마스킹된 정보로 조회한다.",
        "example_utterances": "내 계좌 보여줘 | 생활비 계좌 찾아줘",
        "entry_step_id": "extract_account_list_slots",
        "max_risk_level": "R1",
        "approval_policy": "none",
        "auth_policy": "none",
        "workflow_version": "0.9.0",
        "status": "review",
        "notes": "잔액은 포함하지 않음",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "workflow_name": "잔액 조회",
        "workflow_type": "inquiry",
        "description": "하나 이상의 사용자 계좌 잔액과 출금 가능 금액을 조회한다.",
        "example_utterances": "생활비 통장 잔액 알려줘 | 내 계좌 잔액 전부 보여줘",
        "entry_step_id": "extract_balance_slots",
        "max_risk_level": "R1",
        "approval_policy": "none",
        "auth_policy": "none",
        "workflow_version": "0.9.0",
        "status": "review",
        "notes": "잔액 계산과 송금 가능 판단은 수행하지 않음",
    },
    {
        "workflow_id": "wf_transaction_history",
        "workflow_name": "거래내역 조회",
        "workflow_type": "inquiry",
        "description": "계좌, 기간과 검색 조건에 맞는 거래내역 첫 페이지를 조회한다.",
        "example_utterances": "지난달 거래내역 보여줘 | 이번 주 카드 결제 내역 찾아줘",
        "entry_step_id": "extract_transaction_slots",
        "max_risk_level": "R1",
        "approval_policy": "none",
        "auth_policy": "none",
        "workflow_version": "0.9.0",
        "status": "review",
        "notes": "첫 페이지 이후 탐색은 Frontend와 Backend가 처리",
    },
    {
        "workflow_id": "wf_period_amount_summary",
        "workflow_name": "기간 거래 합계 조회",
        "workflow_type": "inquiry",
        "description": "선택한 계좌와 기간의 입금 또는 출금 합계를 조회한다.",
        "example_utterances": "이번 달 지출 합계 알려줘 | 지난주 입금액 합쳐줘",
        "entry_step_id": "extract_amount_summary_slots",
        "max_risk_level": "R2",
        "approval_policy": "none",
        "auth_policy": "none",
        "workflow_version": "0.9.0",
        "status": "review",
        "notes": "Agent가 거래내역을 직접 합산하지 않음",
    },
    {
        "workflow_id": "wf_set_default_account",
        "workflow_name": "기본 출금 계좌 변경",
        "workflow_type": "setting_change",
        "description": "사용자가 선택한 계좌를 기본 출금 계좌로 변경한다.",
        "example_utterances": "급여 통장을 기본계좌로 바꿔줘",
        "entry_step_id": "extract_default_account_slots",
        "max_risk_level": "R3",
        "approval_policy": "required",
        "auth_policy": "none",
        "workflow_version": "0.9.0",
        "status": "review",
        "notes": "사용자 승인 후 실행",
    },
    {
        "workflow_id": "wf_set_account_alias",
        "workflow_name": "계좌 별칭 변경",
        "workflow_type": "setting_change",
        "description": "사용자 계좌의 별칭을 변경한다.",
        "example_utterances": "급여 계좌 별명을 월급 통장으로 바꿔줘",
        "entry_step_id": "extract_account_alias_slots",
        "max_risk_level": "R3",
        "approval_policy": "required",
        "auth_policy": "none",
        "workflow_version": "0.9.0",
        "status": "review",
        "notes": "사용자 승인 후 실행",
    },
    {
        "workflow_id": "wf_internal_transfer",
        "workflow_name": "본인 계좌 간 이체",
        "workflow_type": "transfer",
        "description": "사용자 본인 소유 계좌 사이에서 자금을 이체한다.",
        "example_utterances": "생활비 통장에서 저축 통장으로 10만 원 옮겨줘",
        "entry_step_id": "extract_internal_transfer_slots",
        "max_risk_level": "R4",
        "approval_policy": "required",
        "auth_policy": "required",
        "workflow_version": "0.9.0",
        "status": "review",
        "notes": "승인과 추가 인증 모두 필수",
    },
    {
        "workflow_id": "wf_external_transfer",
        "workflow_name": "타인송금",
        "workflow_type": "transfer",
        "description": "검증된 수취인에게 사용자 계좌의 자금을 송금한다.",
        "example_utterances": "홍길동에게 5만 원 보내줘",
        "entry_step_id": "extract_external_transfer_slots",
        "max_risk_level": "R4",
        "approval_policy": "required",
        "auth_policy": "required",
        "workflow_version": "0.9.0",
        "status": "review",
        "notes": "승인과 추가 인증 모두 필수",
    },
]

API_CONTRACT_ROWS = [
    ("API-ACCOUNT-LIST", "계좌 목록 조회", "GET /api/v1/agent-tools/accounts", "계좌 후보 조회", "9장"),
    ("API-BALANCE-QUERY", "잔액 조회", "POST /api/v1/agent-tools/accounts/balances:query", "복수 계좌 잔액 조회", "10장"),
    ("API-TRANSACTION-QUERY", "거래내역 조회", "POST /api/v1/agent-tools/transactions:query", "거래내역 첫 페이지 조회", "11장"),
    ("API-TRANSACTION-SUMMARY", "거래 합계 조회", "POST /api/v1/agent-tools/transactions:summary", "기간 거래 합계 조회", "12장"),
    ("API-RECIPIENT-RESOLVE", "기존 수취인 자동 확정", "POST /api/v1/agent-tools/recipients:resolve", "이름 힌트로 기존 거래 수취인 확정", "13장"),
    ("API-EXTERNAL-TRANSFER-PREPARE", "타인송금 Prepare", "POST /api/v1/agent-tools/transfers/external:prepare", "타인송금 조건 사전 평가", "14장"),
    ("API-AUTH-CONTEXT-CREATE", "추가 인증 Context 생성", "POST /api/v1/agent-tools/auth-contexts", "송금 추가 인증 Context 생성", "15장"),
    ("API-EXTERNAL-TRANSFER-EXECUTE", "타인송금 Execute", "POST /api/v1/agent-tools/transfers/external", "승인·인증된 타인송금 실행", "16장"),
    ("API-INTERNAL-TRANSFER-PREPARE", "본인 이체 Prepare", "POST /api/v1/agent-tools/transfers/internal:prepare", "본인 이체 조건 사전 평가", "17장"),
    ("API-INTERNAL-TRANSFER-EXECUTE", "본인 이체 Execute", "POST /api/v1/agent-tools/transfers/internal", "승인·인증된 본인 이체 실행", "18장"),
    ("API-DEFAULT-ACCOUNT-PREPARE", "기본계좌 변경 Prepare", "POST /api/v1/agent-tools/settings/default-account:prepare", "기본 출금 계좌 변경 조건 평가", "19장"),
    ("API-DEFAULT-ACCOUNT-EXECUTE", "기본계좌 변경 Execute", "POST /api/v1/agent-tools/settings/default-account", "승인된 기본 출금 계좌 변경", "20장"),
    ("API-ACCOUNT-ALIAS-PREPARE", "계좌 별칭 변경 Prepare", "POST /api/v1/agent-tools/settings/account-alias:prepare", "계좌 별칭 변경 조건 평가", "21장"),
    ("API-ACCOUNT-ALIAS-EXECUTE", "계좌 별칭 변경 Execute", "POST /api/v1/agent-tools/settings/account-alias", "승인된 계좌 별칭 변경", "22장"),
]

CONTRACT_ROWS = [
    {
        "contract_id": contract_id,
        "contract_type": "agent_tool_api",
        "contract_name": contract_name,
        "transport_target": transport_target,
        "contract_summary": contract_summary,
        "source_document": "agent-tools-api-spec.md",
        "source_section": source_section,
        "contract_version": "0.9.0",
        "status": "review",
    }
    for contract_id, contract_name, transport_target, contract_summary, source_section in API_CONTRACT_ROWS
]

UI_CONTRACT_ROWS = [
    {
        "contract_id": "UI-ACCOUNT-LIST-RESULT",
        "contract_type": "ui_hitl",
        "contract_name": "계좌 목록 결과",
        "transport_target": "component · account_list",
        "contract_summary": "마스킹된 사용자 계좌 목록 표시",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "4.1장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-BALANCE-ACCOUNT-SELECTION",
        "contract_type": "ui_hitl",
        "contract_name": "잔액 조회 계좌 선택",
        "transport_target": "component · account_card_list",
        "contract_summary": "잔액을 조회할 단일·복수 계좌 선택 또는 빈 상태 표시",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "3.3장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-BALANCE-RESULT",
        "contract_type": "ui_hitl",
        "contract_name": "잔액 조회 결과",
        "transport_target": "component · balance_result",
        "contract_summary": "단일·복수 계좌의 잔액과 출금 가능 금액 표시",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "4.2장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-TRANSACTION-ACCOUNT-SELECTION",
        "contract_type": "ui_hitl",
        "contract_name": "거래내역 조회 계좌 선택",
        "transport_target": "component · account_card_list",
        "contract_summary": "거래내역을 조회할 단일·복수 계좌 선택 또는 빈 상태 표시",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "3.3장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-PERIOD-SELECTION",
        "contract_type": "ui_hitl",
        "contract_name": "조회 기간 선택",
        "transport_target": "component · period_input",
        "contract_summary": "기간 프리셋 또는 직접 날짜 범위 선택",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "3.5장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-TRANSACTION-LIST",
        "contract_type": "ui_hitl",
        "contract_name": "거래내역 목록 결과",
        "transport_target": "component · transaction_list",
        "contract_summary": "거래내역 첫 페이지와 이후 조회 Context 표시",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "4.3장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-SUMMARY-ACCOUNT-SELECTION",
        "contract_type": "ui_hitl",
        "contract_name": "합계 조회 계좌 선택",
        "transport_target": "component · account_card_list",
        "contract_summary": "합계를 조회할 단일·복수 계좌 선택 또는 빈 상태 표시",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "3.3장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-SUMMARY-TYPE-SELECTION",
        "contract_type": "ui_hitl",
        "contract_name": "합계 유형 선택",
        "transport_target": "component · option_select",
        "contract_summary": "지출 또는 수입 합계 유형 선택",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "3.6장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-AMOUNT-SUMMARY",
        "contract_type": "ui_hitl",
        "contract_name": "기간 거래 합계 결과",
        "transport_target": "component · amount_summary",
        "contract_summary": "기간·계좌 범위의 지출 또는 수입 합계 표시",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "4.4장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-DEFAULT-ACCOUNT-SELECTION",
        "contract_type": "ui_hitl",
        "contract_name": "기본 출금 계좌 선택",
        "transport_target": "component · account_card_list",
        "contract_summary": "새 기본 출금 계좌 단일 선택 또는 빈 상태 표시",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "3.3장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-DEFAULT-ACCOUNT-CONFIRMATION",
        "contract_type": "ui_hitl",
        "contract_name": "기본 출금 계좌 변경 승인",
        "transport_target": "component · confirm_modal",
        "contract_summary": "현재 계좌와 새 기본 출금 계좌를 확인하고 승인·수정·취소",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "3.7장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-DEFAULT-ACCOUNT-RESULT",
        "contract_type": "ui_hitl",
        "contract_name": "기본 출금 계좌 변경 결과",
        "transport_target": "component · setting_result",
        "contract_summary": "변경 완료 또는 이미 설정된 기본 출금 계좌 결과 표시",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "4.6장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-ACCOUNT-ALIAS-SELECTION",
        "contract_type": "ui_hitl",
        "contract_name": "별칭 변경 계좌 선택",
        "transport_target": "component · account_card_list",
        "contract_summary": "별칭을 변경할 계좌 단일 선택 또는 빈 상태 표시",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "3.3장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-ACCOUNT-ALIAS-INPUT",
        "contract_type": "ui_hitl",
        "contract_name": "새 계좌 별칭 입력",
        "transport_target": "component · text_input",
        "contract_summary": "Backend 정책 검증을 거치는 새 계좌 별칭 입력",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "3.1장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-ACCOUNT-ALIAS-CONFIRMATION",
        "contract_type": "ui_hitl",
        "contract_name": "계좌 별칭 변경 승인",
        "transport_target": "component · confirm_modal",
        "contract_summary": "대상 계좌와 최종 별칭을 확인하고 승인·수정·취소",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "3.7장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-ACCOUNT-ALIAS-RESULT",
        "contract_type": "ui_hitl",
        "contract_name": "계좌 별칭 변경 결과",
        "transport_target": "component · setting_result",
        "contract_summary": "별칭 변경 완료 또는 이미 같은 별칭인 결과 표시",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "4.6장",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-SETTING-BLOCKED",
        "contract_type": "ui_hitl",
        "contract_name": "설정 변경 차단 안내",
        "transport_target": "blocked · blocked_message",
        "contract_summary": "사용자 수정으로 해결할 수 없는 설정 변경 차단 안내",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "UI Component Registry",
        "contract_version": "0.9.0",
        "status": "review",
    },
    {
        "contract_id": "UI-COMMON-ERROR",
        "contract_type": "ui_hitl",
        "contract_name": "공통 오류 안내",
        "transport_target": "error · error_message",
        "contract_summary": "사용자에게 공개 가능한 공통 오류 안내",
        "source_document": "agent-ui-hitl-contract.md",
        "source_section": "UI Component Registry",
        "contract_version": "0.9.0",
        "status": "review",
    },
]

INTERNAL_TRANSFER_UI_CONTRACT_DEFINITIONS = [
    ("UI-INTERNAL-TRANSFER-FROM-ACCOUNT", "본인송금 출금 계좌 선택", "component · account_card_list", "본인송금 출금 계좌 단일 선택 또는 빈 상태 표시"),
    ("UI-INTERNAL-TRANSFER-TO-ACCOUNT", "본인송금 입금 계좌 선택", "component · account_card_list", "출금 계좌를 제외한 본인 입금 계좌 단일 선택 또는 빈 상태 표시"),
    ("UI-TRANSFER-AMOUNT-INPUT", "송금 금액 입력", "component · number_input", "Backend가 검증한 송금 금액 입력 또는 취소"),
    ("UI-INTERNAL-TRANSFER-CONFIRMATION", "본인송금 승인", "component · confirm_modal", "출금 계좌·입금 계좌·금액 확인 후 승인·수정·취소"),
    ("UI-INTERNAL-TRANSFER-CORRECTION", "본인송금 수정 대상 선택", "component · option_select", "Backend가 허용한 복수 수정 대상 중 하나 선택"),
    ("UI-INTERNAL-TRANSFER-AUTH", "본인송금 추가 인증", "component · auth_request", "Backend 추가 인증 시작 및 결과 대기"),
    ("UI-INTERNAL-TRANSFER-AUTH-RETRY", "본인송금 재인증 선택", "component · option_select", "인증 실패·만료 후 재시도 또는 취소 선택"),
    ("UI-INTERNAL-TRANSFER-RESULT", "본인송금 완료 결과", "component · transfer_result", "거래 ID와 Prepare 표시 데이터를 조합한 본인송금 완료 결과"),
    ("UI-TRANSFER-BLOCKED", "송금 차단 안내", "blocked · blocked_message", "Backend가 제공한 사용자 표시용 송금 차단 안내"),
]

UI_CONTRACT_ROWS.extend(
    [
        {
            "contract_id": contract_id,
            "contract_type": "ui_hitl",
            "contract_name": contract_name,
            "transport_target": transport_target,
            "contract_summary": contract_summary,
            "source_document": "agent-ui-hitl-contract.md",
            "source_section": "3장, 5.8장",
            "contract_version": "0.9.0",
            "status": "review",
        }
        for contract_id, contract_name, transport_target, contract_summary in INTERNAL_TRANSFER_UI_CONTRACT_DEFINITIONS
    ]
)

EXTERNAL_TRANSFER_UI_CONTRACT_DEFINITIONS = [
    ("UI-RECIPIENT-SELECT", "타인송금 수취인 선택", "component · recipient_select", "기존 수취인 선택 또는 Backend 신규 계좌 검증·확정"),
    ("UI-EXTERNAL-TRANSFER-FROM-ACCOUNT", "타인송금 출금 계좌 선택", "component · account_card_list", "타인송금 출금 계좌 단일 선택 또는 빈 상태 표시"),
    ("UI-EXTERNAL-TRANSFER-CONFIRMATION", "타인송금 승인", "component · confirm_modal", "출금 계좌·수취인·금액 확인 후 승인·수정·취소"),
    ("UI-EXTERNAL-TRANSFER-CORRECTION", "타인송금 수정 대상 선택", "component · option_select", "Backend가 허용한 복수 수정 대상 중 하나 선택"),
    ("UI-EXTERNAL-TRANSFER-AUTH", "타인송금 추가 인증", "component · auth_request", "Backend 추가 인증 시작 및 결과 대기"),
    ("UI-EXTERNAL-TRANSFER-AUTH-RETRY", "타인송금 재인증 선택", "component · option_select", "인증 실패·만료 후 재시도 또는 취소 선택"),
    ("UI-EXTERNAL-TRANSFER-RESULT", "타인송금 완료 결과", "component · transfer_result", "거래 ID와 Prepare 표시 데이터를 조합한 타인송금 완료 결과"),
]

UI_CONTRACT_ROWS.extend(
    [
        {
            "contract_id": contract_id,
            "contract_type": "ui_hitl",
            "contract_name": contract_name,
            "transport_target": transport_target,
            "contract_summary": contract_summary,
            "source_document": "agent-ui-hitl-contract.md",
            "source_section": "3장, 5.9장",
            "contract_version": "0.9.0",
            "status": "review",
        }
        for contract_id, contract_name, transport_target, contract_summary in EXTERNAL_TRANSFER_UI_CONTRACT_DEFINITIONS
    ]
)

GLOBAL_UI_CONTRACT_DEFINITIONS = [
    (
        "UI-GLOBAL-BLOCKED",
        "전역 정책 차단 안내",
        "blocked · blocked_message",
        "전역 가드레일에서 허용하지 않는 요청에 대한 사용자 안내",
    ),
    (
        "UI-NO-MATCH",
        "지원 Workflow 없음 안내",
        "message · message",
        "지원하는 금융 Workflow와 일치하지 않는 요청에 대한 안내",
    ),
]

UI_CONTRACT_ROWS.extend(
    [
        {
            "contract_id": contract_id,
            "contract_type": "ui_hitl",
            "contract_name": contract_name,
            "transport_target": transport_target,
            "contract_summary": contract_summary,
            "source_document": "agent-ui-hitl-contract.md",
            "source_section": "5.1장",
            "contract_version": "0.9.0",
            "status": "review",
        }
        for contract_id, contract_name, transport_target, contract_summary in GLOBAL_UI_CONTRACT_DEFINITIONS
    ]
)

CONTRACT_ROWS.extend(UI_CONTRACT_ROWS)

GLOBAL_STEP_DEFINITIONS = [
    (1, "run_global_guardrail", "전역 요청 가드레일", "사용자 요청이 Agent가 처리할 수 있는 정책 범위인지 분류한다.", "agent_internal", "run_global_guardrail", "", "R0", "금융 업무의 실행 가능 여부나 원장 상태는 판단하지 않음", "", "", "guardrail_outcome, blocked_view", "allowed → match_workflow | blocked → emit_global_blocked"),
    (2, "match_workflow", "Workflow 매칭", "허용된 사용자 요청을 지원하는 업무 Workflow 중 하나로 분류한다.", "agent_internal", "match_workflow", "", "R0", "업무별 세부 입력값은 하위 Workflow에서 추출", "", "", "workflow_match_outcome, matched_workflow_id", "matched → dispatch_matched_workflow | no_match → emit_no_matching_workflow"),
    (3, "dispatch_matched_workflow", "매칭 Workflow 실행", "매칭된 하위 Workflow를 시작하고 종료 상태를 확인한다.", "agent_internal", "dispatch_matched_workflow", "", "R0", "하위 Workflow가 업무 결과 Webhook을 직접 전송", "", "matched_workflow_id", "dispatch_outcome", "completed → END | failed → emit_workflow_dispatch_error"),
    (4, "emit_global_blocked", "전역 정책 차단 안내", "가드레일 차단 결과를 사용자 표시용 Webhook으로 전송한다.", "webhook", "emit_blocked", "UI-GLOBAL-BLOCKED", "R0", "내부 정책 세부 사유는 노출하지 않음", "blocked · blocked_message", "blocked_view", "", "항상 → END"),
    (5, "emit_no_matching_workflow", "지원 Workflow 없음 안내", "지원 Workflow와 일치하지 않는 요청에 대한 안내를 Webhook으로 전송한다.", "webhook", "emit_message", "UI-NO-MATCH", "R0", "지원 가능한 요청 범위를 정적 안내", "message · message", "", "", "항상 → END"),
    (6, "emit_workflow_dispatch_error", "Workflow 실행 오류 안내", "하위 Workflow 시작 실패 또는 처리되지 않은 인프라 오류를 공통 오류 Webhook으로 전송한다.", "webhook", "emit_error", "UI-COMMON-ERROR", "R0", "하위 Workflow가 정상 종료한 업무 오류에는 중복 전송하지 않음", "error · error_message", "", "", "항상 → END"),
]

GLOBAL_STEP_ROWS = [
    {
        "workflow_id": "wf_global_agent_entry",
        "step_order": step_order,
        "step_id": step_id,
        "step_name": step_name,
        "step_purpose": step_purpose,
        "interaction_mode": interaction_mode,
        "tool_id": tool_id,
        "contract_id": contract_id,
        "step_risk_level": step_risk_level,
        "status": "review",
        "notes": notes,
        "external_action": external_action,
        "input_state_keys": input_state_keys,
        "output_state_keys": output_state_keys,
        "route_summary": route_summary,
        "validation_result": "OK",
    }
    for (
        step_order,
        step_id,
        step_name,
        step_purpose,
        interaction_mode,
        tool_id,
        contract_id,
        step_risk_level,
        notes,
        external_action,
        input_state_keys,
        output_state_keys,
        route_summary,
    ) in GLOBAL_STEP_DEFINITIONS
]

GLOBAL_ROUTE_DEFINITIONS = [
    ("run_global_guardrail", "정책 허용", "guardrail_outcome=allowed인 경우", "match_workflow", ""),
    ("run_global_guardrail", "정책 차단", "guardrail_outcome=blocked이고 사용자 표시용 blocked_view가 생성된 경우", "emit_global_blocked", ""),
    ("match_workflow", "Workflow 매칭 완료", "workflow_match_outcome=matched이고 지원 Workflow ID가 확정된 경우", "dispatch_matched_workflow", ""),
    ("match_workflow", "일치 Workflow 없음", "workflow_match_outcome=no_match인 경우", "emit_no_matching_workflow", ""),
    ("dispatch_matched_workflow", "하위 Workflow 정상 종료", "dispatch_outcome=completed인 경우", "END", "하위 Workflow의 업무 오류 종료도 포함하며 전역 오류를 중복 전송하지 않음"),
    ("dispatch_matched_workflow", "하위 Workflow 실행 실패", "dispatch_outcome=failed인 경우", "emit_workflow_dispatch_error", "시작 실패 또는 처리되지 않은 인프라 예외만 해당"),
    ("emit_global_blocked", "차단 안내 완료", "전역 정책 차단 Webhook 전송을 완료한 경우", "END", ""),
    ("emit_no_matching_workflow", "안내 완료", "지원 Workflow 없음 Webhook 전송을 완료한 경우", "END", ""),
    ("emit_workflow_dispatch_error", "오류 안내 완료", "공통 오류 Webhook 전송을 완료한 경우", "END", ""),
]

GLOBAL_ROUTE_ROWS = [
    {
        "workflow_id": "wf_global_agent_entry",
        "from_step_id": from_step_id,
        "route_name": route_name,
        "condition_description": condition_description,
        "to_step_id": to_step_id,
        "status": "review",
        "notes": notes,
    }
    for from_step_id, route_name, condition_description, to_step_id, notes in GLOBAL_ROUTE_DEFINITIONS
]

ACCOUNT_LIST_STEP_ROWS = [
    {
        "workflow_id": "wf_account_list",
        "step_order": 1,
        "step_id": "extract_account_list_slots",
        "step_name": "계좌 목록 조회 조건 추출",
        "step_purpose": "사용자 발화에서 선택적인 계좌 검색 힌트를 추출한다.",
        "interaction_mode": "agent_internal",
        "tool_id": "extract_account_list_slots",
        "contract_id": "",
        "step_risk_level": "R1",
        "status": "review",
        "notes": "힌트가 없으면 account_hint를 null로 유지",
        "external_action": "",
        "input_state_keys": "",
        "output_state_keys": "account_hint",
        "route_summary": "항상 → fetch_account_list",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_account_list",
        "step_order": 2,
        "step_id": "fetch_account_list",
        "step_name": "계좌 목록 조회",
        "step_purpose": "Backend Tool API로 사용자의 마스킹된 계좌 목록을 조회한다.",
        "interaction_mode": "backend_tool_api",
        "tool_id": "fetch_accounts",
        "contract_id": "API-ACCOUNT-LIST",
        "step_risk_level": "R1",
        "status": "review",
        "notes": "limit=20은 Tool 설정값으로 사용하고 account_capability는 생략",
        "external_action": "GET /api/v1/agent-tools/accounts",
        "input_state_keys": "account_hint",
        "output_state_keys": "account_results",
        "route_summary": "정상 응답 → emit_account_list_result | 오류 → emit_account_list_error",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_account_list",
        "step_order": 3,
        "step_id": "emit_account_list_result",
        "step_name": "계좌 목록 결과 전송",
        "step_purpose": "마스킹된 계좌 목록을 결과 Webhook으로 전송한다.",
        "interaction_mode": "webhook",
        "tool_id": "emit_component",
        "contract_id": "UI-ACCOUNT-LIST-RESULT",
        "step_risk_level": "R0",
        "status": "review",
        "notes": "사용자 회신을 기다리지 않음",
        "external_action": "component · account_list",
        "input_state_keys": "account_results",
        "output_state_keys": "",
        "route_summary": "항상 → END",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_account_list",
        "step_order": 4,
        "step_id": "emit_account_list_error",
        "step_name": "계좌 목록 조회 오류 안내",
        "step_purpose": "조회 요청을 완료하지 못한 경우 사용자에게 공개 가능한 오류를 안내한다.",
        "interaction_mode": "webhook",
        "tool_id": "emit_error",
        "contract_id": "UI-COMMON-ERROR",
        "step_risk_level": "R0",
        "status": "review",
        "notes": "내부 오류와 Stack Trace를 포함하지 않음",
        "external_action": "error · error_message",
        "input_state_keys": "",
        "output_state_keys": "",
        "route_summary": "항상 → END",
        "validation_result": "OK",
    },
]

ACCOUNT_LIST_ROUTE_ROWS = [
    {
        "workflow_id": "wf_account_list",
        "from_step_id": "extract_account_list_slots",
        "route_name": "조회 조건 추출 완료",
        "condition_description": "사용자 발화에서 계좌 검색 힌트 추출을 마친 경우",
        "to_step_id": "fetch_account_list",
        "status": "review",
        "notes": "힌트가 없어도 조회를 계속함",
    },
    {
        "workflow_id": "wf_account_list",
        "from_step_id": "fetch_account_list",
        "route_name": "조회 성공",
        "condition_description": "Backend가 사용자 계좌 목록을 정상적으로 반환한 경우. 계좌가 없으면 accounts는 빈 배열이다.",
        "to_step_id": "emit_account_list_result",
        "status": "review",
        "notes": "accounts=[]도 동일한 account_list UI로 전송",
    },
    {
        "workflow_id": "wf_account_list",
        "from_step_id": "fetch_account_list",
        "route_name": "조회 오류",
        "condition_description": "권한 또는 기술 오류로 Backend가 계약된 계좌 목록 결과를 반환하지 못한 경우",
        "to_step_id": "emit_account_list_error",
        "status": "review",
        "notes": "제한적 재시도 이후 오류 안내",
    },
    {
        "workflow_id": "wf_account_list",
        "from_step_id": "emit_account_list_result",
        "route_name": "결과 전송 완료",
        "condition_description": "계좌 목록 결과 Webhook 전송을 완료한 경우",
        "to_step_id": "END",
        "status": "review",
        "notes": "",
    },
    {
        "workflow_id": "wf_account_list",
        "from_step_id": "emit_account_list_error",
        "route_name": "오류 안내 완료",
        "condition_description": "오류 안내 Webhook 전송을 완료한 경우",
        "to_step_id": "END",
        "status": "review",
        "notes": "",
    },
]

BALANCE_STEP_ROWS = [
    {
        "workflow_id": "wf_balance_inquiry",
        "step_order": 1,
        "step_id": "extract_balance_slots",
        "step_name": "잔액 조회 조건 추출",
        "step_purpose": "사용자 발화에서 계좌 힌트와 전체 계좌 조회 의도를 추출한다.",
        "interaction_mode": "agent_internal",
        "tool_id": "extract_balance_slots",
        "contract_id": "",
        "step_risk_level": "R1",
        "status": "review",
        "notes": "계좌를 직접 확정하거나 검증하지 않음",
        "external_action": "",
        "input_state_keys": "",
        "output_state_keys": "account_hint, all_accounts_requested",
        "route_summary": "항상 → resolve_balance_accounts",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_order": 2,
        "step_id": "resolve_balance_accounts",
        "step_name": "잔액 조회 계좌 확인",
        "step_purpose": "Backend가 조회 가능한 계좌를 검증하고 자동 확정 여부를 결정한다.",
        "interaction_mode": "backend_tool_api",
        "tool_id": "fetch_accounts",
        "contract_id": "API-ACCOUNT-LIST",
        "step_risk_level": "R1",
        "status": "review",
        "notes": "resolve_selection=true, account_capability=inquiry는 Tool 설정값",
        "external_action": "GET /api/v1/agent-tools/accounts",
        "input_state_keys": "account_hint, all_accounts_requested",
        "output_state_keys": "account_resolution_outcome, accounts, account_ids",
        "route_summary": "resolved → query_balances | selection_required → request_balance_account_selection | no_accounts → emit_balance_accounts_empty | 오류 → emit_balance_error",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_order": 3,
        "step_id": "request_balance_account_selection",
        "step_name": "잔액 조회 계좌 선택 요청",
        "step_purpose": "검증된 복수 계좌 후보를 표시하고 사용자 선택 또는 취소 결과를 기다린다.",
        "interaction_mode": "webhook_then_resume",
        "tool_id": "request_component_input",
        "contract_id": "UI-BALANCE-ACCOUNT-SELECTION",
        "step_risk_level": "R1",
        "status": "review",
        "notes": "Backend가 선택값을 검증한 뒤 Agent를 resume",
        "external_action": "component · account_card_list",
        "input_state_keys": "accounts",
        "output_state_keys": "account_selection_outcome, account_ids",
        "route_summary": "selected → query_balances | cancelled → END",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_order": 4,
        "step_id": "emit_balance_accounts_empty",
        "step_name": "잔액 조회 계좌 없음 안내",
        "step_purpose": "조회 가능한 계좌가 없는 빈 상태를 동일한 계좌 선택 UI로 전송한다.",
        "interaction_mode": "webhook",
        "tool_id": "emit_component",
        "contract_id": "UI-BALANCE-ACCOUNT-SELECTION",
        "step_risk_level": "R0",
        "status": "review",
        "notes": "accounts=[]을 전송하고 사용자 회신을 기다리지 않음",
        "external_action": "component · account_card_list",
        "input_state_keys": "accounts",
        "output_state_keys": "",
        "route_summary": "항상 → END",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_order": 5,
        "step_id": "query_balances",
        "step_name": "잔액 일괄 조회",
        "step_purpose": "Backend Tool API로 검증된 단일·복수 계좌의 잔액을 일괄 조회한다.",
        "interaction_mode": "backend_tool_api",
        "tool_id": "query_balances",
        "contract_id": "API-BALANCE-QUERY",
        "step_risk_level": "R1",
        "status": "review",
        "notes": "계좌별 반복 호출을 수행하지 않음",
        "external_action": "POST /api/v1/agent-tools/accounts/balances:query",
        "input_state_keys": "account_ids",
        "output_state_keys": "balance_results",
        "route_summary": "성공 → emit_balance_result | 오류 → emit_balance_error",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_order": 6,
        "step_id": "emit_balance_result",
        "step_name": "잔액 조회 결과 전송",
        "step_purpose": "Backend가 반환한 잔액 결과를 결과 Webhook으로 전송한다.",
        "interaction_mode": "webhook",
        "tool_id": "emit_component",
        "contract_id": "UI-BALANCE-RESULT",
        "step_risk_level": "R0",
        "status": "review",
        "notes": "잔액 원문을 일반 로그에 기록하지 않음",
        "external_action": "component · balance_result",
        "input_state_keys": "balance_results",
        "output_state_keys": "",
        "route_summary": "항상 → END",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_order": 7,
        "step_id": "emit_balance_error",
        "step_name": "잔액 조회 오류 안내",
        "step_purpose": "잔액 조회를 완료하지 못한 경우 Backend의 사용자 공개 가능 오류를 안내한다.",
        "interaction_mode": "webhook",
        "tool_id": "emit_error",
        "contract_id": "UI-COMMON-ERROR",
        "step_risk_level": "R0",
        "status": "review",
        "notes": "Agent가 오류 유형별 복구 흐름을 임의로 선택하지 않음",
        "external_action": "error · error_message",
        "input_state_keys": "",
        "output_state_keys": "",
        "route_summary": "항상 → END",
        "validation_result": "OK",
    },
]

BALANCE_ROUTE_ROWS = [
    {
        "workflow_id": "wf_balance_inquiry",
        "from_step_id": "extract_balance_slots",
        "route_name": "조회 조건 추출 완료",
        "condition_description": "계좌 힌트와 전체 계좌 조회 의도 추출을 마친 경우",
        "to_step_id": "resolve_balance_accounts",
        "status": "review",
        "notes": "힌트가 없어도 Backend 계좌 확인을 계속함",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "from_step_id": "resolve_balance_accounts",
        "route_name": "계좌 자동 확정",
        "condition_description": "Backend가 account_resolution_outcome=resolved와 검증된 account_ids를 반환한 경우",
        "to_step_id": "query_balances",
        "status": "review",
        "notes": "Agent가 후보 개수로 다시 판단하지 않음",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "from_step_id": "resolve_balance_accounts",
        "route_name": "사용자 선택 필요",
        "condition_description": "Backend가 account_resolution_outcome=selection_required와 검증된 계좌 후보를 반환한 경우",
        "to_step_id": "request_balance_account_selection",
        "status": "review",
        "notes": "단일·복수 선택 지원",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "from_step_id": "resolve_balance_accounts",
        "route_name": "조회 가능 계좌 없음",
        "condition_description": "Backend가 account_resolution_outcome=no_accounts와 accounts=[]을 반환한 경우",
        "to_step_id": "emit_balance_accounts_empty",
        "status": "review",
        "notes": "오류가 아닌 정상 빈 상태",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "from_step_id": "resolve_balance_accounts",
        "route_name": "계좌 확인 오류",
        "condition_description": "Backend가 계약된 계좌 확인 결과를 반환하지 못한 경우",
        "to_step_id": "emit_balance_error",
        "status": "review",
        "notes": "사용자 공개 가능 오류만 전달",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "from_step_id": "request_balance_account_selection",
        "route_name": "계좌 선택 완료",
        "condition_description": "Backend가 account_selection_outcome=selected와 검증된 account_ids로 Agent를 resume한 경우",
        "to_step_id": "query_balances",
        "status": "review",
        "notes": "Agent가 선택값을 다시 검증하지 않음",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "from_step_id": "request_balance_account_selection",
        "route_name": "계좌 선택 취소",
        "condition_description": "Backend가 account_selection_outcome=cancelled로 Agent를 resume한 경우",
        "to_step_id": "END",
        "status": "review",
        "notes": "추가 Tool API와 UI Webhook 없이 종료",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "from_step_id": "query_balances",
        "route_name": "잔액 조회 성공",
        "condition_description": "Backend가 모든 요청 계좌의 balance_results를 정상적으로 반환한 경우",
        "to_step_id": "emit_balance_result",
        "status": "review",
        "notes": "부분 결과를 허용하지 않음",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "from_step_id": "query_balances",
        "route_name": "잔액 조회 오류",
        "condition_description": "Backend가 계좌 유효성, 접근 권한 또는 기술 오류를 반환한 경우",
        "to_step_id": "emit_balance_error",
        "status": "review",
        "notes": "Agent가 오류 유형별 복구를 판단하지 않음",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "from_step_id": "emit_balance_accounts_empty",
        "route_name": "빈 상태 전송 완료",
        "condition_description": "조회 가능 계좌 없음 Webhook 전송을 완료한 경우",
        "to_step_id": "END",
        "status": "review",
        "notes": "",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "from_step_id": "emit_balance_result",
        "route_name": "결과 전송 완료",
        "condition_description": "잔액 조회 결과 Webhook 전송을 완료한 경우",
        "to_step_id": "END",
        "status": "review",
        "notes": "",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "from_step_id": "emit_balance_error",
        "route_name": "오류 안내 완료",
        "condition_description": "잔액 조회 오류 안내 Webhook 전송을 완료한 경우",
        "to_step_id": "END",
        "status": "review",
        "notes": "",
    },
]

TRANSACTION_STEP_ROWS = [
    {
        "workflow_id": "wf_transaction_history",
        "step_order": 1,
        "step_id": "extract_transaction_slots",
        "step_name": "거래내역 조회 조건 추출",
        "step_purpose": "사용자 발화에서 계좌, 기간, 검색어와 거래 유형 힌트를 추출한다.",
        "interaction_mode": "agent_internal",
        "tool_id": "extract_transaction_slots",
        "contract_id": "",
        "step_risk_level": "R1",
        "status": "review",
        "notes": "계좌를 직접 확정하거나 금융 조회 조건을 검증하지 않음",
        "external_action": "",
        "input_state_keys": "",
        "output_state_keys": "account_hint, all_accounts_requested, start_date, end_date, keyword, transaction_type",
        "route_summary": "항상 → resolve_transaction_accounts",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_transaction_history",
        "step_order": 2,
        "step_id": "resolve_transaction_accounts",
        "step_name": "거래내역 조회 계좌 확인",
        "step_purpose": "Backend가 조회 가능한 계좌를 검증하고 자동 확정 여부를 결정한다.",
        "interaction_mode": "backend_tool_api",
        "tool_id": "fetch_accounts",
        "contract_id": "API-ACCOUNT-LIST",
        "step_risk_level": "R1",
        "status": "review",
        "notes": "resolve_selection=true, account_capability=inquiry는 Tool 설정값",
        "external_action": "GET /api/v1/agent-tools/accounts",
        "input_state_keys": "account_hint, all_accounts_requested",
        "output_state_keys": "account_resolution_outcome, accounts, account_ids",
        "route_summary": "resolved → check_transaction_period | selection_required → request_transaction_account_selection | no_accounts → emit_transaction_accounts_empty | 오류 → emit_transaction_error",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_transaction_history",
        "step_order": 3,
        "step_id": "request_transaction_account_selection",
        "step_name": "거래내역 조회 계좌 선택 요청",
        "step_purpose": "검증된 복수 계좌 후보를 표시하고 사용자 선택 또는 취소 결과를 기다린다.",
        "interaction_mode": "webhook_then_resume",
        "tool_id": "request_component_input",
        "contract_id": "UI-TRANSACTION-ACCOUNT-SELECTION",
        "step_risk_level": "R1",
        "status": "review",
        "notes": "Backend가 선택값을 검증한 뒤 Agent를 resume",
        "external_action": "component · account_card_list",
        "input_state_keys": "accounts",
        "output_state_keys": "account_selection_outcome, account_ids",
        "route_summary": "selected → check_transaction_period | cancelled → END",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_transaction_history",
        "step_order": 4,
        "step_id": "emit_transaction_accounts_empty",
        "step_name": "거래내역 조회 계좌 없음 안내",
        "step_purpose": "조회 가능한 계좌가 없는 빈 상태를 동일한 계좌 선택 UI로 전송한다.",
        "interaction_mode": "webhook",
        "tool_id": "emit_component",
        "contract_id": "UI-TRANSACTION-ACCOUNT-SELECTION",
        "step_risk_level": "R0",
        "status": "review",
        "notes": "accounts=[]을 전송하고 사용자 회신을 기다리지 않음",
        "external_action": "component · account_card_list",
        "input_state_keys": "accounts",
        "output_state_keys": "",
        "route_summary": "항상 → END",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_transaction_history",
        "step_order": 5,
        "step_id": "check_transaction_period",
        "step_name": "거래내역 조회 기간 정규화",
        "step_purpose": "기간 미입력 시 최근 1개월을 적용하고 해석 실패 시 기간 선택 UI로 연결한다.",
        "interaction_mode": "agent_internal",
        "tool_id": "normalize_transaction_period",
        "contract_id": "",
        "step_risk_level": "R1",
        "status": "review",
        "notes": "최대 기간과 날짜 순서는 Backend가 최종 검증",
        "external_action": "",
        "input_state_keys": "start_date, end_date, requested_at, timezone",
        "output_state_keys": "start_date, end_date",
        "route_summary": "정규화 또는 기본값 적용 → query_transactions | 해석 실패 → request_period_selection",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_transaction_history",
        "step_order": 6,
        "step_id": "request_period_selection",
        "step_name": "거래내역 조회 기간 선택 요청",
        "step_purpose": "기간 프리셋 또는 직접 입력 UI를 표시하고 Backend가 정규화한 결과를 기다린다.",
        "interaction_mode": "webhook_then_resume",
        "tool_id": "request_component_input",
        "contract_id": "UI-PERIOD-SELECTION",
        "step_risk_level": "R1",
        "status": "review",
        "notes": "Agent에는 preset 원문이 아니라 정규화된 날짜만 resume",
        "external_action": "component · period_input",
        "input_state_keys": "",
        "output_state_keys": "period_selection_outcome, start_date, end_date",
        "route_summary": "selected → query_transactions | cancelled → END",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_transaction_history",
        "step_order": 7,
        "step_id": "query_transactions",
        "step_name": "거래내역 첫 페이지 조회",
        "step_purpose": "Backend Tool API로 최근 1개월 범위의 최신순 10건 또는 사용자가 지정한 기간의 첫 페이지를 조회한다.",
        "interaction_mode": "backend_tool_api",
        "tool_id": "query_transactions",
        "contract_id": "API-TRANSACTION-QUERY",
        "step_risk_level": "R1",
        "status": "review",
        "notes": "limit=10은 Tool 설정값이며 cursor는 보내지 않음",
        "external_action": "POST /api/v1/agent-tools/transactions:query",
        "input_state_keys": "account_ids, start_date, end_date, keyword, transaction_type",
        "output_state_keys": "transaction_results, transaction_query_id, next_cursor",
        "route_summary": "성공 → emit_transaction_result | 오류 → emit_transaction_error",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_transaction_history",
        "step_order": 8,
        "step_id": "emit_transaction_result",
        "step_name": "거래내역 결과 전송",
        "step_purpose": "첫 페이지 거래내역과 이후 페이지 조회 Context를 결과 Webhook으로 전송한다.",
        "interaction_mode": "webhook",
        "tool_id": "emit_component",
        "contract_id": "UI-TRANSACTION-LIST",
        "step_risk_level": "R0",
        "status": "review",
        "notes": "transaction_results=[]도 동일한 UI의 정상 빈 상태",
        "external_action": "component · transaction_list",
        "input_state_keys": "account_ids, start_date, end_date, keyword, transaction_results, transaction_query_id, next_cursor",
        "output_state_keys": "",
        "route_summary": "항상 → END",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_transaction_history",
        "step_order": 9,
        "step_id": "emit_transaction_error",
        "step_name": "거래내역 조회 오류 안내",
        "step_purpose": "거래내역 조회를 완료하지 못한 경우 Backend의 사용자 공개 가능 오류를 안내한다.",
        "interaction_mode": "webhook",
        "tool_id": "emit_error",
        "contract_id": "UI-COMMON-ERROR",
        "step_risk_level": "R0",
        "status": "review",
        "notes": "Agent가 오류 유형별 복구 흐름을 임의로 선택하지 않음",
        "external_action": "error · error_message",
        "input_state_keys": "",
        "output_state_keys": "",
        "route_summary": "항상 → END",
        "validation_result": "OK",
    },
]

TRANSACTION_ROUTE_ROWS = [
    ("extract_transaction_slots", "조건 추출 완료", "거래내역 조회 조건 추출을 마친 경우", "resolve_transaction_accounts", "힌트가 없어도 계속함"),
    ("resolve_transaction_accounts", "계좌 자동 확정", "Backend가 account_resolution_outcome=resolved와 검증된 account_ids를 반환한 경우", "check_transaction_period", "Agent가 후보 개수로 다시 판단하지 않음"),
    ("resolve_transaction_accounts", "계좌 선택 필요", "Backend가 account_resolution_outcome=selection_required와 검증된 계좌 후보를 반환한 경우", "request_transaction_account_selection", "단일·복수 선택 지원"),
    ("resolve_transaction_accounts", "조회 가능 계좌 없음", "Backend가 account_resolution_outcome=no_accounts와 accounts=[]을 반환한 경우", "emit_transaction_accounts_empty", "오류가 아닌 정상 빈 상태"),
    ("resolve_transaction_accounts", "계좌 확인 오류", "Backend가 계약된 계좌 확인 결과를 반환하지 못한 경우", "emit_transaction_error", "사용자 공개 가능 오류만 전달"),
    ("request_transaction_account_selection", "계좌 선택 완료", "Backend가 account_selection_outcome=selected와 검증된 account_ids로 Agent를 resume한 경우", "check_transaction_period", "선택값을 다시 검증하지 않음"),
    ("request_transaction_account_selection", "계좌 선택 취소", "Backend가 account_selection_outcome=cancelled로 Agent를 resume한 경우", "END", "추가 호출 없이 종료"),
    ("emit_transaction_accounts_empty", "빈 상태 전송 완료", "조회 가능 계좌 없음 Webhook 전송을 완료한 경우", "END", ""),
    ("check_transaction_period", "기간 정규화 완료", "명시 기간을 정규화했거나 기간 미입력에 최근 1개월 기본값을 적용한 경우", "query_transactions", "requested_at과 timezone 기준"),
    ("check_transaction_period", "기간 해석 실패", "사용자가 기간을 언급했지만 Agent가 안전하게 날짜로 정규화하지 못한 경우", "request_period_selection", "기간 미입력은 UI 요청 대상이 아님"),
    ("request_period_selection", "기간 선택 완료", "Backend가 period_selection_outcome=selected와 정규화된 날짜로 Agent를 resume한 경우", "query_transactions", "Agent가 preset을 다시 해석하지 않음"),
    ("request_period_selection", "기간 선택 취소", "Backend가 period_selection_outcome=cancelled로 Agent를 resume한 경우", "END", "추가 호출 없이 종료"),
    ("query_transactions", "거래내역 조회 성공", "Backend가 첫 페이지 결과를 정상 반환한 경우. transaction_results가 빈 배열인 경우도 포함한다.", "emit_transaction_result", "최신순 10건"),
    ("query_transactions", "거래내역 조회 오류", "Backend가 요청 검증, 접근 권한 또는 기술 오류를 반환한 경우", "emit_transaction_error", "오류 유형별 자동 복구 없음"),
    ("emit_transaction_result", "결과 전송 완료", "거래내역 결과 Webhook 전송을 완료한 경우", "END", "이후 페이지는 Frontend와 Backend가 처리"),
    ("emit_transaction_error", "오류 안내 완료", "거래내역 조회 오류 안내 Webhook 전송을 완료한 경우", "END", ""),
]

TRANSACTION_ROUTE_ROWS = [
    {
        "workflow_id": "wf_transaction_history",
        "from_step_id": from_step_id,
        "route_name": route_name,
        "condition_description": condition_description,
        "to_step_id": to_step_id,
        "status": "review",
        "notes": notes,
    }
    for from_step_id, route_name, condition_description, to_step_id, notes in TRANSACTION_ROUTE_ROWS
]

SUMMARY_STEP_DEFINITIONS = [
    (1, "extract_amount_summary_slots", "합계 조회 조건 추출", "사용자 발화에서 계좌, 기간, 합계 유형과 검색어 힌트를 추출한다.", "agent_internal", "extract_amount_summary_slots", "", "R2", "계좌 확정과 원장 거래 분류를 수행하지 않음", "", "", "account_hint, all_accounts_requested, start_date, end_date, summary_type, keyword", "항상 → resolve_summary_accounts"),
    (2, "resolve_summary_accounts", "합계 조회 계좌 확인", "Backend가 집계 가능한 계좌를 검증하고 전체 또는 단일 계좌 자동 확정 여부를 결정한다.", "backend_tool_api", "fetch_accounts", "API-ACCOUNT-LIST", "R2", "resolve_selection=true, account_capability=inquiry는 Tool 설정값", "GET /api/v1/agent-tools/accounts", "account_hint, all_accounts_requested", "account_resolution_outcome, accounts, account_ids", "resolved → check_summary_period | selection_required → request_summary_account_selection | no_accounts → emit_summary_accounts_empty | 오류 → emit_amount_summary_error"),
    (3, "request_summary_account_selection", "합계 조회 계좌 선택 요청", "검증된 복수 계좌 후보를 표시하고 사용자 선택 또는 취소 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-SUMMARY-ACCOUNT-SELECTION", "R1", "Backend가 선택값을 검증한 뒤 Agent를 resume", "component · account_card_list", "accounts", "account_selection_outcome, account_ids", "selected → check_summary_period | cancelled → END"),
    (4, "emit_summary_accounts_empty", "합계 조회 계좌 없음 안내", "집계 가능한 계좌가 없는 빈 상태를 동일한 계좌 선택 UI로 전송한다.", "webhook", "emit_component", "UI-SUMMARY-ACCOUNT-SELECTION", "R0", "accounts=[]을 전송하고 사용자 회신을 기다리지 않음", "component · account_card_list", "accounts", "", "항상 → END"),
    (5, "check_summary_period", "합계 조회 기간 정규화", "기간 미입력 시 최근 1개월을 적용하고 해석 실패 시 기간 선택 UI로 연결한다.", "agent_internal", "normalize_summary_period", "", "R1", "최대 기간과 날짜 순서는 Backend가 최종 검증", "", "start_date, end_date, requested_at, timezone", "start_date, end_date", "정규화 또는 기본값 적용 → check_summary_type | 해석 실패 → request_period_selection"),
    (6, "request_period_selection", "합계 조회 기간 선택 요청", "기간 프리셋 또는 직접 입력 UI를 표시하고 Backend가 정규화한 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-PERIOD-SELECTION", "R1", "Agent에는 preset 원문이 아니라 정규화된 날짜만 resume", "component · period_input", "", "period_selection_outcome, start_date, end_date", "selected → check_summary_type | cancelled → END"),
    (7, "check_summary_type", "합계 유형 확인", "사용자 발화에서 지출 또는 수입 합계 목적이 확정되었는지 확인한다.", "agent_internal", "check_summary_type", "", "R1", "원장 거래를 지출·수입으로 분류하지 않음", "", "summary_type", "summary_type", "확정 → query_transaction_summary | 불명확 → request_summary_type"),
    (8, "request_summary_type", "합계 유형 선택 요청", "지출 또는 수입 중 하나를 선택받고 Backend가 검증한 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-SUMMARY-TYPE-SELECTION", "R1", "Backend가 summary_type Enum을 검증한 뒤 Agent를 resume", "component · option_select", "", "summary_type_selection_outcome, summary_type", "selected → query_transaction_summary | cancelled → END"),
    (9, "query_transaction_summary", "기간 거래 합계 조회", "Backend Tool API로 검증된 계좌와 기간의 지출 또는 수입 합계를 조회한다.", "backend_tool_api", "query_transaction_summary", "API-TRANSACTION-SUMMARY", "R2", "Agent가 거래내역을 받거나 직접 합산하지 않음", "POST /api/v1/agent-tools/transactions:summary", "account_ids, start_date, end_date, summary_type, keyword", "summary_result", "성공 → emit_amount_summary | 오류 → emit_amount_summary_error"),
    (10, "emit_amount_summary", "기간 거래 합계 결과 전송", "Backend가 반환한 합계 결과를 결과 Webhook으로 전송한다.", "webhook", "emit_component", "UI-AMOUNT-SUMMARY", "R0", "total_amount=0도 동일한 UI의 정상 결과", "component · amount_summary", "account_ids, keyword, summary_result", "", "항상 → END"),
    (11, "emit_amount_summary_error", "기간 거래 합계 오류 안내", "합계 조회를 완료하지 못한 경우 Backend의 사용자 공개 가능 오류를 안내한다.", "webhook", "emit_error", "UI-COMMON-ERROR", "R0", "Agent가 오류 유형별 복구 흐름을 임의로 선택하지 않음", "error · error_message", "", "", "항상 → END"),
]

SUMMARY_STEP_ROWS = [
    {
        "workflow_id": "wf_period_amount_summary",
        "step_order": step_order,
        "step_id": step_id,
        "step_name": step_name,
        "step_purpose": step_purpose,
        "interaction_mode": interaction_mode,
        "tool_id": tool_id,
        "contract_id": contract_id,
        "step_risk_level": step_risk_level,
        "status": "review",
        "notes": notes,
        "external_action": external_action,
        "input_state_keys": input_state_keys,
        "output_state_keys": output_state_keys,
        "route_summary": route_summary,
        "validation_result": "OK",
    }
    for (
        step_order,
        step_id,
        step_name,
        step_purpose,
        interaction_mode,
        tool_id,
        contract_id,
        step_risk_level,
        notes,
        external_action,
        input_state_keys,
        output_state_keys,
        route_summary,
    ) in SUMMARY_STEP_DEFINITIONS
]

SUMMARY_ROUTE_DEFINITIONS = [
    ("extract_amount_summary_slots", "조건 추출 완료", "합계 조회 조건 추출을 마친 경우", "resolve_summary_accounts", "계좌 힌트가 없으면 all_accounts_requested=true"),
    ("resolve_summary_accounts", "계좌 자동 확정", "Backend가 account_resolution_outcome=resolved와 검증된 account_ids를 반환한 경우", "check_summary_period", "전체 계좌 또는 단일 계좌"),
    ("resolve_summary_accounts", "계좌 선택 필요", "Backend가 account_resolution_outcome=selection_required와 검증된 계좌 후보를 반환한 경우", "request_summary_account_selection", "단일·복수 선택 지원"),
    ("resolve_summary_accounts", "집계 가능 계좌 없음", "Backend가 account_resolution_outcome=no_accounts와 accounts=[]을 반환한 경우", "emit_summary_accounts_empty", "오류가 아닌 정상 빈 상태"),
    ("resolve_summary_accounts", "계좌 확인 오류", "Backend가 계약된 계좌 확인 결과를 반환하지 못한 경우", "emit_amount_summary_error", "사용자 공개 가능 오류만 전달"),
    ("request_summary_account_selection", "계좌 선택 완료", "Backend가 account_selection_outcome=selected와 검증된 account_ids로 Agent를 resume한 경우", "check_summary_period", "선택값을 다시 검증하지 않음"),
    ("request_summary_account_selection", "계좌 선택 취소", "Backend가 account_selection_outcome=cancelled로 Agent를 resume한 경우", "END", "추가 호출 없이 종료"),
    ("emit_summary_accounts_empty", "빈 상태 전송 완료", "집계 가능 계좌 없음 Webhook 전송을 완료한 경우", "END", ""),
    ("check_summary_period", "기간 정규화 완료", "명시 기간을 정규화했거나 기간 미입력에 최근 1개월 기본값을 적용한 경우", "check_summary_type", "requested_at과 timezone 기준"),
    ("check_summary_period", "기간 해석 실패", "사용자가 기간을 언급했지만 Agent가 안전하게 날짜로 정규화하지 못한 경우", "request_period_selection", "기간 미입력은 UI 요청 대상이 아님"),
    ("request_period_selection", "기간 선택 완료", "Backend가 period_selection_outcome=selected와 정규화된 날짜로 Agent를 resume한 경우", "check_summary_type", "Agent가 preset을 다시 해석하지 않음"),
    ("request_period_selection", "기간 선택 취소", "Backend가 period_selection_outcome=cancelled로 Agent를 resume한 경우", "END", "추가 호출 없이 종료"),
    ("check_summary_type", "합계 유형 확정", "사용자 발화에서 summary_type이 spending 또는 income으로 확정된 경우", "query_transaction_summary", ""),
    ("check_summary_type", "합계 유형 선택 필요", "사용자 발화만으로 summary_type을 안전하게 확정하지 못한 경우", "request_summary_type", "기본값을 추측하지 않음"),
    ("request_summary_type", "합계 유형 선택 완료", "Backend가 summary_type_selection_outcome=selected와 검증된 summary_type으로 Agent를 resume한 경우", "query_transaction_summary", ""),
    ("request_summary_type", "합계 유형 선택 취소", "Backend가 summary_type_selection_outcome=cancelled로 Agent를 resume한 경우", "END", "추가 호출 없이 종료"),
    ("query_transaction_summary", "합계 조회 성공", "Backend가 summary_result를 정상 반환한 경우. total_amount=0인 경우도 포함한다.", "emit_amount_summary", "원장 거래를 Agent가 직접 합산하지 않음"),
    ("query_transaction_summary", "합계 조회 오류", "Backend가 요청 검증, 접근 권한 또는 기술 오류를 반환한 경우", "emit_amount_summary_error", "오류 유형별 자동 복구 없음"),
    ("emit_amount_summary", "결과 전송 완료", "기간 거래 합계 결과 Webhook 전송을 완료한 경우", "END", ""),
    ("emit_amount_summary_error", "오류 안내 완료", "기간 거래 합계 오류 안내 Webhook 전송을 완료한 경우", "END", ""),
]

SUMMARY_ROUTE_ROWS = [
    {
        "workflow_id": "wf_period_amount_summary",
        "from_step_id": from_step_id,
        "route_name": route_name,
        "condition_description": condition_description,
        "to_step_id": to_step_id,
        "status": "review",
        "notes": notes,
    }
    for from_step_id, route_name, condition_description, to_step_id, notes in SUMMARY_ROUTE_DEFINITIONS
]

DEFAULT_ACCOUNT_STEP_DEFINITIONS = [
    (1, "extract_default_account_slots", "기본계좌 대상 힌트 추출", "사용자 발화에서 새 기본 출금 계좌의 선택적 힌트를 추출한다.", "agent_internal", "extract_default_account_slots", "", "R2", "계좌를 직접 확정하거나 검증하지 않음", "", "", "account_hint", "항상 → resolve_default_account"),
    (2, "resolve_default_account", "기본계좌 대상 확인", "Backend가 설정 가능한 계좌를 검증하고 자동 확정 여부를 결정한다.", "backend_tool_api", "fetch_accounts", "API-ACCOUNT-LIST", "R2", "resolve_selection=true, account_capability=settings는 Tool 설정값", "GET /api/v1/agent-tools/accounts", "account_hint", "account_resolution_outcome, accounts, account_id", "resolved → start_default_account_prepare | selection_required → request_default_account_selection | no_accounts → emit_default_account_selection_empty | 오류 → emit_default_account_error"),
    (3, "request_default_account_selection", "기본 출금 계좌 선택 요청", "검증된 계좌 후보를 표시하고 단일 선택 또는 취소 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-DEFAULT-ACCOUNT-SELECTION", "R2", "Backend가 검증한 account_ids 배열의 단일 값을 account_id로 저장", "component · account_card_list", "accounts", "account_selection_outcome, account_id", "selected → start_default_account_prepare | cancelled → END | 계약 오류 → emit_default_account_error"),
    (4, "emit_default_account_selection_empty", "설정 가능 계좌 없음 안내", "설정 가능한 계좌가 없는 빈 상태를 동일한 계좌 선택 UI로 전송한다.", "webhook", "emit_component", "UI-DEFAULT-ACCOUNT-SELECTION", "R0", "accounts=[]을 전송하고 사용자 회신을 기다리지 않음", "component · account_card_list", "accounts", "", "항상 → END"),
    (5, "start_default_account_prepare", "기본계좌 Prepare 시도 시작", "이전 수정 사유를 제거하고 새로운 논리 Prepare 요청의 시도 번호를 증가시켜 Checkpoint에 저장한다.", "agent_internal", "start_prepare_attempt", "", "R2", "통신 재시도에서는 prepare_attempt를 증가시키지 않음", "", "prepare_attempt, correction_view", "prepare_attempt, correction_view", "항상 → prepare_default_account_change"),
    (6, "prepare_default_account_change", "기본계좌 변경 준비", "Backend가 계좌 소유권·활성·출금 가능 여부와 현재 설정을 검증하고 Confirmation을 준비한다.", "backend_tool_api", "prepare_default_account_change", "API-DEFAULT-ACCOUNT-PREPARE", "R3", "Timeout·502·503·504에 한해 같은 멱등성 키로 최대 1회 재시도", "POST /api/v1/agent-tools/settings/default-account:prepare", "account_id, prepare_attempt", "confirmation_id, confirmation_view, correction_view", "ready_for_confirmation → request_default_account_approval | unchanged → emit_default_account_unchanged | correction_required → reset_default_account_target | blocked → emit_default_account_blocked | 오류 → emit_default_account_error"),
    (7, "emit_default_account_unchanged", "기본계좌 변경 없음 결과", "대상 계좌가 이미 기본계좌인 정상 결과를 setting_result Webhook으로 전송한다.", "webhook", "emit_component", "UI-DEFAULT-ACCOUNT-RESULT", "R0", "outcome=unchanged이며 Confirmation을 생성하지 않음", "component · setting_result", "account_id", "", "항상 → END"),
    (8, "emit_default_account_blocked", "기본계좌 변경 차단 안내", "사용자 수정으로 해결할 수 없는 설정 변경 차단 사유를 안내한다.", "webhook", "emit_blocked", "UI-SETTING-BLOCKED", "R0", "내부 정책 세부정보는 노출하지 않음", "blocked · blocked_message", "", "", "항상 → END"),
    (9, "request_default_account_approval", "기본계좌 변경 승인 요청", "Backend가 생성한 변경 전후 표시 데이터를 보여주고 승인·수정·취소 결과를 기다린다.", "webhook_then_resume", "request_confirmation", "UI-DEFAULT-ACCOUNT-CONFIRMATION", "R3", "Backend가 Confirmation을 검증하고 approval_outcome으로 resume", "component · confirm_modal", "confirmation_id, confirmation_view", "approval_outcome", "approved → execute_default_account_change | change_requested → reset_default_account_target | cancelled → END | 계약 오류 → emit_default_account_error"),
    (10, "reset_default_account_target", "기본계좌 대상 초기화", "수정 또는 계좌 상태 변경 시 기존 대상·후보·승인 임시 State를 제거한다.", "agent_internal", "reset_default_account_target", "", "R2", "prepare_attempt와 correction_view는 유지하고 최신 계좌 확인을 다시 수행", "", "account_hint, account_resolution_outcome, accounts, account_id, account_selection_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome", "account_hint, account_resolution_outcome, accounts, account_id, account_selection_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome", "항상 → resolve_default_account"),
    (11, "execute_default_account_change", "기본계좌 변경 실행", "Backend가 승인된 Confirmation을 검증하고 기본 출금 계좌 변경을 실행한다.", "backend_tool_api", "execute_default_account_change", "API-DEFAULT-ACCOUNT-EXECUTE", "R3", "Timeout·502·503·504에 한해 같은 멱등성 키로 최대 1회 재시도", "POST /api/v1/agent-tools/settings/default-account", "confirmation_id", "account_id, completed_at, correction_view", "completed → emit_default_account_result | correction_required → reset_default_account_target | blocked → emit_default_account_blocked | 오류 → emit_default_account_error"),
    (12, "emit_default_account_result", "기본계좌 변경 결과 전송", "Backend가 실제 반영한 기본 출금 계좌 변경 결과를 Webhook으로 전송한다.", "webhook", "emit_component", "UI-DEFAULT-ACCOUNT-RESULT", "R0", "outcome=completed", "component · setting_result", "account_id, completed_at", "", "항상 → END"),
    (13, "emit_default_account_error", "기본계좌 변경 오류 안내", "기본계좌 변경을 완료하지 못한 경우 사용자에게 공개 가능한 오류를 안내한다.", "webhook", "emit_error", "UI-COMMON-ERROR", "R0", "재시도 대상은 공통 Adapter가 최대 1회 처리한 뒤 진입", "error · error_message", "", "", "항상 → END"),
]

DEFAULT_ACCOUNT_STEP_ROWS = [
    {
        "workflow_id": "wf_set_default_account",
        "step_order": step_order,
        "step_id": step_id,
        "step_name": step_name,
        "step_purpose": step_purpose,
        "interaction_mode": interaction_mode,
        "tool_id": tool_id,
        "contract_id": contract_id,
        "step_risk_level": step_risk_level,
        "status": "review",
        "notes": notes,
        "external_action": external_action,
        "input_state_keys": input_state_keys,
        "output_state_keys": output_state_keys,
        "route_summary": route_summary,
        "validation_result": "OK",
    }
    for (
        step_order,
        step_id,
        step_name,
        step_purpose,
        interaction_mode,
        tool_id,
        contract_id,
        step_risk_level,
        notes,
        external_action,
        input_state_keys,
        output_state_keys,
        route_summary,
    ) in DEFAULT_ACCOUNT_STEP_DEFINITIONS
]

DEFAULT_ACCOUNT_ROUTE_DEFINITIONS = [
    ("extract_default_account_slots", "대상 힌트 추출 완료", "사용자 발화에서 선택적 계좌 힌트 추출을 마친 경우", "resolve_default_account", "힌트가 없어도 최신 계좌 확인을 계속함"),
    ("resolve_default_account", "계좌 자동 확정", "Backend가 account_resolution_outcome=resolved와 정확히 하나의 account_ids를 반환한 경우", "start_default_account_prepare", "Agent가 account_ids[0]을 account_id로 저장"),
    ("resolve_default_account", "사용자 선택 필요", "Backend가 account_resolution_outcome=selection_required와 검증된 최신 계좌 후보를 반환한 경우", "request_default_account_selection", "단일 선택 UI"),
    ("resolve_default_account", "설정 가능 계좌 없음", "Backend가 account_resolution_outcome=no_accounts와 accounts=[]을 반환한 경우", "emit_default_account_selection_empty", "오류가 아닌 정상 빈 상태"),
    ("resolve_default_account", "계좌 확인 오류", "Backend 오류이거나 resolved의 account_ids가 정확히 하나가 아닌 계약 오류인 경우", "emit_default_account_error", "공통 재시도 이후 오류 Route"),
    ("request_default_account_selection", "계좌 선택 완료", "Backend가 account_selection_outcome=selected와 정확히 하나의 account_ids로 Agent를 resume한 경우", "start_default_account_prepare", "account_ids[0]을 account_id로 저장"),
    ("request_default_account_selection", "계좌 선택 취소", "Backend가 account_selection_outcome=cancelled로 Agent를 resume한 경우", "END", "추가 Tool API와 UI Webhook 없이 종료"),
    ("request_default_account_selection", "계좌 선택 계약 오류", "selected인데 account_ids가 정확히 하나가 아닌 경우", "emit_default_account_error", "임의로 첫 번째 값을 선택하지 않음"),
    ("emit_default_account_selection_empty", "빈 상태 전송 완료", "설정 가능 계좌 없음 Webhook 전송을 완료한 경우", "END", ""),
    ("start_default_account_prepare", "Prepare 시도 저장 완료", "prepare_attempt 증가와 Checkpoint 저장을 완료한 경우", "prepare_default_account_change", "새로운 논리 요청에만 증가"),
    ("prepare_default_account_change", "승인 준비 완료", "Backend가 outcome=ready_for_confirmation과 Confirmation을 반환한 경우", "request_default_account_approval", ""),
    ("prepare_default_account_change", "이미 기본계좌", "Backend가 outcome=unchanged를 반환한 경우", "emit_default_account_unchanged", "Confirmation을 생성하지 않음"),
    ("prepare_default_account_change", "다른 계좌 필요", "Backend가 outcome=correction_required와 allowed_change_targets에 account를 반환한 경우", "reset_default_account_target", "최신 계좌를 다시 조회"),
    ("prepare_default_account_change", "설정 변경 차단", "Backend가 outcome=blocked를 반환한 경우", "emit_default_account_blocked", "사용자 수정으로 해결할 수 없음"),
    ("prepare_default_account_change", "Prepare 오류", "재시도 대상이 아니거나 공통 최대 1회 재시도 후에도 실패한 경우", "emit_default_account_error", ""),
    ("emit_default_account_unchanged", "변경 없음 결과 전송 완료", "변경 없음 결과 Webhook 전송을 완료한 경우", "END", ""),
    ("emit_default_account_blocked", "차단 안내 완료", "설정 변경 차단 Webhook 전송을 완료한 경우", "END", ""),
    ("request_default_account_approval", "변경 승인 완료", "Backend가 approval_outcome=approved로 Agent를 resume한 경우", "execute_default_account_change", "confirmation_id만 유지"),
    ("request_default_account_approval", "계좌 수정 요청", "Backend가 기존 Confirmation을 무효화하고 approval_outcome=change_requested로 Agent를 resume한 경우", "reset_default_account_target", "최신 계좌를 다시 조회"),
    ("request_default_account_approval", "변경 취소", "Backend가 Confirmation을 취소하고 approval_outcome=cancelled로 Agent를 resume한 경우", "END", "추가 UI Webhook 없이 종료"),
    ("request_default_account_approval", "승인 결과 계약 오류", "approval_outcome이 계약된 Enum과 일치하지 않는 경우", "emit_default_account_error", "Agent가 승인 결과를 추측하지 않음"),
    ("reset_default_account_target", "대상 초기화 완료", "기존 대상·후보·승인 임시 State 제거를 완료한 경우", "resolve_default_account", "prepare_attempt는 유지"),
    ("execute_default_account_change", "변경 실행 완료", "Backend가 outcome=completed와 최종 account_id를 반환한 경우", "emit_default_account_result", ""),
    ("execute_default_account_change", "실행 시점 계좌 수정 필요", "Backend가 outcome=correction_required와 allowed_change_targets에 account를 반환한 경우", "reset_default_account_target", "기존 Confirmation은 재사용하지 않음"),
    ("execute_default_account_change", "실행 차단", "Backend가 outcome=blocked를 반환한 경우", "emit_default_account_blocked", ""),
    ("execute_default_account_change", "Execute 오류", "재시도 대상이 아니거나 공통 최대 1회 재시도 후에도 실패한 경우", "emit_default_account_error", "같은 멱등성 키와 Body 유지"),
    ("emit_default_account_result", "변경 결과 전송 완료", "기본계좌 변경 완료 Webhook 전송을 완료한 경우", "END", ""),
    ("emit_default_account_error", "오류 안내 완료", "기본계좌 변경 오류 안내 Webhook 전송을 완료한 경우", "END", ""),
]

DEFAULT_ACCOUNT_ROUTE_ROWS = [
    {
        "workflow_id": "wf_set_default_account",
        "from_step_id": from_step_id,
        "route_name": route_name,
        "condition_description": condition_description,
        "to_step_id": to_step_id,
        "status": "review",
        "notes": notes,
    }
    for from_step_id, route_name, condition_description, to_step_id, notes in DEFAULT_ACCOUNT_ROUTE_DEFINITIONS
]

ACCOUNT_ALIAS_STEP_DEFINITIONS = [
    (1, "extract_account_alias_slots", "별칭 변경 조건 추출", "사용자 발화에서 대상 계좌 힌트와 새 별칭을 추출한다.", "agent_internal", "extract_account_alias_slots", "", "R2", "계좌 확정과 별칭 정책 검증은 수행하지 않음", "", "", "account_hint, alias", "항상 → resolve_account_alias_target"),
    (2, "resolve_account_alias_target", "별칭 변경 계좌 확인", "Backend가 설정 가능한 계좌를 검증하고 자동 확정 여부를 결정한다.", "backend_tool_api", "fetch_accounts", "API-ACCOUNT-LIST", "R2", "resolve_selection=true, account_capability=settings는 Tool 설정값", "GET /api/v1/agent-tools/accounts", "account_hint", "account_resolution_outcome, accounts, account_id", "resolved → check_account_alias_value | selection_required → request_account_alias_selection | no_accounts → emit_account_alias_selection_empty | 오류 → emit_account_alias_error"),
    (3, "request_account_alias_selection", "별칭 변경 계좌 선택 요청", "검증된 최신 계좌 후보를 표시하고 단일 선택 또는 취소 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-ACCOUNT-ALIAS-SELECTION", "R2", "correction_view가 있으면 수정 사유도 표시", "component · account_card_list", "accounts, correction_view", "account_selection_outcome, account_id", "selected → check_account_alias_value | cancelled → END | 계약 오류 → emit_account_alias_error"),
    (4, "emit_account_alias_selection_empty", "별칭 변경 가능 계좌 없음 안내", "별칭을 변경할 수 있는 계좌가 없는 빈 상태를 동일한 선택 UI로 전송한다.", "webhook", "emit_component", "UI-ACCOUNT-ALIAS-SELECTION", "R0", "accounts=[]을 전송하고 사용자 회신을 기다리지 않음", "component · account_card_list", "accounts", "", "항상 → END"),
    (5, "check_account_alias_value", "새 별칭 존재 여부 확인", "State에 새 별칭이 존재하는지 구조적으로 확인한다.", "agent_internal", "check_account_alias_value", "", "R2", "별칭 정책 검증은 Backend가 수행", "", "alias", "", "alias 존재 → start_account_alias_prepare | alias 없음 → request_account_alias_input"),
    (6, "request_account_alias_input", "새 계좌 별칭 입력 요청", "새 별칭 입력 UI를 표시하고 Backend가 검증·정규화한 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-ACCOUNT-ALIAS-INPUT", "R2", "검증 실패는 Agent를 resume하지 않고 같은 UI에서 처리", "component · text_input", "correction_view", "alias_input_outcome, alias", "submitted → start_account_alias_prepare | cancelled → END | 계약 오류 → emit_account_alias_error"),
    (7, "start_account_alias_prepare", "별칭 변경 Prepare 시도 시작", "이전 수정 사유를 제거하고 새로운 논리 Prepare의 시도 번호를 증가시켜 Checkpoint에 저장한다.", "agent_internal", "start_prepare_attempt", "", "R2", "통신 재시도에서는 prepare_attempt를 증가시키지 않음", "", "prepare_attempt, correction_view", "prepare_attempt, correction_view", "항상 → prepare_account_alias_change"),
    (8, "prepare_account_alias_change", "계좌 별칭 변경 준비", "Backend가 계좌 소유권·상태와 별칭 정책을 검증하고 Confirmation을 준비한다.", "backend_tool_api", "prepare_account_alias_change", "API-ACCOUNT-ALIAS-PREPARE", "R3", "Timeout·502·503·504에 한해 같은 멱등성 키로 최대 1회 재시도", "POST /api/v1/agent-tools/settings/account-alias:prepare", "account_id, alias, prepare_attempt", "confirmation_id, confirmation_view, correction_view", "ready_for_confirmation → request_account_alias_approval | unchanged → emit_account_alias_unchanged | correction/account → reset_account_alias_target | correction/alias → reset_account_alias_value | blocked → emit_account_alias_blocked | 오류 → emit_account_alias_error"),
    (9, "emit_account_alias_unchanged", "계좌 별칭 변경 없음 결과", "이미 같은 별칭인 정상 결과를 setting_result Webhook으로 전송한다.", "webhook", "emit_component", "UI-ACCOUNT-ALIAS-RESULT", "R0", "outcome=unchanged이며 Confirmation을 생성하지 않음", "component · setting_result", "account_id, alias", "", "항상 → END"),
    (10, "emit_account_alias_blocked", "계좌 별칭 변경 차단 안내", "사용자 수정으로 해결할 수 없는 설정 변경 차단 사유를 안내한다.", "webhook", "emit_blocked", "UI-SETTING-BLOCKED", "R0", "내부 정책 세부정보는 노출하지 않음", "blocked · blocked_message", "", "", "항상 → END"),
    (11, "request_account_alias_approval", "계좌 별칭 변경 승인 요청", "대상 계좌와 최종 별칭을 표시하고 승인·수정·취소 결과를 기다린다.", "webhook_then_resume", "request_confirmation", "UI-ACCOUNT-ALIAS-CONFIRMATION", "R3", "Backend가 Confirmation과 change_target을 검증한 뒤 resume", "component · confirm_modal", "confirmation_id, confirmation_view", "approval_outcome, change_target", "approved → execute_account_alias_change | change/account → reset_account_alias_target | change/alias → reset_account_alias_value | cancelled → END | 계약 오류 → emit_account_alias_error"),
    (12, "reset_account_alias_target", "별칭 변경 계좌 초기화", "계좌 수정 시 이전 대상·후보·승인 임시 State를 제거한다.", "agent_internal", "reset_account_alias_target", "", "R2", "alias, prepare_attempt, correction_view는 유지", "", "account_hint, account_resolution_outcome, accounts, account_id, account_selection_outcome, alias_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target", "account_hint, account_resolution_outcome, accounts, account_id, account_selection_outcome, alias_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target", "항상 → resolve_account_alias_target"),
    (13, "reset_account_alias_value", "새 별칭 값 초기화", "별칭 수정 시 이전 별칭·승인 임시 State를 제거한다.", "agent_internal", "reset_account_alias_value", "", "R2", "account_id, prepare_attempt, correction_view는 유지", "", "alias, alias_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target", "alias, alias_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target", "항상 → request_account_alias_input"),
    (14, "execute_account_alias_change", "계좌 별칭 변경 실행", "Backend가 승인된 Confirmation을 검증하고 계좌 별칭 변경을 실행한다.", "backend_tool_api", "execute_account_alias_change", "API-ACCOUNT-ALIAS-EXECUTE", "R3", "Timeout·502·503·504에 한해 같은 멱등성 키로 최대 1회 재시도", "POST /api/v1/agent-tools/settings/account-alias", "confirmation_id", "account_id, alias, completed_at, correction_view", "completed → emit_account_alias_result | correction/account → reset_account_alias_target | correction/alias → reset_account_alias_value | blocked → emit_account_alias_blocked | 오류 → emit_account_alias_error"),
    (15, "emit_account_alias_result", "계좌 별칭 변경 결과 전송", "Backend가 실제 반영한 계좌 별칭 변경 결과를 Webhook으로 전송한다.", "webhook", "emit_component", "UI-ACCOUNT-ALIAS-RESULT", "R0", "outcome=completed", "component · setting_result", "account_id, alias, completed_at", "", "항상 → END"),
    (16, "emit_account_alias_error", "계좌 별칭 변경 오류 안내", "계좌 별칭 변경을 완료하지 못한 경우 사용자에게 공개 가능한 오류를 안내한다.", "webhook", "emit_error", "UI-COMMON-ERROR", "R0", "공통 Adapter 재시도와 계약 검사를 마친 뒤 진입", "error · error_message", "", "", "항상 → END"),
]

ACCOUNT_ALIAS_STEP_ROWS = [
    {
        "workflow_id": "wf_set_account_alias",
        "step_order": step_order,
        "step_id": step_id,
        "step_name": step_name,
        "step_purpose": step_purpose,
        "interaction_mode": interaction_mode,
        "tool_id": tool_id,
        "contract_id": contract_id,
        "step_risk_level": step_risk_level,
        "status": "review",
        "notes": notes,
        "external_action": external_action,
        "input_state_keys": input_state_keys,
        "output_state_keys": output_state_keys,
        "route_summary": route_summary,
        "validation_result": "OK",
    }
    for (
        step_order,
        step_id,
        step_name,
        step_purpose,
        interaction_mode,
        tool_id,
        contract_id,
        step_risk_level,
        notes,
        external_action,
        input_state_keys,
        output_state_keys,
        route_summary,
    ) in ACCOUNT_ALIAS_STEP_DEFINITIONS
]

ACCOUNT_ALIAS_ROUTE_DEFINITIONS = [
    ("extract_account_alias_slots", "조건 추출 완료", "대상 계좌 힌트와 선택적 새 별칭 추출을 마친 경우", "resolve_account_alias_target", "alias가 없어도 Workflow에 진입"),
    ("resolve_account_alias_target", "계좌 자동 확정", "Backend가 resolved와 정확히 하나의 account_ids를 반환한 경우", "check_account_alias_value", "account_ids[0]을 account_id로 저장"),
    ("resolve_account_alias_target", "계좌 선택 필요", "Backend가 selection_required와 검증된 최신 계좌 후보를 반환한 경우", "request_account_alias_selection", "단일 선택 UI"),
    ("resolve_account_alias_target", "변경 가능 계좌 없음", "Backend가 no_accounts와 accounts=[]을 반환한 경우", "emit_account_alias_selection_empty", "정상 빈 상태"),
    ("resolve_account_alias_target", "계좌 확인 오류", "Backend 오류이거나 resolved의 account_ids가 정확히 하나가 아닌 경우", "emit_account_alias_error", "임의로 첫 값을 선택하지 않음"),
    ("request_account_alias_selection", "계좌 선택 완료", "Backend가 selected와 정확히 하나의 account_ids로 Agent를 resume한 경우", "check_account_alias_value", "account_ids[0]을 account_id로 저장"),
    ("request_account_alias_selection", "계좌 선택 취소", "Backend가 cancelled로 Agent를 resume한 경우", "END", "추가 호출 없이 종료"),
    ("request_account_alias_selection", "계좌 선택 계약 오류", "selected인데 account_ids가 정확히 하나가 아닌 경우", "emit_account_alias_error", ""),
    ("emit_account_alias_selection_empty", "빈 상태 전송 완료", "변경 가능 계좌 없음 Webhook 전송을 완료한 경우", "END", ""),
    ("check_account_alias_value", "별칭 존재", "정규화 대상 alias가 State에 존재하는 경우", "start_account_alias_prepare", "정책 검증은 Prepare에서 수행"),
    ("check_account_alias_value", "별칭 입력 필요", "alias가 null이거나 비어 있어 새 입력이 필요한 경우", "request_account_alias_input", "Agent가 기본 별칭을 추측하지 않음"),
    ("request_account_alias_input", "별칭 입력 완료", "Backend가 submitted와 정규화된 alias로 Agent를 resume한 경우", "start_account_alias_prepare", ""),
    ("request_account_alias_input", "별칭 입력 취소", "Backend가 cancelled와 alias=null로 Agent를 resume한 경우", "END", "추가 호출 없이 종료"),
    ("request_account_alias_input", "별칭 입력 계약 오류", "alias_input_outcome이나 alias가 계약과 일치하지 않는 경우", "emit_account_alias_error", ""),
    ("start_account_alias_prepare", "Prepare 시도 저장 완료", "correction_view 제거, prepare_attempt 증가와 Checkpoint 저장을 완료한 경우", "prepare_account_alias_change", "새로운 논리 요청에만 증가"),
    ("prepare_account_alias_change", "승인 준비 완료", "Backend가 ready_for_confirmation과 Confirmation을 반환한 경우", "request_account_alias_approval", ""),
    ("prepare_account_alias_change", "이미 같은 별칭", "Backend가 unchanged를 반환한 경우", "emit_account_alias_unchanged", "Confirmation을 생성하지 않음"),
    ("prepare_account_alias_change", "계좌 수정 필요", "correction_required이며 allowed_change_targets가 account 하나인 경우", "reset_account_alias_target", "최신 계좌를 다시 조회"),
    ("prepare_account_alias_change", "별칭 수정 필요", "correction_required이며 allowed_change_targets가 alias 하나인 경우", "reset_account_alias_value", "수정 사유를 입력 UI까지 유지"),
    ("prepare_account_alias_change", "수정 대상 계약 오류", "allowed_change_targets가 비었거나 두 개 이상인 경우", "emit_account_alias_error", "우선순위를 추측하지 않음"),
    ("prepare_account_alias_change", "설정 변경 차단", "Backend가 blocked를 반환한 경우", "emit_account_alias_blocked", ""),
    ("prepare_account_alias_change", "Prepare 오류", "재시도 대상이 아니거나 최대 1회 재시도 후에도 실패한 경우", "emit_account_alias_error", ""),
    ("emit_account_alias_unchanged", "변경 없음 결과 전송 완료", "변경 없음 Webhook 전송을 완료한 경우", "END", ""),
    ("emit_account_alias_blocked", "차단 안내 완료", "설정 변경 차단 Webhook 전송을 완료한 경우", "END", ""),
    ("request_account_alias_approval", "변경 승인 완료", "Backend가 approved로 Agent를 resume한 경우", "execute_account_alias_change", "confirmation_id만 유지"),
    ("request_account_alias_approval", "계좌 수정 요청", "Backend가 change_requested와 change_target=account로 Agent를 resume한 경우", "reset_account_alias_target", "기존 Confirmation은 무효화됨"),
    ("request_account_alias_approval", "별칭 수정 요청", "Backend가 change_requested와 change_target=alias로 Agent를 resume한 경우", "reset_account_alias_value", "기존 Confirmation은 무효화됨"),
    ("request_account_alias_approval", "변경 취소", "Backend가 cancelled로 Agent를 resume한 경우", "END", "추가 UI Webhook 없이 종료"),
    ("request_account_alias_approval", "승인 결과 계약 오류", "approval_outcome 또는 change_target이 계약과 일치하지 않는 경우", "emit_account_alias_error", ""),
    ("reset_account_alias_target", "계좌 초기화 완료", "계좌·승인 임시 State 제거를 완료한 경우", "resolve_account_alias_target", "alias와 correction_view 유지"),
    ("reset_account_alias_value", "별칭 초기화 완료", "별칭·승인 임시 State 제거를 완료한 경우", "request_account_alias_input", "account_id와 correction_view 유지"),
    ("execute_account_alias_change", "변경 실행 완료", "Backend가 completed와 최종 account_id, alias를 반환한 경우", "emit_account_alias_result", ""),
    ("execute_account_alias_change", "실행 시점 계좌 수정 필요", "correction_required이며 allowed_change_targets가 account 하나인 경우", "reset_account_alias_target", "기존 Confirmation은 재사용하지 않음"),
    ("execute_account_alias_change", "실행 시점 별칭 수정 필요", "correction_required이며 allowed_change_targets가 alias 하나인 경우", "reset_account_alias_value", "기존 Confirmation은 재사용하지 않음"),
    ("execute_account_alias_change", "실행 수정 대상 계약 오류", "allowed_change_targets가 비었거나 두 개 이상인 경우", "emit_account_alias_error", ""),
    ("execute_account_alias_change", "실행 차단", "Backend가 blocked를 반환한 경우", "emit_account_alias_blocked", ""),
    ("execute_account_alias_change", "Execute 오류", "재시도 대상이 아니거나 최대 1회 재시도 후에도 실패한 경우", "emit_account_alias_error", "같은 멱등성 키와 Body 유지"),
    ("emit_account_alias_result", "변경 결과 전송 완료", "계좌 별칭 변경 완료 Webhook 전송을 완료한 경우", "END", ""),
    ("emit_account_alias_error", "오류 안내 완료", "계좌 별칭 변경 오류 안내 Webhook 전송을 완료한 경우", "END", ""),
]

ACCOUNT_ALIAS_ROUTE_ROWS = [
    {
        "workflow_id": "wf_set_account_alias",
        "from_step_id": from_step_id,
        "route_name": route_name,
        "condition_description": condition_description,
        "to_step_id": to_step_id,
        "status": "review",
        "notes": notes,
    }
    for from_step_id, route_name, condition_description, to_step_id, notes in ACCOUNT_ALIAS_ROUTE_DEFINITIONS
]

INTERNAL_TRANSFER_STEP_DEFINITIONS = [
    (1, "extract_internal_transfer_slots", "본인송금 조건 추출", "사용자 발화에서 출금 계좌·입금 계좌 힌트와 금액을 추출한다.", "agent_internal", "extract_internal_transfer_slots", "", "R2", "계좌 확정과 금융 검증은 수행하지 않음", "", "", "from_account_hint, to_account_hint, amount", "항상 → resolve_internal_from_account"),
    (2, "resolve_internal_from_account", "출금 계좌 확인", "Backend가 출금 가능한 본인 계좌를 검증하고 자동 확정 여부를 결정한다.", "backend_tool_api", "fetch_accounts", "API-ACCOUNT-LIST", "R2", "to_account_id가 있으면 exclude_account_ids로 전달", "GET /api/v1/agent-tools/accounts", "from_account_hint, to_account_id", "account_resolution_outcome, accounts, from_account_id", "resolved → resolve_internal_to_account | selection_required → request_from_account_selection | no_accounts → emit_internal_from_accounts_empty | 오류 → emit_internal_transfer_error"),
    (3, "request_from_account_selection", "출금 계좌 선택 요청", "검증된 출금 계좌 후보를 표시하고 단일 선택 또는 취소 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-INTERNAL-TRANSFER-FROM-ACCOUNT", "R2", "correction_view가 있으면 수정 사유도 표시", "component · account_card_list", "accounts, correction_view", "account_selection_outcome, from_account_id", "selected → resolve_internal_to_account | cancelled → END | 계약 오류 → emit_internal_transfer_error"),
    (4, "emit_internal_from_accounts_empty", "출금 가능 계좌 없음 안내", "출금 가능한 본인 계좌가 없는 빈 상태를 동일한 계좌 선택 UI로 전송한다.", "webhook", "emit_component", "UI-INTERNAL-TRANSFER-FROM-ACCOUNT", "R0", "accounts=[]을 전송하고 사용자 회신을 기다리지 않음", "component · account_card_list", "accounts", "", "항상 → END"),
    (5, "resolve_internal_to_account", "입금 계좌 확인", "Backend가 출금 계좌를 제외한 입금 가능한 본인 계좌를 검증하고 자동 확정 여부를 결정한다.", "backend_tool_api", "fetch_accounts", "API-ACCOUNT-LIST", "R2", "from_account_id를 exclude_account_ids로 전달", "GET /api/v1/agent-tools/accounts", "to_account_hint, from_account_id", "account_resolution_outcome, accounts, to_account_id", "resolved → check_internal_transfer_amount | selection_required → request_to_account_selection | no_accounts → emit_internal_to_accounts_empty | 오류 → emit_internal_transfer_error"),
    (6, "request_to_account_selection", "입금 계좌 선택 요청", "검증된 입금 계좌 후보를 표시하고 단일 선택 또는 취소 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-INTERNAL-TRANSFER-TO-ACCOUNT", "R2", "correction_view가 있으면 수정 사유도 표시", "component · account_card_list", "accounts, correction_view", "account_selection_outcome, to_account_id", "selected → check_internal_transfer_amount | cancelled → END | 계약 오류 → emit_internal_transfer_error"),
    (7, "emit_internal_to_accounts_empty", "입금 가능 계좌 없음 안내", "입금 가능한 본인 계좌가 없는 빈 상태를 동일한 계좌 선택 UI로 전송한다.", "webhook", "emit_component", "UI-INTERNAL-TRANSFER-TO-ACCOUNT", "R0", "accounts=[]을 전송하고 사용자 회신을 기다리지 않음", "component · account_card_list", "accounts", "", "항상 → END"),
    (8, "check_internal_transfer_amount", "송금 금액 구조 확인", "금액이 존재하고 정수이며 0보다 큰지 구조적으로 확인한다.", "agent_internal", "check_internal_transfer_amount", "", "R2", "잔액·한도·정책은 Backend가 검증", "", "amount", "", "유효 → start_internal_transfer_prepare | 없거나 구조 오류 → request_internal_transfer_amount"),
    (9, "request_internal_transfer_amount", "송금 금액 입력 요청", "금액 입력 UI를 표시하고 Backend가 검증한 제출 또는 취소 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-TRANSFER-AMOUNT-INPUT", "R2", "검증 실패는 Agent를 재개하지 않고 같은 UI에서 처리", "component · number_input", "correction_view, currency", "amount_input_outcome, amount", "submitted → start_internal_transfer_prepare | cancelled → END | 계약 오류 → emit_internal_transfer_error"),
    (10, "start_internal_transfer_prepare", "본인송금 Prepare 시도 시작", "이전 수정 사유를 제거하고 새 Prepare 시도 번호를 증가시켜 Checkpoint에 저장한다.", "agent_internal", "start_prepare_attempt", "", "R2", "통신 재시도에서는 prepare_attempt를 증가시키지 않음", "", "prepare_attempt, correction_view", "prepare_attempt, correction_view", "항상 → prepare_internal_transfer"),
    (11, "prepare_internal_transfer", "본인송금 준비", "Backend가 계좌 소유권·상태, 잔액·한도·정책을 검증하고 Confirmation을 준비한다.", "backend_tool_api", "prepare_internal_transfer", "API-INTERNAL-TRANSFER-PREPARE", "R3", "Timeout·502·503·504에 한해 같은 멱등성 키로 최대 1회 재시도", "POST /api/v1/agent-tools/transfers/internal:prepare", "from_account_id, to_account_id, amount, currency, prepare_attempt", "confirmation_id, confirmation_view, correction_view, blocked_view", "ready_for_confirmation → request_internal_transfer_approval | correction_required → route_internal_transfer_correction | blocked → emit_internal_transfer_blocked | 오류 → emit_internal_transfer_error"),
    (12, "request_internal_transfer_approval", "본인송금 승인 요청", "Backend 표시 데이터로 송금 조건을 보여주고 승인·수정·취소 결과를 기다린다.", "webhook_then_resume", "request_confirmation", "UI-INTERNAL-TRANSFER-CONFIRMATION", "R3", "Backend가 Confirmation과 change_target을 검증한 뒤 재개", "component · confirm_modal", "confirmation_id, confirmation_view", "approval_outcome, change_target", "approved → start_internal_auth | 수정 → 대상별 reset Step | cancelled → END | 계약 오류 → emit_internal_transfer_error"),
    (13, "reset_internal_from_account", "출금 계좌 초기화", "출금 계좌 수정 시 관련 계좌 후보와 승인·인증 임시 State를 제거한다.", "agent_internal", "reset_internal_from_account", "", "R2", "to_account_id, amount, correction_view, prepare_attempt는 유지", "", "from_account_hint, from_account_id, account_resolution_outcome, accounts, account_selection_outcome, amount_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target, correction_selection_outcome, blocked_view, auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "from_account_hint, from_account_id, account_resolution_outcome, accounts, account_selection_outcome, amount_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target, correction_selection_outcome, blocked_view, auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "항상 → resolve_internal_from_account"),
    (14, "reset_internal_to_account", "입금 계좌 초기화", "입금 계좌 수정 시 관련 계좌 후보와 승인·인증 임시 State를 제거한다.", "agent_internal", "reset_internal_to_account", "", "R2", "from_account_id, amount, correction_view, prepare_attempt는 유지", "", "to_account_hint, to_account_id, account_resolution_outcome, accounts, account_selection_outcome, amount_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target, correction_selection_outcome, blocked_view, auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "to_account_hint, to_account_id, account_resolution_outcome, accounts, account_selection_outcome, amount_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target, correction_selection_outcome, blocked_view, auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "항상 → resolve_internal_to_account"),
    (15, "reset_internal_transfer_amount", "송금 금액 초기화", "금액 수정 시 금액과 승인·인증 임시 State를 제거한다.", "agent_internal", "reset_internal_transfer_amount", "", "R2", "두 계좌, correction_view, prepare_attempt는 유지", "", "amount, amount_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target, correction_selection_outcome, blocked_view, auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "amount, amount_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target, correction_selection_outcome, blocked_view, auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "항상 → request_internal_transfer_amount"),
    (16, "route_internal_transfer_correction", "본인송금 수정 대상 분기", "Backend가 허용한 수정 대상의 개수와 값을 구조적으로 확인한다.", "agent_internal", "route_internal_transfer_correction", "", "R2", "수정 대상 하나는 바로 이동하고 복수이면 사용자 선택", "", "correction_view", "", "단일 대상 → 해당 reset Step | 복수 → request_internal_transfer_correction | 계약 오류 → emit_internal_transfer_error"),
    (17, "request_internal_transfer_correction", "본인송금 수정 대상 선택", "Backend가 허용한 복수 수정 대상 중 하나를 선택하거나 취소하도록 요청한다.", "webhook_then_resume", "request_component_input", "UI-INTERNAL-TRANSFER-CORRECTION", "R2", "Backend가 allowed_change_targets와 선택값을 검증한 뒤 재개", "component · option_select", "correction_view", "correction_selection_outcome, change_target", "selected → 대상별 reset Step | cancelled → END | 계약 오류 → emit_internal_transfer_error"),
    (18, "start_internal_auth", "본인송금 인증 시도 시작", "이전 인증 State를 제거하고 auth_attempt를 증가시켜 Checkpoint에 저장한다.", "agent_internal", "start_auth_attempt", "", "R3", "통신 재시도에서는 auth_attempt를 증가시키지 않음", "", "auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "항상 → create_internal_auth_context"),
    (19, "create_internal_auth_context", "본인송금 추가 인증 준비", "Backend가 승인된 Confirmation을 확인하고 추가 인증 Context를 생성한다.", "backend_tool_api", "create_auth_context", "API-AUTH-CONTEXT-CREATE", "R3", "인증 원문은 Agent에 반환하지 않음", "POST /api/v1/agent-tools/auth-contexts", "confirmation_id, auth_attempt", "auth_context_id, auth_request_view, blocked_view", "authentication_required → request_internal_authentication | blocked → emit_internal_transfer_blocked | 오류 → emit_internal_transfer_error"),
    (20, "request_internal_authentication", "본인송금 추가 인증 요청", "인증 UI를 표시하고 Backend가 검증한 인증 결과로 Workflow 재개를 기다린다.", "webhook_then_resume", "request_authentication", "UI-INTERNAL-TRANSFER-AUTH", "R3", "PIN·생체인증 Assertion 원문은 Agent가 받지 않음", "component · auth_request", "auth_context_id, auth_request_view", "auth_status", "verified → execute_internal_transfer | failed·expired → request_internal_auth_retry | cancelled → END | 계약 오류 → emit_internal_transfer_error"),
    (21, "request_internal_auth_retry", "본인송금 재인증 선택", "인증 실패·만료 후 재시도 또는 취소 선택 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-INTERNAL-TRANSFER-AUTH-RETRY", "R3", "취소 시 추가 Webhook 없이 종료", "component · option_select", "auth_request_view", "auth_retry_outcome", "retry → start_internal_auth | cancelled → END | 계약 오류 → emit_internal_transfer_error"),
    (22, "execute_internal_transfer", "본인송금 실행", "Backend가 승인·인증·잔액·한도·정책을 재검증하고 원장 이체를 실행한다.", "backend_tool_api", "execute_internal_transfer", "API-INTERNAL-TRANSFER-EXECUTE", "R4", "멱등성 키는 confirmation_id와 auth_attempt로 생성", "POST /api/v1/agent-tools/transfers/internal", "confirmation_id, auth_context_id, auth_attempt", "transaction_id, completed_at, correction_view, blocked_view", "completed → emit_internal_transfer_result | correction_required → route_internal_transfer_correction | reauthentication_required → start_internal_auth | blocked → emit_internal_transfer_blocked | 오류 → emit_internal_transfer_error"),
    (23, "emit_internal_transfer_result", "본인송금 완료 결과 전송", "Execute 결과와 Prepare 표시 데이터를 조합해 완료 Webhook을 전송한다.", "webhook", "emit_component", "UI-INTERNAL-TRANSFER-RESULT", "R0", "계좌 표시정보를 다시 조회하거나 생성하지 않음", "component · transfer_result", "transaction_id, completed_at, confirmation_view", "", "항상 → END"),
    (24, "emit_internal_transfer_blocked", "본인송금 차단 안내", "Backend가 제공한 사용자 표시용 차단 안내를 그대로 전송한다.", "webhook", "emit_blocked", "UI-TRANSFER-BLOCKED", "R0", "자동 재시도하지 않고 내부 정책 사유를 해석하지 않음", "blocked · blocked_message", "blocked_view", "", "항상 → END"),
    (25, "emit_internal_transfer_error", "본인송금 오류 안내", "공통 재시도 이후에도 완료하지 못한 기술·계약 오류를 안내한다.", "webhook", "emit_error", "UI-COMMON-ERROR", "R0", "내부 예외 내용을 사용자에게 노출하지 않음", "error · error_message", "", "", "항상 → END"),
]

INTERNAL_TRANSFER_STEP_ROWS = [
    {
        "workflow_id": "wf_internal_transfer",
        "step_order": step_order,
        "step_id": step_id,
        "step_name": step_name,
        "step_purpose": step_purpose,
        "interaction_mode": interaction_mode,
        "tool_id": tool_id,
        "contract_id": contract_id,
        "step_risk_level": step_risk_level,
        "status": "review",
        "notes": notes,
        "external_action": external_action,
        "input_state_keys": input_state_keys,
        "output_state_keys": output_state_keys,
        "route_summary": route_summary,
        "validation_result": "OK",
    }
    for (
        step_order,
        step_id,
        step_name,
        step_purpose,
        interaction_mode,
        tool_id,
        contract_id,
        step_risk_level,
        notes,
        external_action,
        input_state_keys,
        output_state_keys,
        route_summary,
    ) in INTERNAL_TRANSFER_STEP_DEFINITIONS
]

INTERNAL_TRANSFER_ROUTE_DEFINITIONS = [
    ("extract_internal_transfer_slots", "조건 추출 완료", "사용자 발화에서 선택적 계좌 힌트와 금액 추출을 마친 경우", "resolve_internal_from_account", "값이 없어도 계좌 확인을 계속함"),
    ("resolve_internal_from_account", "출금 계좌 자동 확정", "Backend가 resolved와 정확히 하나의 account_ids를 반환한 경우", "resolve_internal_to_account", "account_ids[0]을 from_account_id로 저장"),
    ("resolve_internal_from_account", "출금 계좌 선택 필요", "Backend가 selection_required와 최신 계좌 후보를 반환한 경우", "request_from_account_selection", ""),
    ("resolve_internal_from_account", "출금 가능 계좌 없음", "Backend가 no_accounts와 accounts=[]을 반환한 경우", "emit_internal_from_accounts_empty", "정상 빈 상태"),
    ("resolve_internal_from_account", "출금 계좌 확인 오류", "Backend 오류이거나 resolved 배열 길이가 정확히 하나가 아닌 경우", "emit_internal_transfer_error", "공통 재시도 이후 오류 Route"),
    ("request_from_account_selection", "출금 계좌 선택 완료", "Backend가 selected와 정확히 하나의 account_ids로 재개한 경우", "resolve_internal_to_account", "account_ids[0]을 from_account_id로 저장"),
    ("request_from_account_selection", "출금 계좌 선택 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_from_account_selection", "출금 계좌 선택 계약 오류", "selected인데 account_ids가 정확히 하나가 아닌 경우", "emit_internal_transfer_error", ""),
    ("emit_internal_from_accounts_empty", "출금 계좌 빈 상태 완료", "빈 상태 Webhook 전송을 완료한 경우", "END", ""),
    ("resolve_internal_to_account", "입금 계좌 자동 확정", "Backend가 resolved와 정확히 하나의 account_ids를 반환한 경우", "check_internal_transfer_amount", "account_ids[0]을 to_account_id로 저장"),
    ("resolve_internal_to_account", "입금 계좌 선택 필요", "Backend가 selection_required와 최신 계좌 후보를 반환한 경우", "request_to_account_selection", "출금 계좌 제외"),
    ("resolve_internal_to_account", "입금 가능 계좌 없음", "Backend가 no_accounts와 accounts=[]을 반환한 경우", "emit_internal_to_accounts_empty", "정상 빈 상태"),
    ("resolve_internal_to_account", "입금 계좌 확인 오류", "Backend 오류이거나 resolved 배열 길이가 정확히 하나가 아닌 경우", "emit_internal_transfer_error", "공통 재시도 이후 오류 Route"),
    ("request_to_account_selection", "입금 계좌 선택 완료", "Backend가 selected와 정확히 하나의 account_ids로 재개한 경우", "check_internal_transfer_amount", "account_ids[0]을 to_account_id로 저장"),
    ("request_to_account_selection", "입금 계좌 선택 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_to_account_selection", "입금 계좌 선택 계약 오류", "selected인데 account_ids가 정확히 하나가 아닌 경우", "emit_internal_transfer_error", ""),
    ("emit_internal_to_accounts_empty", "입금 계좌 빈 상태 완료", "빈 상태 Webhook 전송을 완료한 경우", "END", ""),
    ("check_internal_transfer_amount", "금액 사용 가능", "amount가 정수이고 0보다 큰 경우", "start_internal_transfer_prepare", "금융 정책 검증은 Backend가 수행"),
    ("check_internal_transfer_amount", "금액 입력 필요", "amount가 없거나 구조적으로 유효하지 않은 경우", "request_internal_transfer_amount", ""),
    ("request_internal_transfer_amount", "금액 입력 완료", "Backend가 submitted와 유효한 amount로 재개한 경우", "start_internal_transfer_prepare", ""),
    ("request_internal_transfer_amount", "금액 입력 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_internal_transfer_amount", "금액 입력 계약 오류", "재개 결과와 amount가 계약에 맞지 않는 경우", "emit_internal_transfer_error", ""),
    ("start_internal_transfer_prepare", "Prepare 시도 저장 완료", "prepare_attempt 증가와 Checkpoint 저장을 완료한 경우", "prepare_internal_transfer", ""),
    ("prepare_internal_transfer", "승인 준비 완료", "Backend가 ready_for_confirmation과 Confirmation을 반환한 경우", "request_internal_transfer_approval", ""),
    ("prepare_internal_transfer", "입력 수정 필요", "Backend가 correction_required와 correction_view를 반환한 경우", "route_internal_transfer_correction", ""),
    ("prepare_internal_transfer", "송금 차단", "Backend가 blocked와 blocked_view를 반환한 경우", "emit_internal_transfer_blocked", "자동 재시도 없음"),
    ("prepare_internal_transfer", "Prepare 오류", "재시도 대상이 아니거나 최대 1회 재시도 후에도 실패한 경우", "emit_internal_transfer_error", ""),
    ("request_internal_transfer_approval", "송금 승인 완료", "Backend가 approved로 재개한 경우", "start_internal_auth", "모든 본인송금에 추가 인증 필수"),
    ("request_internal_transfer_approval", "출금 계좌 수정", "Backend가 change_requested와 from_account로 재개한 경우", "reset_internal_from_account", "기존 Confirmation 무효화"),
    ("request_internal_transfer_approval", "입금 계좌 수정", "Backend가 change_requested와 to_account로 재개한 경우", "reset_internal_to_account", "기존 Confirmation 무효화"),
    ("request_internal_transfer_approval", "금액 수정", "Backend가 change_requested와 amount로 재개한 경우", "reset_internal_transfer_amount", "기존 Confirmation 무효화"),
    ("request_internal_transfer_approval", "송금 승인 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_internal_transfer_approval", "승인 결과 계약 오류", "승인 결과 또는 수정 대상이 계약과 일치하지 않는 경우", "emit_internal_transfer_error", ""),
    ("reset_internal_from_account", "출금 계좌 초기화 완료", "관련 임시 State 제거를 완료한 경우", "resolve_internal_from_account", "to_account_id를 제외 조건으로 사용"),
    ("reset_internal_to_account", "입금 계좌 초기화 완료", "관련 임시 State 제거를 완료한 경우", "resolve_internal_to_account", "from_account_id를 제외 조건으로 사용"),
    ("reset_internal_transfer_amount", "금액 초기화 완료", "관련 임시 State 제거를 완료한 경우", "request_internal_transfer_amount", "두 계좌 유지"),
    ("route_internal_transfer_correction", "출금 계좌 단일 수정", "허용 수정 대상이 from_account 하나인 경우", "reset_internal_from_account", "사용자 선택 UI 생략"),
    ("route_internal_transfer_correction", "입금 계좌 단일 수정", "허용 수정 대상이 to_account 하나인 경우", "reset_internal_to_account", "사용자 선택 UI 생략"),
    ("route_internal_transfer_correction", "금액 단일 수정", "허용 수정 대상이 amount 하나인 경우", "reset_internal_transfer_amount", "사용자 선택 UI 생략"),
    ("route_internal_transfer_correction", "복수 수정 대상", "허용 수정 대상이 두 개 이상인 경우", "request_internal_transfer_correction", ""),
    ("route_internal_transfer_correction", "수정 대상 계약 오류", "허용 목록이 비었거나 허용되지 않은 값이 포함된 경우", "emit_internal_transfer_error", "Agent가 수정 대상을 추측하지 않음"),
    ("request_internal_transfer_correction", "출금 계좌 수정 선택", "Backend가 selected와 from_account로 재개한 경우", "reset_internal_from_account", ""),
    ("request_internal_transfer_correction", "입금 계좌 수정 선택", "Backend가 selected와 to_account로 재개한 경우", "reset_internal_to_account", ""),
    ("request_internal_transfer_correction", "금액 수정 선택", "Backend가 selected와 amount로 재개한 경우", "reset_internal_transfer_amount", ""),
    ("request_internal_transfer_correction", "수정 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_internal_transfer_correction", "수정 선택 계약 오류", "선택 결과나 change_target이 허용 목록과 일치하지 않는 경우", "emit_internal_transfer_error", ""),
    ("start_internal_auth", "인증 시도 저장 완료", "기존 인증 State 제거와 auth_attempt 증가를 완료한 경우", "create_internal_auth_context", ""),
    ("create_internal_auth_context", "인증 준비 완료", "Backend가 authentication_required와 Auth Context를 반환한 경우", "request_internal_authentication", ""),
    ("create_internal_auth_context", "인증 진행 차단", "Backend가 blocked와 blocked_view를 반환한 경우", "emit_internal_transfer_blocked", ""),
    ("create_internal_auth_context", "인증 준비 오류", "재시도 대상이 아니거나 최대 1회 재시도 후에도 실패한 경우", "emit_internal_transfer_error", ""),
    ("request_internal_authentication", "인증 완료", "Backend가 verified로 재개한 경우", "execute_internal_transfer", ""),
    ("request_internal_authentication", "재인증 선택 필요", "Backend가 failed 또는 expired로 재개한 경우", "request_internal_auth_retry", ""),
    ("request_internal_authentication", "인증 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_internal_authentication", "인증 결과 계약 오류", "auth_status가 계약된 Enum과 일치하지 않는 경우", "emit_internal_transfer_error", ""),
    ("request_internal_auth_retry", "재인증", "Backend가 retry로 재개한 경우", "start_internal_auth", "새 Auth Context 생성"),
    ("request_internal_auth_retry", "재인증 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_internal_auth_retry", "재인증 선택 계약 오류", "auth_retry_outcome이 계약된 Enum과 일치하지 않는 경우", "emit_internal_transfer_error", ""),
    ("execute_internal_transfer", "송금 실행 완료", "Backend가 completed와 거래 결과를 반환한 경우", "emit_internal_transfer_result", ""),
    ("execute_internal_transfer", "실행 시점 입력 수정 필요", "Backend가 correction_required와 correction_view를 반환한 경우", "route_internal_transfer_correction", "Prepare와 승인·인증을 다시 수행"),
    ("execute_internal_transfer", "재인증 필요", "Confirmation은 유효하지만 Backend가 reauthentication_required를 반환한 경우", "start_internal_auth", "Prepare와 승인 생략"),
    ("execute_internal_transfer", "실행 차단", "Backend가 blocked와 blocked_view를 반환한 경우", "emit_internal_transfer_blocked", "자동 재시도 없음"),
    ("execute_internal_transfer", "Execute 오류", "재시도 대상이 아니거나 최대 1회 재시도 후에도 실패한 경우", "emit_internal_transfer_error", "같은 인증 시도에서는 같은 키와 Body 유지"),
    ("emit_internal_transfer_result", "완료 결과 전송 완료", "본인송금 완료 Webhook 전송을 완료한 경우", "END", ""),
    ("emit_internal_transfer_blocked", "차단 안내 완료", "본인송금 차단 Webhook 전송을 완료한 경우", "END", ""),
    ("emit_internal_transfer_error", "오류 안내 완료", "본인송금 오류 Webhook 전송을 완료한 경우", "END", ""),
]

INTERNAL_TRANSFER_ROUTE_ROWS = [
    {
        "workflow_id": "wf_internal_transfer",
        "from_step_id": from_step_id,
        "route_name": route_name,
        "condition_description": condition_description,
        "to_step_id": to_step_id,
        "status": "review",
        "notes": notes,
    }
    for from_step_id, route_name, condition_description, to_step_id, notes in INTERNAL_TRANSFER_ROUTE_DEFINITIONS
]

EXTERNAL_TRANSFER_STEP_DEFINITIONS = [
    (1, "extract_external_transfer_slots", "타인송금 조건 추출", "사용자 발화에서 출금 계좌 힌트, 수취인 이름 힌트와 금액을 추출한다.", "agent_internal", "extract_external_transfer_slots", "", "R2", "수취인·계좌 확정과 금융 검증은 수행하지 않음", "", "", "from_account_hint, recipient_name_hint, amount", "이름 힌트 있음 → resolve_recipient_hint | 없음 → request_recipient_selection"),
    (2, "resolve_recipient_hint", "기존 거래 수취인 자동 확정", "Backend가 완료된 기존 타인송금 거래에서 이름과 정확히 일치하는 고유 수취인을 확인한다.", "backend_tool_api", "resolve_recipient_hint", "API-RECIPIENT-RESOLVE", "R2", "후보 목록은 Agent에 반환하지 않음", "POST /api/v1/agent-tools/recipients:resolve", "recipient_name_hint", "recipient_resolution_outcome, recipient_selection_reason, to_recipient_id", "resolved → resolve_external_from_account | selection_required → request_recipient_selection | 오류 → emit_external_transfer_error"),
    (3, "request_recipient_selection", "타인송금 수취인 선택 요청", "수취인 선택 UI를 요청하고 Backend가 검증한 기존 또는 신규 수취인 참조를 기다린다.", "webhook_then_resume", "request_component_input", "UI-RECIPIENT-SELECT", "R2", "검색·신규 계좌 검증 중에는 Agent를 재개하지 않음", "component · recipient_select", "recipient_name_hint, recipient_selection_reason, correction_view", "recipient_selection_outcome, to_recipient_id, to_recipient_candidate_id", "selected → resolve_external_from_account | cancelled → END | 계약 오류 → emit_external_transfer_error"),
    (4, "resolve_external_from_account", "타인송금 출금 계좌 확인", "Backend가 출금 가능한 본인 계좌를 검증하고 자동 확정 여부를 결정한다.", "backend_tool_api", "fetch_accounts", "API-ACCOUNT-LIST", "R2", "account_capability=withdraw, resolve_selection=true", "GET /api/v1/agent-tools/accounts", "from_account_hint", "account_resolution_outcome, accounts, from_account_id", "resolved → check_external_transfer_amount | selection_required → request_external_from_account_selection | no_accounts → emit_external_from_accounts_empty | 오류 → emit_external_transfer_error"),
    (5, "request_external_from_account_selection", "타인송금 출금 계좌 선택", "검증된 출금 계좌 후보를 표시하고 단일 선택 또는 취소 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-EXTERNAL-TRANSFER-FROM-ACCOUNT", "R2", "account_ids 배열 길이가 정확히 하나여야 함", "component · account_card_list", "accounts, correction_view", "account_selection_outcome, from_account_id", "selected → check_external_transfer_amount | cancelled → END | 계약 오류 → emit_external_transfer_error"),
    (6, "emit_external_from_accounts_empty", "타인송금 출금 가능 계좌 없음", "출금 가능한 계좌가 없는 빈 상태를 동일한 계좌 선택 UI로 전송한다.", "webhook", "emit_component", "UI-EXTERNAL-TRANSFER-FROM-ACCOUNT", "R0", "accounts=[]을 전송하고 사용자 회신을 기다리지 않음", "component · account_card_list", "accounts", "", "항상 → END"),
    (7, "check_external_transfer_amount", "타인송금 금액 구조 확인", "금액이 존재하고 정수이며 0보다 큰지 구조적으로 확인한다.", "agent_internal", "check_external_transfer_amount", "", "R2", "잔액·한도·정책은 Backend가 검증", "", "amount", "", "유효 → start_external_transfer_prepare | 없거나 구조 오류 → request_external_transfer_amount"),
    (8, "request_external_transfer_amount", "타인송금 금액 입력", "금액 입력 UI를 표시하고 Backend가 검증한 제출 또는 취소 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-TRANSFER-AMOUNT-INPUT", "R2", "입력 검증 실패는 Agent를 재개하지 않고 같은 UI에서 처리", "component · number_input", "correction_view, currency", "amount_input_outcome, amount", "submitted → start_external_transfer_prepare | cancelled → END | 계약 오류 → emit_external_transfer_error"),
    (9, "start_external_transfer_prepare", "타인송금 Prepare 시도 시작", "이전 수정 사유를 제거하고 새 Prepare 시도 번호를 증가시켜 Checkpoint에 저장한다.", "agent_internal", "start_prepare_attempt", "", "R2", "통신 재시도에서는 prepare_attempt를 증가시키지 않음", "", "prepare_attempt, correction_view", "prepare_attempt, correction_view", "항상 → prepare_external_transfer"),
    (10, "prepare_external_transfer", "타인송금 준비", "Backend가 출금 계좌·수취인·금액, 잔액·한도·정책을 검증하고 Confirmation을 준비한다.", "backend_tool_api", "prepare_external_transfer", "API-EXTERNAL-TRANSFER-PREPARE", "R3", "두 수취인 참조 중 정확히 하나만 전달", "POST /api/v1/agent-tools/transfers/external:prepare", "from_account_id, to_recipient_id, to_recipient_candidate_id, amount, currency, prepare_attempt", "confirmation_id, confirmation_view, correction_view, blocked_view", "ready_for_confirmation → request_external_transfer_approval | correction_required → route_external_transfer_correction | blocked → emit_external_transfer_blocked | 오류 → emit_external_transfer_error"),
    (11, "request_external_transfer_approval", "타인송금 승인 요청", "Backend 표시 데이터로 송금 조건을 보여주고 승인·수정·취소 결과를 기다린다.", "webhook_then_resume", "request_confirmation", "UI-EXTERNAL-TRANSFER-CONFIRMATION", "R3", "경고가 있으면 동일 승인 화면의 warning variant 사용", "component · confirm_modal", "confirmation_id, confirmation_view", "approval_outcome, change_target", "approved → start_external_auth | 수정 → 대상별 reset Step | cancelled → END | 계약 오류 → emit_external_transfer_error"),
    (12, "reset_external_from_account", "타인송금 출금 계좌 초기화", "출금 계좌 수정 시 계좌 후보와 승인·인증 임시 State를 제거한다.", "agent_internal", "reset_external_from_account", "", "R2", "수취인 참조, amount, correction_view, prepare_attempt 유지", "", "from_account_hint, account_resolution_outcome, accounts, account_selection_outcome, from_account_id, amount_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target, correction_selection_outcome, blocked_view, auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "from_account_hint, account_resolution_outcome, accounts, account_selection_outcome, from_account_id, amount_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target, correction_selection_outcome, blocked_view, auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "항상 → resolve_external_from_account"),
    (13, "reset_external_recipient", "타인송금 수취인 초기화", "수취인 수정 시 이름 힌트·수취인 참조와 승인·인증 임시 State를 제거한다.", "agent_internal", "reset_external_recipient", "", "R2", "from_account_id, amount, correction_view, prepare_attempt 유지", "", "recipient_name_hint, recipient_resolution_outcome, recipient_selection_reason, recipient_selection_outcome, to_recipient_id, to_recipient_candidate_id, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target, correction_selection_outcome, blocked_view, auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "recipient_name_hint, recipient_resolution_outcome, recipient_selection_reason, recipient_selection_outcome, to_recipient_id, to_recipient_candidate_id, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target, correction_selection_outcome, blocked_view, auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "항상 → request_recipient_selection"),
    (14, "reset_external_transfer_amount", "타인송금 금액 초기화", "금액 수정 시 금액과 승인·인증 임시 State를 제거한다.", "agent_internal", "reset_external_transfer_amount", "", "R2", "출금 계좌·수취인 참조, correction_view, prepare_attempt 유지", "", "amount, amount_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target, correction_selection_outcome, blocked_view, auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "amount, amount_input_outcome, input_request_id, confirmation_id, confirmation_view, approval_outcome, change_target, correction_selection_outcome, blocked_view, auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "항상 → request_external_transfer_amount"),
    (15, "route_external_transfer_correction", "타인송금 수정 대상 분기", "Backend가 허용한 수정 대상의 개수와 값을 구조적으로 확인한다.", "agent_internal", "route_external_transfer_correction", "", "R2", "단일 대상은 바로 이동하고 복수이면 사용자 선택", "", "correction_view", "", "단일 대상 → reset Step | 복수 → request_external_transfer_correction | 계약 오류 → emit_external_transfer_error"),
    (16, "request_external_transfer_correction", "타인송금 수정 대상 선택", "Backend가 허용한 복수 수정 대상 중 하나를 선택하거나 취소하도록 요청한다.", "webhook_then_resume", "request_component_input", "UI-EXTERNAL-TRANSFER-CORRECTION", "R2", "Backend가 allowed_change_targets와 선택값을 검증", "component · option_select", "correction_view", "correction_selection_outcome, change_target", "selected → 대상별 reset Step | cancelled → END | 계약 오류 → emit_external_transfer_error"),
    (17, "start_external_auth", "타인송금 인증 시도 시작", "이전 인증 State를 제거하고 auth_attempt를 증가시켜 Checkpoint에 저장한다.", "agent_internal", "start_auth_attempt", "", "R3", "통신 재시도에서는 auth_attempt를 증가시키지 않음", "", "auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "auth_context_id, auth_request_view, auth_status, auth_retry_outcome, auth_attempt", "항상 → create_external_auth_context"),
    (18, "create_external_auth_context", "타인송금 추가 인증 준비", "Backend가 승인된 Confirmation을 확인하고 추가 인증 Context를 생성한다.", "backend_tool_api", "create_auth_context", "API-AUTH-CONTEXT-CREATE", "R3", "인증 원문은 Agent에 반환하지 않음", "POST /api/v1/agent-tools/auth-contexts", "confirmation_id, auth_attempt", "auth_context_id, auth_request_view, blocked_view", "authentication_required → request_external_authentication | blocked → emit_external_transfer_blocked | 오류 → emit_external_transfer_error"),
    (19, "request_external_authentication", "타인송금 추가 인증 요청", "인증 UI를 표시하고 Backend가 검증한 인증 결과로 Workflow 재개를 기다린다.", "webhook_then_resume", "request_authentication", "UI-EXTERNAL-TRANSFER-AUTH", "R3", "PIN·생체인증 Assertion 원문은 Agent가 받지 않음", "component · auth_request", "auth_context_id, auth_request_view", "auth_status", "verified → execute_external_transfer | failed·expired → request_external_auth_retry | cancelled → END | 계약 오류 → emit_external_transfer_error"),
    (20, "request_external_auth_retry", "타인송금 재인증 선택", "인증 실패·만료 후 재시도 또는 취소 선택 결과를 기다린다.", "webhook_then_resume", "request_component_input", "UI-EXTERNAL-TRANSFER-AUTH-RETRY", "R3", "취소 시 추가 Webhook 없이 종료", "component · option_select", "auth_request_view", "auth_retry_outcome", "retry → start_external_auth | cancelled → END | 계약 오류 → emit_external_transfer_error"),
    (21, "execute_external_transfer", "타인송금 실행", "Backend가 승인·인증·잔액·한도·정책을 재검증하고 원장 송금을 실행한다.", "backend_tool_api", "execute_external_transfer", "API-EXTERNAL-TRANSFER-EXECUTE", "R4", "멱등성 키는 confirmation_id와 auth_attempt로 생성", "POST /api/v1/agent-tools/transfers/external", "confirmation_id, auth_context_id, auth_attempt", "transaction_id, completed_at, correction_view, blocked_view", "completed → emit_external_transfer_result | correction_required → route_external_transfer_correction | reauthentication_required → start_external_auth | blocked → emit_external_transfer_blocked | 오류 → emit_external_transfer_error"),
    (22, "emit_external_transfer_result", "타인송금 완료 결과 전송", "Execute 결과와 Prepare 표시 데이터를 조합해 완료 Webhook을 전송한다.", "webhook", "emit_component", "UI-EXTERNAL-TRANSFER-RESULT", "R0", "계좌와 수취인 표시정보를 다시 조회하지 않음", "component · transfer_result", "transaction_id, completed_at, confirmation_view", "", "항상 → END"),
    (23, "emit_external_transfer_blocked", "타인송금 차단 안내", "Backend가 제공한 사용자 표시용 차단 안내를 그대로 전송한다.", "webhook", "emit_blocked", "UI-TRANSFER-BLOCKED", "R0", "자동 재시도하지 않고 내부 정책 사유를 해석하지 않음", "blocked · blocked_message", "blocked_view", "", "항상 → END"),
    (24, "emit_external_transfer_error", "타인송금 오류 안내", "공통 재시도 이후에도 완료하지 못한 기술·계약 오류를 안내한다.", "webhook", "emit_error", "UI-COMMON-ERROR", "R0", "내부 예외 내용을 사용자에게 노출하지 않음", "error · error_message", "", "", "항상 → END"),
]

EXTERNAL_TRANSFER_STEP_ROWS = [
    {
        "workflow_id": "wf_external_transfer", "step_order": step_order, "step_id": step_id,
        "step_name": step_name, "step_purpose": step_purpose, "interaction_mode": interaction_mode,
        "tool_id": tool_id, "contract_id": contract_id, "step_risk_level": step_risk_level,
        "status": "review", "notes": notes, "external_action": external_action,
        "input_state_keys": input_state_keys, "output_state_keys": output_state_keys,
        "route_summary": route_summary, "validation_result": "OK",
    }
    for (step_order, step_id, step_name, step_purpose, interaction_mode, tool_id, contract_id,
         step_risk_level, notes, external_action, input_state_keys, output_state_keys, route_summary)
    in EXTERNAL_TRANSFER_STEP_DEFINITIONS
]

EXTERNAL_TRANSFER_ROUTE_DEFINITIONS = [
    ("extract_external_transfer_slots", "이름 힌트 있음", "recipient_name_hint가 존재하는 경우", "resolve_recipient_hint", ""),
    ("extract_external_transfer_slots", "이름 힌트 없음", "recipient_name_hint가 없는 경우", "request_recipient_selection", "초기 수취인 선택 UI"),
    ("resolve_recipient_hint", "수취인 자동 확정", "Backend가 resolved와 to_recipient_id를 반환한 경우", "resolve_external_from_account", "별도 수취인 확인 UI 생략"),
    ("resolve_recipient_hint", "수취인 선택 필요", "Backend가 selection_required와 허용된 selection_reason을 반환한 경우", "request_recipient_selection", "후보 목록은 Agent가 받지 않음"),
    ("resolve_recipient_hint", "수취인 확인 오류", "Backend 오류 또는 outcome과 참조값이 계약에 맞지 않는 경우", "emit_external_transfer_error", ""),
    ("request_recipient_selection", "수취인 선택 완료", "selected이며 두 수취인 참조 중 정확히 하나가 존재하는 경우", "resolve_external_from_account", ""),
    ("request_recipient_selection", "수취인 선택 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_recipient_selection", "수취인 선택 계약 오류", "selected인데 수취인 참조가 없거나 두 개 모두 존재하는 경우", "emit_external_transfer_error", ""),
    ("resolve_external_from_account", "출금 계좌 자동 확정", "Backend가 resolved와 정확히 하나의 account_ids를 반환한 경우", "check_external_transfer_amount", ""),
    ("resolve_external_from_account", "출금 계좌 선택 필요", "Backend가 selection_required와 최신 계좌 후보를 반환한 경우", "request_external_from_account_selection", ""),
    ("resolve_external_from_account", "출금 가능 계좌 없음", "Backend가 no_accounts와 accounts=[]을 반환한 경우", "emit_external_from_accounts_empty", "정상 빈 상태"),
    ("resolve_external_from_account", "출금 계좌 확인 오류", "Backend 오류이거나 resolved 배열 길이가 정확히 하나가 아닌 경우", "emit_external_transfer_error", ""),
    ("request_external_from_account_selection", "출금 계좌 선택 완료", "Backend가 selected와 정확히 하나의 account_ids로 재개한 경우", "check_external_transfer_amount", ""),
    ("request_external_from_account_selection", "출금 계좌 선택 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_external_from_account_selection", "출금 계좌 선택 계약 오류", "selected인데 account_ids가 정확히 하나가 아닌 경우", "emit_external_transfer_error", ""),
    ("emit_external_from_accounts_empty", "빈 상태 전송 완료", "출금 가능 계좌 없음 Webhook 전송을 완료한 경우", "END", ""),
    ("check_external_transfer_amount", "금액 사용 가능", "amount가 정수이고 0보다 큰 경우", "start_external_transfer_prepare", ""),
    ("check_external_transfer_amount", "금액 입력 필요", "amount가 없거나 구조적으로 유효하지 않은 경우", "request_external_transfer_amount", ""),
    ("request_external_transfer_amount", "금액 입력 완료", "Backend가 submitted와 유효한 amount로 재개한 경우", "start_external_transfer_prepare", ""),
    ("request_external_transfer_amount", "금액 입력 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_external_transfer_amount", "금액 입력 계약 오류", "재개 결과와 amount가 계약에 맞지 않는 경우", "emit_external_transfer_error", ""),
    ("start_external_transfer_prepare", "Prepare 시도 저장 완료", "prepare_attempt 증가와 Checkpoint 저장을 완료한 경우", "prepare_external_transfer", ""),
    ("prepare_external_transfer", "승인 준비 완료", "Backend가 ready_for_confirmation과 Confirmation을 반환한 경우", "request_external_transfer_approval", ""),
    ("prepare_external_transfer", "입력 수정 필요", "Backend가 correction_required와 correction_view를 반환한 경우", "route_external_transfer_correction", ""),
    ("prepare_external_transfer", "송금 차단", "Backend가 blocked와 blocked_view를 반환한 경우", "emit_external_transfer_blocked", "자동 재시도 없음"),
    ("prepare_external_transfer", "Prepare 오류", "재시도 대상이 아니거나 최대 1회 재시도 후에도 실패한 경우", "emit_external_transfer_error", ""),
    ("request_external_transfer_approval", "송금 승인 완료", "Backend가 approved로 재개한 경우", "start_external_auth", "모든 타인송금에 추가 인증 필수"),
    ("request_external_transfer_approval", "출금 계좌 수정", "Backend가 change_requested와 from_account로 재개한 경우", "reset_external_from_account", ""),
    ("request_external_transfer_approval", "수취인 수정", "Backend가 change_requested와 recipient로 재개한 경우", "reset_external_recipient", ""),
    ("request_external_transfer_approval", "금액 수정", "Backend가 change_requested와 amount로 재개한 경우", "reset_external_transfer_amount", ""),
    ("request_external_transfer_approval", "송금 승인 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_external_transfer_approval", "승인 결과 계약 오류", "승인 결과 또는 수정 대상이 계약과 일치하지 않는 경우", "emit_external_transfer_error", ""),
    ("reset_external_from_account", "출금 계좌 초기화 완료", "관련 임시 State 제거를 완료한 경우", "resolve_external_from_account", ""),
    ("reset_external_recipient", "수취인 초기화 완료", "관련 임시 State 제거를 완료한 경우", "request_recipient_selection", "이름 자동 확정은 다시 수행하지 않음"),
    ("reset_external_transfer_amount", "금액 초기화 완료", "관련 임시 State 제거를 완료한 경우", "request_external_transfer_amount", ""),
    ("route_external_transfer_correction", "출금 계좌 단일 수정", "허용 수정 대상이 from_account 하나인 경우", "reset_external_from_account", ""),
    ("route_external_transfer_correction", "수취인 단일 수정", "허용 수정 대상이 recipient 하나인 경우", "reset_external_recipient", ""),
    ("route_external_transfer_correction", "금액 단일 수정", "허용 수정 대상이 amount 하나인 경우", "reset_external_transfer_amount", ""),
    ("route_external_transfer_correction", "복수 수정 대상", "허용 수정 대상이 두 개 이상인 경우", "request_external_transfer_correction", ""),
    ("route_external_transfer_correction", "수정 대상 계약 오류", "허용 목록이 비었거나 허용되지 않은 값이 포함된 경우", "emit_external_transfer_error", ""),
    ("request_external_transfer_correction", "출금 계좌 수정 선택", "Backend가 selected와 from_account로 재개한 경우", "reset_external_from_account", ""),
    ("request_external_transfer_correction", "수취인 수정 선택", "Backend가 selected와 recipient로 재개한 경우", "reset_external_recipient", ""),
    ("request_external_transfer_correction", "금액 수정 선택", "Backend가 selected와 amount로 재개한 경우", "reset_external_transfer_amount", ""),
    ("request_external_transfer_correction", "수정 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_external_transfer_correction", "수정 선택 계약 오류", "선택 결과나 change_target이 허용 목록과 일치하지 않는 경우", "emit_external_transfer_error", ""),
    ("start_external_auth", "인증 시도 저장 완료", "기존 인증 State 제거와 auth_attempt 증가를 완료한 경우", "create_external_auth_context", ""),
    ("create_external_auth_context", "인증 준비 완료", "Backend가 authentication_required와 Auth Context를 반환한 경우", "request_external_authentication", ""),
    ("create_external_auth_context", "인증 진행 차단", "Backend가 blocked와 blocked_view를 반환한 경우", "emit_external_transfer_blocked", ""),
    ("create_external_auth_context", "인증 준비 오류", "재시도 대상이 아니거나 최대 1회 재시도 후에도 실패한 경우", "emit_external_transfer_error", ""),
    ("request_external_authentication", "인증 완료", "Backend가 verified로 재개한 경우", "execute_external_transfer", ""),
    ("request_external_authentication", "재인증 선택 필요", "Backend가 failed 또는 expired로 재개한 경우", "request_external_auth_retry", ""),
    ("request_external_authentication", "인증 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_external_authentication", "인증 결과 계약 오류", "auth_status가 계약된 Enum과 일치하지 않는 경우", "emit_external_transfer_error", ""),
    ("request_external_auth_retry", "재인증", "Backend가 retry로 재개한 경우", "start_external_auth", ""),
    ("request_external_auth_retry", "재인증 취소", "Backend가 cancelled로 재개한 경우", "END", "추가 Webhook 없이 종료"),
    ("request_external_auth_retry", "재인증 선택 계약 오류", "auth_retry_outcome이 계약된 Enum과 일치하지 않는 경우", "emit_external_transfer_error", ""),
    ("execute_external_transfer", "송금 실행 완료", "Backend가 completed와 거래 결과를 반환한 경우", "emit_external_transfer_result", ""),
    ("execute_external_transfer", "실행 시점 입력 수정 필요", "Backend가 correction_required와 correction_view를 반환한 경우", "route_external_transfer_correction", ""),
    ("execute_external_transfer", "재인증 필요", "Confirmation은 유효하지만 Backend가 reauthentication_required를 반환한 경우", "start_external_auth", "Prepare와 승인 생략"),
    ("execute_external_transfer", "실행 차단", "Backend가 blocked와 blocked_view를 반환한 경우", "emit_external_transfer_blocked", ""),
    ("execute_external_transfer", "Execute 오류", "재시도 대상이 아니거나 최대 1회 재시도 후에도 실패한 경우", "emit_external_transfer_error", ""),
    ("emit_external_transfer_result", "완료 결과 전송 완료", "타인송금 완료 Webhook 전송을 완료한 경우", "END", ""),
    ("emit_external_transfer_blocked", "차단 안내 완료", "타인송금 차단 Webhook 전송을 완료한 경우", "END", ""),
    ("emit_external_transfer_error", "오류 안내 완료", "타인송금 오류 Webhook 전송을 완료한 경우", "END", ""),
]

EXTERNAL_TRANSFER_ROUTE_ROWS = [
    {"workflow_id": "wf_external_transfer", "from_step_id": from_step_id, "route_name": route_name,
     "condition_description": condition_description, "to_step_id": to_step_id, "status": "review", "notes": notes}
    for from_step_id, route_name, condition_description, to_step_id, notes in EXTERNAL_TRANSFER_ROUTE_DEFINITIONS
]

COMMON_SCHEMA_ROWS = [
    ("chat_session_id", "string", False, "Backend Chat Session 식별자", True, "masked"),
    ("execution_context_id", "string", False, "Backend가 발급한 실행 Context 식별자", True, "masked"),
    ("agent_thread_id", "string", False, "중단과 재개에 사용하는 LangGraph Thread 식별자", True, "masked"),
    ("workflow_id", "string", False, "현재 실행 Workflow 식별자", False, "allow"),
    ("workflow_version", "string", False, "현재 Workflow 계약 버전", False, "allow"),
    ("requested_at", "datetime", False, "사용자 요청 접수 시각", False, "allow"),
    ("timezone", "string", False, "날짜와 기간 해석에 사용하는 시간대", False, "allow"),
]

SCHEMA_ROWS = [
    {
        "schema_scope": "common",
        "workflow_id": "",
        "state_key": state_key,
        "data_type": data_type,
        "nullable": str(nullable).lower(),
        "default_value": "",
        "description": description,
        "retention_scope": "workflow",
        "clear_when": "Workflow 종료",
        "sensitive": str(sensitive).lower(),
        "log_policy": log_policy,
        "notes": "모든 Workflow에 병합",
    }
    for state_key, data_type, nullable, description, sensitive, log_policy in COMMON_SCHEMA_ROWS
]

GLOBAL_SCHEMA_DEFINITIONS = [
    ("guardrail_outcome", "string", "true", "null", "전역 가드레일 분류 결과", "interaction", "Route 결정 또는 Workflow 종료", "false", "allow", "allowed, blocked"),
    ("blocked_view", "GlobalBlockedView", "true", "null", "전역 정책 차단 시 사용자에게 표시할 안전한 안내 데이터", "interaction", "차단 안내 전송 또는 Workflow 종료", "true", "exclude", "내부 정책 세부 사유는 포함하지 않음"),
    ("workflow_match_outcome", "string", "true", "null", "지원 Workflow 매칭 결과", "interaction", "Route 결정 또는 Workflow 종료", "false", "allow", "matched, no_match"),
    ("matched_workflow_id", "string", "true", "null", "매칭되어 실행할 지원 Workflow 식별자", "workflow", "하위 Workflow 실행 완료 또는 Workflow 종료", "false", "allow", "Workflow Catalog에 등록된 업무 Workflow만 허용"),
    ("dispatch_outcome", "string", "true", "null", "하위 Workflow 디스패치 종료 결과", "interaction", "Route 결정 또는 Workflow 종료", "false", "allow", "completed, failed"),
]

SCHEMA_ROWS.extend(
    [
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_global_agent_entry",
            "state_key": state_key,
            "data_type": data_type,
            "nullable": nullable,
            "default_value": default_value,
            "description": description,
            "retention_scope": retention_scope,
            "clear_when": clear_when,
            "sensitive": sensitive,
            "log_policy": log_policy,
            "notes": notes,
        }
        for (
            state_key,
            data_type,
            nullable,
            default_value,
            description,
            retention_scope,
            clear_when,
            sensitive,
            log_policy,
            notes,
        ) in GLOBAL_SCHEMA_DEFINITIONS
    ]
)

SCHEMA_ROWS.extend(
    [
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_account_list",
            "state_key": "account_hint",
            "data_type": "string",
            "nullable": "true",
            "default_value": "null",
            "description": "사용자 발화에서 추출한 선택적 계좌 검색 힌트",
            "retention_scope": "workflow",
            "clear_when": "Workflow 종료",
            "sensitive": "true",
            "log_policy": "masked",
            "notes": "은행명, 계좌 별칭 또는 계좌 유형 힌트",
        },
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_account_list",
            "state_key": "account_results",
            "data_type": "list[AccountSummary]",
            "nullable": "false",
            "default_value": "[]",
            "description": "Backend가 반환한 마스킹된 최종 계좌 목록",
            "retention_scope": "result",
            "clear_when": "결과 보존 정책에 따름",
            "sensitive": "true",
            "log_policy": "exclude",
            "notes": "전체 계좌번호와 잔액을 포함하지 않음",
        },
    ]
)

SCHEMA_ROWS.extend(
    [
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_balance_inquiry",
            "state_key": "account_hint",
            "data_type": "string",
            "nullable": "true",
            "default_value": "null",
            "description": "사용자 발화에서 추출한 선택적 계좌 검색 힌트",
            "retention_scope": "workflow",
            "clear_when": "Workflow 종료",
            "sensitive": "true",
            "log_policy": "masked",
            "notes": "은행명, 계좌 별칭 또는 계좌 유형 힌트",
        },
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_balance_inquiry",
            "state_key": "all_accounts_requested",
            "data_type": "boolean",
            "nullable": "false",
            "default_value": "false",
            "description": "사용자가 전체 계좌 잔액 조회를 명시했는지 여부",
            "retention_scope": "workflow",
            "clear_when": "Workflow 종료",
            "sensitive": "false",
            "log_policy": "allow",
            "notes": "Backend 자동 확정 요청에 그대로 전달",
        },
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_balance_inquiry",
            "state_key": "account_resolution_outcome",
            "data_type": "string",
            "nullable": "true",
            "default_value": "null",
            "description": "Backend가 반환한 계좌 자동 확정 결과",
            "retention_scope": "interaction",
            "clear_when": "계좌 확정, 사용자 취소 또는 Workflow 종료",
            "sensitive": "false",
            "log_policy": "allow",
            "notes": "resolved, selection_required, no_accounts",
        },
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_balance_inquiry",
            "state_key": "accounts",
            "data_type": "list[AccountCandidate]",
            "nullable": "false",
            "default_value": "[]",
            "description": "Backend가 검증하여 반환한 마스킹된 계좌 선택 후보",
            "retention_scope": "interaction",
            "clear_when": "계좌 확정, 사용자 취소 또는 Workflow 종료",
            "sensitive": "true",
            "log_policy": "exclude",
            "notes": "전체 계좌번호와 잔액을 포함하지 않음",
        },
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_balance_inquiry",
            "state_key": "account_ids",
            "data_type": "list[string]",
            "nullable": "false",
            "default_value": "[]",
            "description": "Backend가 자동 확정하거나 사용자 선택 후 검증한 잔액 조회 계좌 ID",
            "retention_scope": "workflow",
            "clear_when": "Workflow 종료",
            "sensitive": "true",
            "log_policy": "masked",
            "notes": "단일 계좌도 배열 사용",
        },
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_balance_inquiry",
            "state_key": "account_selection_outcome",
            "data_type": "string",
            "nullable": "true",
            "default_value": "null",
            "description": "계좌 선택 UI의 검증된 재개 결과",
            "retention_scope": "interaction",
            "clear_when": "계좌 선택 Route 결정 후",
            "sensitive": "false",
            "log_policy": "allow",
            "notes": "selected, cancelled",
        },
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_balance_inquiry",
            "state_key": "input_request_id",
            "data_type": "string",
            "nullable": "true",
            "default_value": "null",
            "description": "계좌 선택 요청과 Backend resume을 연결하는 식별자",
            "retention_scope": "interaction",
            "clear_when": "Backend resume 검증 완료 또는 Workflow 종료",
            "sensitive": "true",
            "log_policy": "masked",
            "notes": "사용자 입력 대기 중에만 보존",
        },
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_balance_inquiry",
            "state_key": "balance_results",
            "data_type": "list[BalanceResult]",
            "nullable": "false",
            "default_value": "[]",
            "description": "Backend가 반환한 계좌별 잔액 조회 결과",
            "retention_scope": "result",
            "clear_when": "결과 보존 정책에 따름",
            "sensitive": "true",
            "log_policy": "exclude",
            "notes": "잔액 원문을 Agent Trace와 일반 로그에 기록하지 않음",
        },
    ]
)

TRANSACTION_SCHEMA_DEFINITIONS = [
    ("account_hint", "string", "true", "null", "사용자 발화에서 추출한 선택적 계좌 검색 힌트", "workflow", "Workflow 종료", "true", "masked", "은행명, 계좌 별칭 또는 계좌 유형 힌트"),
    ("all_accounts_requested", "boolean", "false", "false", "사용자가 전체 계좌 거래내역 조회를 명시했는지 여부", "workflow", "Workflow 종료", "false", "allow", "Backend 자동 확정 요청에 그대로 전달"),
    ("account_resolution_outcome", "string", "true", "null", "Backend가 반환한 계좌 자동 확정 결과", "interaction", "계좌 확정, 사용자 취소 또는 Workflow 종료", "false", "allow", "resolved, selection_required, no_accounts"),
    ("accounts", "list[AccountCandidate]", "false", "[]", "Backend가 검증하여 반환한 마스킹된 계좌 선택 후보", "interaction", "계좌 확정, 사용자 취소 또는 Workflow 종료", "true", "exclude", "전체 계좌번호와 잔액을 포함하지 않음"),
    ("account_ids", "list[string]", "false", "[]", "Backend가 자동 확정하거나 사용자 선택 후 검증한 거래내역 조회 계좌 ID", "workflow", "Workflow 종료", "true", "masked", "단일 계좌도 배열 사용"),
    ("account_selection_outcome", "string", "true", "null", "계좌 선택 UI의 검증된 재개 결과", "interaction", "계좌 선택 Route 결정 후", "false", "allow", "selected, cancelled"),
    ("input_request_id", "string", "true", "null", "계좌 또는 기간 입력 요청과 Backend resume을 연결하는 식별자", "interaction", "Backend resume 검증 완료 또는 Workflow 종료", "true", "masked", "사용자 입력 대기 중에만 보존"),
    ("start_date", "date", "true", "null", "거래내역 조회 시작일", "workflow", "Workflow 종료", "false", "allow", "기간 미입력 시 최근 1개월 기본값 적용"),
    ("end_date", "date", "true", "null", "거래내역 조회 종료일", "workflow", "Workflow 종료", "false", "allow", "requested_at과 timezone 기준"),
    ("period_selection_outcome", "string", "true", "null", "기간 선택 UI의 검증된 재개 결과", "interaction", "기간 선택 Route 결정 후", "false", "allow", "selected, cancelled"),
    ("keyword", "string", "true", "null", "선택적 거래내역 검색어", "workflow", "Workflow 종료", "true", "exclude", "최대 100자이며 Backend가 정규화"),
    ("transaction_type", "string", "true", "null", "선택적 TransactionType 거래 유형 필터", "workflow", "Workflow 종료", "false", "allow", "Enum Registry transaction_type 참조"),
    ("transaction_results", "list[TransactionItem]", "false", "[]", "Backend가 반환한 첫 페이지 거래내역", "result", "결과 보존 정책에 따름", "true", "exclude", "거래 금액과 제목 원문을 일반 로그에 기록하지 않음"),
    ("transaction_query_id", "string", "true", "null", "Frontend와 Backend의 이후 페이지 조회용 Query Context", "result", "Query Context 만료 또는 결과 보존 정책에 따름", "true", "masked", "Agent는 이후 페이지 조회에 사용하지 않음"),
    ("next_cursor", "string", "true", "null", "Frontend의 다음 페이지 조회용 불투명 Cursor", "result", "Query Context 만료 또는 결과 보존 정책에 따름", "true", "exclude", "다음 페이지가 없으면 null"),
]

SCHEMA_ROWS.extend(
    [
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_transaction_history",
            "state_key": state_key,
            "data_type": data_type,
            "nullable": nullable,
            "default_value": default_value,
            "description": description,
            "retention_scope": retention_scope,
            "clear_when": clear_when,
            "sensitive": sensitive,
            "log_policy": log_policy,
            "notes": notes,
        }
        for (
            state_key,
            data_type,
            nullable,
            default_value,
            description,
            retention_scope,
            clear_when,
            sensitive,
            log_policy,
            notes,
        ) in TRANSACTION_SCHEMA_DEFINITIONS
    ]
)

SUMMARY_SCHEMA_DEFINITIONS = [
    ("account_hint", "string", "true", "null", "사용자 발화에서 추출한 선택적 계좌 검색 힌트", "workflow", "Workflow 종료", "true", "masked", "힌트가 없으면 전체 계좌 집계"),
    ("all_accounts_requested", "boolean", "false", "true", "계좌 힌트가 없을 때 전체 계좌 집계 여부", "workflow", "Workflow 종료", "false", "allow", "합계 조회의 기본값은 true"),
    ("account_resolution_outcome", "string", "true", "null", "Backend가 반환한 계좌 자동 확정 결과", "interaction", "계좌 확정, 사용자 취소 또는 Workflow 종료", "false", "allow", "resolved, selection_required, no_accounts"),
    ("accounts", "list[AccountCandidate]", "false", "[]", "Backend가 검증하여 반환한 마스킹된 계좌 선택 후보", "interaction", "계좌 확정, 사용자 취소 또는 Workflow 종료", "true", "exclude", "전체 계좌번호와 잔액을 포함하지 않음"),
    ("account_ids", "list[string]", "false", "[]", "Backend가 자동 확정하거나 사용자 선택 후 검증한 집계 대상 계좌 ID", "workflow", "Workflow 종료", "true", "masked", "단일 계좌도 배열 사용"),
    ("account_selection_outcome", "string", "true", "null", "계좌 선택 UI의 검증된 재개 결과", "interaction", "계좌 선택 Route 결정 후", "false", "allow", "selected, cancelled"),
    ("input_request_id", "string", "true", "null", "계좌·기간·합계 유형 입력 요청과 Backend resume을 연결하는 식별자", "interaction", "Backend resume 검증 완료 또는 Workflow 종료", "true", "masked", "사용자 입력 대기 중에만 보존"),
    ("start_date", "date", "true", "null", "합계 조회 시작일", "workflow", "Workflow 종료", "false", "allow", "기간 미입력 시 최근 1개월 기본값 적용"),
    ("end_date", "date", "true", "null", "합계 조회 종료일", "workflow", "Workflow 종료", "false", "allow", "requested_at과 timezone 기준"),
    ("period_selection_outcome", "string", "true", "null", "기간 선택 UI의 검증된 재개 결과", "interaction", "기간 선택 Route 결정 후", "false", "allow", "selected, cancelled"),
    ("summary_type", "string", "true", "null", "지출 또는 수입 합계 유형", "workflow", "Workflow 종료", "false", "allow", "spending, income"),
    ("summary_type_selection_outcome", "string", "true", "null", "합계 유형 선택 UI의 검증된 재개 결과", "interaction", "합계 유형 Route 결정 후", "false", "allow", "selected, cancelled"),
    ("keyword", "string", "true", "null", "선택적 가맹점·상대방 검색어", "workflow", "Workflow 종료", "true", "exclude", "최대 100자이며 Backend가 정규화"),
    ("summary_result", "AmountSummary", "true", "null", "Backend가 계산한 기간 거래 합계 결과", "result", "결과 보존 정책에 따름", "true", "exclude", "Agent가 거래내역을 받거나 직접 합산하지 않음"),
]

SCHEMA_ROWS.extend(
    [
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_period_amount_summary",
            "state_key": state_key,
            "data_type": data_type,
            "nullable": nullable,
            "default_value": default_value,
            "description": description,
            "retention_scope": retention_scope,
            "clear_when": clear_when,
            "sensitive": sensitive,
            "log_policy": log_policy,
            "notes": notes,
        }
        for (
            state_key,
            data_type,
            nullable,
            default_value,
            description,
            retention_scope,
            clear_when,
            sensitive,
            log_policy,
            notes,
        ) in SUMMARY_SCHEMA_DEFINITIONS
    ]
)

DEFAULT_ACCOUNT_SCHEMA_DEFINITIONS = [
    ("account_hint", "string", "true", "null", "사용자 발화에서 추출한 새 기본 출금 계좌 힌트", "workflow", "대상 초기화 또는 Workflow 종료", "true", "masked", "은행명, 계좌 별칭 또는 계좌 유형 힌트"),
    ("account_resolution_outcome", "string", "true", "null", "Backend가 반환한 계좌 자동 확정 결과", "interaction", "대상 확정, 대상 초기화 또는 Workflow 종료", "false", "allow", "resolved, selection_required, no_accounts"),
    ("accounts", "list[AccountCandidate]", "false", "[]", "Backend가 검증하여 반환한 마스킹 계좌 선택 후보", "interaction", "대상 확정, 대상 초기화 또는 Workflow 종료", "true", "exclude", "항상 최신 API 응답으로 교체"),
    ("account_id", "string", "true", "null", "새 기본 출금 계좌로 설정할 검증된 계좌 ID", "workflow", "대상 초기화 또는 Workflow 종료", "true", "masked", "API와 resume의 account_ids 배열에서 정확히 한 값을 저장"),
    ("account_selection_outcome", "string", "true", "null", "계좌 선택 UI의 검증된 재개 결과", "interaction", "Route 결정, 대상 초기화 또는 Workflow 종료", "false", "allow", "selected, cancelled"),
    ("input_request_id", "string", "true", "null", "계좌 선택 요청과 Backend resume을 연결하는 식별자", "interaction", "Backend resume 검증, 대상 초기화 또는 Workflow 종료", "true", "masked", "사용자 입력 대기 중에만 보존"),
    ("confirmation_id", "string", "true", "null", "Backend가 발급한 기본계좌 변경 승인 식별자", "workflow", "대상 초기화, 실행 완료 또는 Workflow 종료", "true", "masked", "Execute에는 승인된 식별자만 전달"),
    ("confirmation_view", "ConfirmationView", "true", "null", "승인 화면에 표시할 변경 전후 마스킹 데이터", "interaction", "승인 결과 처리, 대상 초기화 또는 Workflow 종료", "true", "exclude", "승인 대기 중에만 보존"),
    ("approval_outcome", "string", "true", "null", "Backend가 검증한 승인 재개 결과", "interaction", "Route 결정, 대상 초기화 또는 Workflow 종료", "false", "allow", "approved, change_requested, cancelled"),
    ("correction_view", "SettingCorrectionView", "true", "null", "Backend가 반환한 수정 가능 대상 표시 데이터", "interaction", "새 Prepare 시작 또는 Workflow 종료", "true", "exclude", "다음 계좌 선택 UI까지 유지하며 허용 수정 대상은 account"),
    ("prepare_attempt", "integer", "false", "0", "Prepare 멱등성 키 생성을 위한 논리 시도 번호", "workflow", "Workflow 종료", "false", "allow", "통신 재시도에서는 증가시키지 않음"),
    ("completed_at", "datetime", "true", "null", "기본 출금 계좌 변경 완료 시각", "result", "결과 보존 정책에 따름", "false", "allow", "Backend Execute 응답 기준"),
]

SCHEMA_ROWS.extend(
    [
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_set_default_account",
            "state_key": state_key,
            "data_type": data_type,
            "nullable": nullable,
            "default_value": default_value,
            "description": description,
            "retention_scope": retention_scope,
            "clear_when": clear_when,
            "sensitive": sensitive,
            "log_policy": log_policy,
            "notes": notes,
        }
        for (
            state_key,
            data_type,
            nullable,
            default_value,
            description,
            retention_scope,
            clear_when,
            sensitive,
            log_policy,
            notes,
        ) in DEFAULT_ACCOUNT_SCHEMA_DEFINITIONS
    ]
)

ACCOUNT_ALIAS_SCHEMA_DEFINITIONS = [
    ("account_hint", "string", "true", "null", "사용자 발화에서 추출한 별칭 변경 대상 계좌 힌트", "workflow", "계좌 초기화 또는 Workflow 종료", "true", "masked", "은행명, 계좌 별칭 또는 계좌 유형 힌트"),
    ("account_resolution_outcome", "string", "true", "null", "Backend가 반환한 계좌 자동 확정 결과", "interaction", "계좌 확정, 계좌 초기화 또는 Workflow 종료", "false", "allow", "resolved, selection_required, no_accounts"),
    ("accounts", "list[AccountCandidate]", "false", "[]", "Backend가 검증하여 반환한 마스킹 계좌 선택 후보", "interaction", "계좌 확정, 계좌 초기화 또는 Workflow 종료", "true", "exclude", "항상 최신 API 응답으로 교체"),
    ("account_id", "string", "true", "null", "별칭을 변경할 검증된 계좌 ID", "workflow", "계좌 초기화 또는 Workflow 종료", "true", "masked", "account_ids 배열에서 정확히 한 값을 저장"),
    ("account_selection_outcome", "string", "true", "null", "계좌 선택 UI의 검증된 재개 결과", "interaction", "Route 결정, 계좌 초기화 또는 Workflow 종료", "false", "allow", "selected, cancelled"),
    ("alias", "string", "true", "null", "Backend가 검증·정규화한 새 계좌 별칭", "workflow", "별칭 초기화 또는 Workflow 종료", "true", "exclude", "account_label, current_alias, new_alias를 별도 저장하지 않음"),
    ("alias_input_outcome", "string", "true", "null", "별칭 입력 UI의 검증된 재개 결과", "interaction", "Route 결정, 값 초기화 또는 Workflow 종료", "false", "allow", "submitted, cancelled"),
    ("input_request_id", "string", "true", "null", "계좌 선택 또는 별칭 입력과 Backend resume을 연결하는 식별자", "interaction", "Backend resume 검증, 초기화 또는 Workflow 종료", "true", "masked", "사용자 입력 대기 중에만 보존"),
    ("confirmation_id", "string", "true", "null", "Backend가 발급한 계좌 별칭 변경 승인 식별자", "workflow", "초기화, 실행 완료 또는 Workflow 종료", "true", "masked", "Execute에는 승인된 식별자만 전달"),
    ("confirmation_view", "ConfirmationView", "true", "null", "승인 화면에 표시할 계좌와 최종 별칭 데이터", "interaction", "승인 결과 처리, 초기화 또는 Workflow 종료", "true", "exclude", "account 객체와 alias만 포함"),
    ("approval_outcome", "string", "true", "null", "Backend가 검증한 승인 재개 결과", "interaction", "Route 결정, 초기화 또는 Workflow 종료", "false", "allow", "approved, change_requested, cancelled"),
    ("change_target", "string", "true", "null", "사용자가 수정할 계좌 별칭 변경 대상", "interaction", "수정 Route 진입 또는 Workflow 종료", "false", "allow", "account, alias"),
    ("correction_view", "SettingCorrectionView", "true", "null", "Backend가 반환한 단일 수정 대상과 표시 사유", "interaction", "새 Prepare 시작 또는 Workflow 종료", "true", "exclude", "다음 선택·입력 UI까지 유지"),
    ("prepare_attempt", "integer", "false", "0", "Prepare 멱등성 키 생성을 위한 논리 시도 번호", "workflow", "Workflow 종료", "false", "allow", "통신 재시도에서는 증가시키지 않음"),
    ("completed_at", "datetime", "true", "null", "계좌 별칭 변경 완료 시각", "result", "결과 보존 정책에 따름", "false", "allow", "Backend Execute 응답 기준"),
]

SCHEMA_ROWS.extend(
    [
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_set_account_alias",
            "state_key": state_key,
            "data_type": data_type,
            "nullable": nullable,
            "default_value": default_value,
            "description": description,
            "retention_scope": retention_scope,
            "clear_when": clear_when,
            "sensitive": sensitive,
            "log_policy": log_policy,
            "notes": notes,
        }
        for (
            state_key,
            data_type,
            nullable,
            default_value,
            description,
            retention_scope,
            clear_when,
            sensitive,
            log_policy,
            notes,
        ) in ACCOUNT_ALIAS_SCHEMA_DEFINITIONS
    ]
)

INTERNAL_TRANSFER_SCHEMA_DEFINITIONS = [
    ("from_account_hint", "string", "true", "null", "사용자 발화에서 추출한 출금 계좌 힌트", "workflow", "출금 계좌 초기화 또는 Workflow 종료", "true", "masked", "은행명, 계좌 별칭 또는 계좌 유형 힌트"),
    ("to_account_hint", "string", "true", "null", "사용자 발화에서 추출한 입금 계좌 힌트", "workflow", "입금 계좌 초기화 또는 Workflow 종료", "true", "masked", "은행명, 계좌 별칭 또는 계좌 유형 힌트"),
    ("account_resolution_outcome", "string", "true", "null", "현재 계좌 확인 Step의 Backend 자동 확정 결과", "interaction", "계좌 확정, 계좌 초기화 또는 Workflow 종료", "false", "allow", "resolved, selection_required, no_accounts"),
    ("accounts", "list[AccountCandidate]", "false", "[]", "현재 계좌 선택 단계의 Backend 검증 UI 후보", "interaction", "계좌 확정, 계좌 초기화 또는 Workflow 종료", "true", "exclude", "출금과 입금 후보를 동시에 저장하지 않고 최신 응답으로 교체"),
    ("account_selection_outcome", "string", "true", "null", "현재 계좌 선택 UI의 검증된 재개 결과", "interaction", "Route 결정, 계좌 초기화 또는 Workflow 종료", "false", "allow", "selected, cancelled"),
    ("from_account_id", "string", "true", "null", "출금할 검증된 본인 계좌 ID", "workflow", "출금 계좌 초기화 또는 Workflow 종료", "true", "masked", "account_ids 배열에서 정확히 한 값을 저장"),
    ("to_account_id", "string", "true", "null", "입금할 검증된 본인 계좌 ID", "workflow", "입금 계좌 초기화 또는 Workflow 종료", "true", "masked", "from_account_id와 달라야 함"),
    ("amount", "integer", "true", "null", "KRW 최소 화폐 단위의 이체 금액", "workflow", "금액 초기화 또는 Workflow 종료", "true", "exclude", "Agent는 정수이고 0보다 큰지만 확인"),
    ("amount_input_outcome", "string", "true", "null", "금액 입력 UI의 검증된 재개 결과", "interaction", "Route 결정, 값 초기화 또는 Workflow 종료", "false", "allow", "submitted, cancelled"),
    ("currency", "string", "false", "KRW", "본인송금 통화", "workflow", "Workflow 종료", "false", "allow", "현재 KRW만 허용"),
    ("input_request_id", "string", "true", "null", "계좌·금액·수정 입력과 Backend 재개를 연결하는 식별자", "interaction", "Backend 재개 검증, 초기화 또는 Workflow 종료", "true", "masked", "사용자 입력 대기 중에만 보존"),
    ("confirmation_id", "string", "true", "null", "Backend가 발급한 본인송금 승인 식별자", "workflow", "입력 초기화, 실행 완료 또는 Workflow 종료", "true", "masked", "Prepare에서 고정한 송금 조건 참조"),
    ("confirmation_view", "ConfirmationView", "true", "null", "승인 화면과 완료 결과에서 재사용할 Backend 표시 데이터", "workflow", "입력 초기화, 결과 전송 또는 Workflow 종료", "true", "exclude", "마스킹 계좌, 금액과 통화 포함"),
    ("approval_outcome", "string", "true", "null", "Backend가 검증한 승인 재개 결과", "interaction", "Route 결정, 입력 초기화 또는 Workflow 종료", "false", "allow", "approved, change_requested, cancelled"),
    ("change_target", "string", "true", "null", "사용자가 수정할 본인송금 조건", "interaction", "수정 Route 진입, 초기화 또는 Workflow 종료", "false", "allow", "from_account, to_account, amount"),
    ("correction_view", "TransferCorrectionView", "true", "null", "Backend가 반환한 수정 가능 대상과 표시 사유", "interaction", "새 Prepare 시작 또는 Workflow 종료", "true", "exclude", "다음 계좌·금액 UI까지 유지"),
    ("correction_selection_outcome", "string", "true", "null", "복수 수정 대상 선택 UI의 검증된 재개 결과", "interaction", "수정 Route 진입, 초기화 또는 Workflow 종료", "false", "allow", "selected, cancelled"),
    ("blocked_view", "BlockedView", "true", "null", "Backend가 제공한 사용자 표시용 송금 차단 안내", "interaction", "차단 결과 전송 또는 Workflow 종료", "true", "exclude", "Agent가 내부 정책 사유를 해석하지 않음"),
    ("auth_context_id", "string", "true", "null", "Backend가 발급한 추가 인증 Context ID", "workflow", "재인증, 입력 초기화, 실행 완료 또는 Workflow 종료", "true", "masked", "인증 원문은 저장하지 않음"),
    ("auth_request_view", "AuthRequestView", "true", "null", "인증 UI에 표시할 Backend 제공 데이터", "interaction", "인증 결과 처리, 재인증 또는 Workflow 종료", "true", "exclude", "인증 방식과 만료 시각 포함"),
    ("auth_status", "string", "true", "null", "Backend가 검증한 추가 인증 결과", "interaction", "Route 결정, 재인증 또는 Workflow 종료", "false", "allow", "verified, failed, cancelled, expired"),
    ("auth_retry_outcome", "string", "true", "null", "재인증 선택 UI의 검증된 재개 결과", "interaction", "Route 결정, 재인증 또는 Workflow 종료", "false", "allow", "retry, cancelled"),
    ("prepare_attempt", "integer", "false", "0", "Prepare 멱등성 키 생성을 위한 논리 시도 번호", "workflow", "Workflow 종료", "false", "allow", "통신 재시도에서는 증가시키지 않음"),
    ("auth_attempt", "integer", "false", "0", "Auth Context와 송금 Execute 멱등성 키 생성을 위한 인증 시도 번호", "workflow", "입력 초기화 또는 Workflow 종료", "false", "allow", "재인증에서만 증가하고 통신 재시도에서는 유지"),
    ("transaction_id", "string", "true", "null", "완료된 본인송금 거래 ID", "result", "결과 보존 정책에 따름", "true", "masked", "Backend Execute 응답 기준"),
    ("completed_at", "datetime", "true", "null", "본인송금 완료 시각", "result", "결과 보존 정책에 따름", "false", "allow", "Backend Execute 응답 기준"),
]

SCHEMA_ROWS.extend(
    [
        {
            "schema_scope": "workflow",
            "workflow_id": "wf_internal_transfer",
            "state_key": state_key,
            "data_type": data_type,
            "nullable": nullable,
            "default_value": default_value,
            "description": description,
            "retention_scope": retention_scope,
            "clear_when": clear_when,
            "sensitive": sensitive,
            "log_policy": log_policy,
            "notes": notes,
        }
        for (
            state_key,
            data_type,
            nullable,
            default_value,
            description,
            retention_scope,
            clear_when,
            sensitive,
            log_policy,
            notes,
        ) in INTERNAL_TRANSFER_SCHEMA_DEFINITIONS
    ]
)

EXTERNAL_TRANSFER_SCHEMA_DEFINITIONS = [
    ("from_account_hint", "string", "true", "null", "사용자 발화에서 추출한 출금 계좌 힌트", "workflow", "출금 계좌 초기화 또는 Workflow 종료", "true", "masked", "은행명, 별칭 또는 계좌 유형 힌트"),
    ("account_resolution_outcome", "string", "true", "null", "Backend 출금 계좌 자동 확정 결과", "interaction", "계좌 확정·초기화 또는 Workflow 종료", "false", "allow", "resolved, selection_required, no_accounts"),
    ("accounts", "list[AccountCandidate]", "false", "[]", "Backend가 검증한 출금 계좌 선택 후보", "interaction", "계좌 확정·초기화 또는 Workflow 종료", "true", "exclude", "항상 최신 응답으로 교체"),
    ("account_selection_outcome", "string", "true", "null", "출금 계좌 선택 UI 재개 결과", "interaction", "Route 결정·초기화 또는 Workflow 종료", "false", "allow", "selected, cancelled"),
    ("from_account_id", "string", "true", "null", "송금액을 출금할 검증된 사용자 계좌 ID", "workflow", "출금 계좌 초기화 또는 Workflow 종료", "true", "masked", "account_ids 배열에서 정확히 한 값을 저장"),
    ("recipient_name_hint", "string", "true", "null", "최초 발화에서 추출한 수취인 이름 힌트", "workflow", "수취인 초기화 또는 Workflow 종료", "true", "masked", "기존 완료 타인송금 거래 자동 확정에만 사용"),
    ("recipient_resolution_outcome", "string", "true", "null", "기존 거래 수취인 자동 확정 결과", "interaction", "수취인 확정·초기화 또는 Workflow 종료", "false", "allow", "resolved, selection_required"),
    ("recipient_selection_reason", "string", "true", "null", "수취인 선택 UI가 필요한 사유", "interaction", "수취인 확정·초기화 또는 Workflow 종료", "false", "allow", "multiple_matches, no_match"),
    ("recipient_selection_outcome", "string", "true", "null", "수취인 선택 UI 재개 결과", "interaction", "Route 결정·초기화 또는 Workflow 종료", "false", "allow", "selected, cancelled"),
    ("to_recipient_id", "string", "true", "null", "Backend가 검증한 기존 수취인 ID", "workflow", "수취인 초기화 또는 Workflow 종료", "true", "masked", "신규 후보 ID와 동시에 존재할 수 없음"),
    ("to_recipient_candidate_id", "string", "true", "null", "Backend가 신규 계좌 검증 후 발급한 수취인 후보 ID", "workflow", "수취인 초기화 또는 Workflow 종료", "true", "masked", "계좌번호 원문은 저장하지 않음"),
    ("amount", "integer", "true", "null", "KRW 최소 화폐 단위의 송금 금액", "workflow", "금액 초기화 또는 Workflow 종료", "true", "exclude", "Agent는 정수이고 0보다 큰지만 확인"),
    ("amount_input_outcome", "string", "true", "null", "금액 입력 UI 재개 결과", "interaction", "Route 결정·초기화 또는 Workflow 종료", "false", "allow", "submitted, cancelled"),
    ("currency", "string", "false", "KRW", "타인송금 통화", "workflow", "Workflow 종료", "false", "allow", "현재 KRW만 허용"),
    ("input_request_id", "string", "true", "null", "수취인·계좌·금액·수정 입력과 재개를 연결하는 식별자", "interaction", "재개 검증·초기화 또는 Workflow 종료", "true", "masked", "입력 대기 중에만 보존"),
    ("confirmation_id", "string", "true", "null", "Backend가 발급한 타인송금 승인 식별자", "workflow", "입력 초기화·실행 완료 또는 Workflow 종료", "true", "masked", "Prepare에서 고정한 송금 조건 참조"),
    ("confirmation_view", "ConfirmationView", "true", "null", "승인 화면과 완료 결과에서 재사용할 Backend 표시 데이터", "workflow", "입력 초기화·결과 전송 또는 Workflow 종료", "true", "exclude", "마스킹 계좌·수취인, 금액과 통화 포함"),
    ("approval_outcome", "string", "true", "null", "Backend가 검증한 승인 재개 결과", "interaction", "Route 결정·초기화 또는 Workflow 종료", "false", "allow", "approved, change_requested, cancelled"),
    ("change_target", "string", "true", "null", "사용자가 수정할 타인송금 조건", "interaction", "수정 Route·초기화 또는 Workflow 종료", "false", "allow", "from_account, recipient, amount"),
    ("correction_view", "TransferCorrectionView", "true", "null", "Backend가 반환한 수정 가능 대상과 표시 사유", "interaction", "새 Prepare 시작 또는 Workflow 종료", "true", "exclude", "다음 입력 UI까지 유지"),
    ("correction_selection_outcome", "string", "true", "null", "복수 수정 대상 선택 UI 재개 결과", "interaction", "수정 Route·초기화 또는 Workflow 종료", "false", "allow", "selected, cancelled"),
    ("blocked_view", "BlockedView", "true", "null", "Backend 사용자 표시용 송금 차단 안내", "interaction", "차단 결과 전송 또는 Workflow 종료", "true", "exclude", "Agent가 내부 정책 사유를 해석하지 않음"),
    ("auth_context_id", "string", "true", "null", "Backend가 발급한 추가 인증 Context ID", "workflow", "재인증·입력 초기화·실행 완료 또는 Workflow 종료", "true", "masked", "인증 원문은 저장하지 않음"),
    ("auth_request_view", "AuthRequestView", "true", "null", "인증 UI에 표시할 Backend 제공 데이터", "interaction", "인증 결과·재인증 또는 Workflow 종료", "true", "exclude", "인증 방식과 만료 시각 포함"),
    ("auth_status", "string", "true", "null", "Backend가 검증한 추가 인증 결과", "interaction", "Route 결정·재인증 또는 Workflow 종료", "false", "allow", "verified, failed, cancelled, expired"),
    ("auth_retry_outcome", "string", "true", "null", "재인증 선택 UI 재개 결과", "interaction", "Route 결정·재인증 또는 Workflow 종료", "false", "allow", "retry, cancelled"),
    ("prepare_attempt", "integer", "false", "0", "Prepare 멱등성 키 논리 시도 번호", "workflow", "Workflow 종료", "false", "allow", "통신 재시도에서는 증가시키지 않음"),
    ("auth_attempt", "integer", "false", "0", "Auth Context와 Execute 멱등성 키 인증 시도 번호", "workflow", "입력 초기화 또는 Workflow 종료", "false", "allow", "재인증에서만 증가"),
    ("transaction_id", "string", "true", "null", "완료된 타인송금 거래 ID", "result", "결과 보존 정책에 따름", "true", "masked", "Backend Execute 응답 기준"),
    ("completed_at", "datetime", "true", "null", "타인송금 완료 시각", "result", "결과 보존 정책에 따름", "false", "allow", "Backend Execute 응답 기준"),
]

SCHEMA_ROWS.extend(
    [
        {"schema_scope": "workflow", "workflow_id": "wf_external_transfer", "state_key": state_key,
         "data_type": data_type, "nullable": nullable, "default_value": default_value,
         "description": description, "retention_scope": retention_scope, "clear_when": clear_when,
         "sensitive": sensitive, "log_policy": log_policy, "notes": notes}
        for (state_key, data_type, nullable, default_value, description, retention_scope,
             clear_when, sensitive, log_policy, notes) in EXTERNAL_TRANSFER_SCHEMA_DEFINITIONS
    ]
)

GLOBAL_MAPPING_DEFINITIONS = [
    ("run_global_guardrail", "output", "guardrail_outcome", "", "true", "사용자 요청의 전역 정책 허용 또는 차단 결과를 저장한다.", "업무 실행 가능성 검증과 구분"),
    ("run_global_guardrail", "output", "blocked_view", "", "false", "차단 시 사용자에게 표시할 안전한 안내 데이터를 생성한다.", "blocked일 때 필수"),
    ("match_workflow", "output", "workflow_match_outcome", "", "true", "지원 Workflow 매칭 여부를 저장한다.", "matched, no_match"),
    ("match_workflow", "output", "matched_workflow_id", "", "false", "매칭된 지원 Workflow ID를 저장한다.", "matched일 때 필수"),
    ("dispatch_matched_workflow", "input", "matched_workflow_id", "", "true", "매칭된 하위 Workflow를 실행 대상으로 사용한다.", "Workflow Catalog 등록값만 허용"),
    ("dispatch_matched_workflow", "output", "dispatch_outcome", "", "true", "하위 Workflow의 정상 종료 또는 디스패치 실패 결과를 저장한다.", "업무 오류를 처리하고 종료한 경우도 completed"),
    ("emit_global_blocked", "input", "blocked_view", "webhook.metadata.ui.payload", "true", "전역 정책 차단 안내 데이터를 Webhook Payload로 전달한다.", "내부 정책 세부 사유 제외"),
]

GLOBAL_MAPPING_ROWS = [
    {
        "workflow_id": "wf_global_agent_entry",
        "step_id": step_id,
        "direction": direction,
        "state_key": state_key,
        "contract_field_path": contract_field_path,
        "required_at_step": required_at_step,
        "mapping_description": mapping_description,
        "notes": notes,
        "validation_result": "OK",
    }
    for (
        step_id,
        direction,
        state_key,
        contract_field_path,
        required_at_step,
        mapping_description,
        notes,
    ) in GLOBAL_MAPPING_DEFINITIONS
]

ACCOUNT_LIST_MAPPING_ROWS = [
    {
        "workflow_id": "wf_account_list",
        "step_id": "extract_account_list_slots",
        "direction": "output",
        "state_key": "account_hint",
        "contract_field_path": "",
        "required_at_step": "false",
        "mapping_description": "사용자 발화에서 선택적인 계좌 검색 힌트를 추출한다.",
        "notes": "힌트가 없으면 null",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_account_list",
        "step_id": "fetch_account_list",
        "direction": "input",
        "state_key": "account_hint",
        "contract_field_path": "query.account_hint",
        "required_at_step": "false",
        "mapping_description": "계좌 검색 힌트를 Backend API Query Parameter로 전달한다.",
        "notes": "limit=20은 Tool 설정값",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_account_list",
        "step_id": "fetch_account_list",
        "direction": "output",
        "state_key": "account_results",
        "contract_field_path": "response.data.accounts",
        "required_at_step": "true",
        "mapping_description": "Backend가 반환한 계좌 목록을 최종 결과 State에 저장한다.",
        "notes": "빈 목록도 정상 결과",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_account_list",
        "step_id": "emit_account_list_result",
        "direction": "input",
        "state_key": "account_results",
        "contract_field_path": "webhook.metadata.ui.payload.accounts",
        "required_at_step": "true",
        "mapping_description": "마스킹된 계좌 목록을 결과 UI Payload로 전달한다.",
        "notes": "",
        "validation_result": "OK",
    },
]

BALANCE_MAPPING_ROWS = [
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "extract_balance_slots",
        "direction": "output",
        "state_key": "account_hint",
        "contract_field_path": "",
        "required_at_step": "false",
        "mapping_description": "사용자 발화에서 선택적인 계좌 검색 힌트를 추출한다.",
        "notes": "힌트가 없으면 null",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "extract_balance_slots",
        "direction": "output",
        "state_key": "all_accounts_requested",
        "contract_field_path": "",
        "required_at_step": "true",
        "mapping_description": "사용자가 전체 계좌 잔액 조회를 명시했는지 추출한다.",
        "notes": "기본값 false",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "resolve_balance_accounts",
        "direction": "input",
        "state_key": "account_hint",
        "contract_field_path": "query.account_hint",
        "required_at_step": "false",
        "mapping_description": "계좌 검색 힌트를 Backend 계좌 확인 요청에 전달한다.",
        "notes": "resolve_selection=true, account_capability=inquiry는 Tool 설정값",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "resolve_balance_accounts",
        "direction": "input",
        "state_key": "all_accounts_requested",
        "contract_field_path": "query.all_accounts_requested",
        "required_at_step": "true",
        "mapping_description": "전체 계좌 조회 의도를 Backend 자동 확정 요청에 전달한다.",
        "notes": "Agent와 API에서 같은 필드명 사용",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "resolve_balance_accounts",
        "direction": "output",
        "state_key": "account_resolution_outcome",
        "contract_field_path": "response.data.account_resolution_outcome",
        "required_at_step": "true",
        "mapping_description": "Backend의 계좌 자동 확정 결과를 Route 값으로 저장한다.",
        "notes": "Agent가 후보 개수로 다시 판단하지 않음",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "resolve_balance_accounts",
        "direction": "output",
        "state_key": "accounts",
        "contract_field_path": "response.data.accounts",
        "required_at_step": "true",
        "mapping_description": "Backend가 검증한 마스킹 계좌 후보를 저장한다.",
        "notes": "no_accounts이면 빈 배열",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "resolve_balance_accounts",
        "direction": "output",
        "state_key": "account_ids",
        "contract_field_path": "response.data.account_ids",
        "required_at_step": "false",
        "mapping_description": "Backend가 자동 확정한 검증된 계좌 ID를 저장한다.",
        "notes": "resolved일 때 필수",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "request_balance_account_selection",
        "direction": "input",
        "state_key": "accounts",
        "contract_field_path": "webhook.metadata.ui.payload.accounts",
        "required_at_step": "true",
        "mapping_description": "검증된 복수 계좌 후보를 선택 UI Payload로 전달한다.",
        "notes": "selection_mode=multiple은 UI 계약값",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "request_balance_account_selection",
        "direction": "output",
        "state_key": "account_selection_outcome",
        "contract_field_path": "resume.value.account_selection_outcome",
        "required_at_step": "true",
        "mapping_description": "Backend가 검증한 선택 또는 취소 결과를 저장한다.",
        "notes": "selected, cancelled",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "request_balance_account_selection",
        "direction": "output",
        "state_key": "account_ids",
        "contract_field_path": "resume.value.account_ids",
        "required_at_step": "false",
        "mapping_description": "선택 완료 시 Backend가 검증한 계좌 ID를 저장한다.",
        "notes": "selected일 때 필수, cancelled이면 빈 배열",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "emit_balance_accounts_empty",
        "direction": "input",
        "state_key": "accounts",
        "contract_field_path": "webhook.metadata.ui.payload.accounts",
        "required_at_step": "true",
        "mapping_description": "빈 계좌 목록을 동일한 선택 UI의 빈 상태로 전달한다.",
        "notes": "accounts=[]이며 resume을 기다리지 않음",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "query_balances",
        "direction": "input",
        "state_key": "account_ids",
        "contract_field_path": "request.account_ids",
        "required_at_step": "true",
        "mapping_description": "검증된 단일·복수 계좌 ID를 잔액 일괄 조회 요청에 전달한다.",
        "notes": "최소 1개, 최대 20개, 중복 불가",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "query_balances",
        "direction": "output",
        "state_key": "balance_results",
        "contract_field_path": "response.data.balance_results",
        "required_at_step": "true",
        "mapping_description": "Backend가 반환한 계좌별 잔액 결과를 저장한다.",
        "notes": "부분 결과를 허용하지 않음",
        "validation_result": "OK",
    },
    {
        "workflow_id": "wf_balance_inquiry",
        "step_id": "emit_balance_result",
        "direction": "input",
        "state_key": "balance_results",
        "contract_field_path": "webhook.metadata.ui.payload.accounts",
        "required_at_step": "true",
        "mapping_description": "계좌별 잔액 결과를 결과 UI Payload로 전달한다.",
        "notes": "잔액 원문은 일반 로그에서 제외",
        "validation_result": "OK",
    },
]

TRANSACTION_MAPPING_DEFINITIONS = [
    ("extract_transaction_slots", "output", "account_hint", "", "false", "사용자 발화에서 선택적인 계좌 검색 힌트를 추출한다.", "힌트가 없으면 null"),
    ("extract_transaction_slots", "output", "all_accounts_requested", "", "true", "전체 계좌 거래내역 조회 의도를 추출한다.", "기본값 false"),
    ("extract_transaction_slots", "output", "start_date", "", "false", "사용자 발화의 조회 시작일을 정규화한다.", "기간 미입력 또는 해석 실패이면 null"),
    ("extract_transaction_slots", "output", "end_date", "", "false", "사용자 발화의 조회 종료일을 정규화한다.", "기간 미입력 또는 해석 실패이면 null"),
    ("extract_transaction_slots", "output", "keyword", "", "false", "선택적인 거래 검색어를 추출한다.", "원문은 일반 로그에서 제외"),
    ("extract_transaction_slots", "output", "transaction_type", "", "false", "선택적인 TransactionType 필터를 추출한다.", "Enum Registry 참조"),
    ("resolve_transaction_accounts", "input", "account_hint", "query.account_hint", "false", "계좌 검색 힌트를 Backend 계좌 확인 요청에 전달한다.", "resolve_selection=true, account_capability=inquiry는 Tool 설정값"),
    ("resolve_transaction_accounts", "input", "all_accounts_requested", "query.all_accounts_requested", "true", "전체 계좌 조회 의도를 Backend 자동 확정 요청에 전달한다.", "Agent와 API에서 같은 필드명 사용"),
    ("resolve_transaction_accounts", "output", "account_resolution_outcome", "response.data.account_resolution_outcome", "true", "Backend의 계좌 자동 확정 결과를 Route 값으로 저장한다.", "후보 개수로 다시 판단하지 않음"),
    ("resolve_transaction_accounts", "output", "accounts", "response.data.accounts", "true", "Backend가 검증한 마스킹 계좌 후보를 저장한다.", "no_accounts이면 빈 배열"),
    ("resolve_transaction_accounts", "output", "account_ids", "response.data.account_ids", "false", "Backend가 자동 확정한 검증된 계좌 ID를 저장한다.", "resolved일 때 필수"),
    ("request_transaction_account_selection", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "검증된 복수 계좌 후보를 선택 UI Payload로 전달한다.", "selection_mode=multiple은 UI 계약값"),
    ("request_transaction_account_selection", "output", "account_selection_outcome", "resume.value.account_selection_outcome", "true", "Backend가 검증한 선택 또는 취소 결과를 저장한다.", "selected, cancelled"),
    ("request_transaction_account_selection", "output", "account_ids", "resume.value.account_ids", "false", "선택 완료 시 Backend가 검증한 계좌 ID를 저장한다.", "selected일 때 필수"),
    ("emit_transaction_accounts_empty", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "빈 계좌 목록을 동일한 선택 UI의 빈 상태로 전달한다.", "accounts=[]이며 resume을 기다리지 않음"),
    ("check_transaction_period", "input", "start_date", "", "false", "추출된 조회 시작일의 존재와 해석 결과를 확인한다.", "업무 제한 검증은 하지 않음"),
    ("check_transaction_period", "input", "end_date", "", "false", "추출된 조회 종료일의 존재와 해석 결과를 확인한다.", "업무 제한 검증은 하지 않음"),
    ("check_transaction_period", "output", "start_date", "", "false", "기간 미입력 시 최근 1개월 시작일을 적용한다.", "requested_at과 timezone 기준"),
    ("check_transaction_period", "output", "end_date", "", "false", "기간 미입력 시 조회 기준일을 종료일로 적용한다.", "requested_at과 timezone 기준"),
    ("request_period_selection", "output", "period_selection_outcome", "resume.value.period_selection_outcome", "true", "Backend가 검증한 기간 선택 또는 취소 결과를 저장한다.", "selected, cancelled"),
    ("request_period_selection", "output", "start_date", "resume.value.start_date", "false", "Backend가 정규화한 조회 시작일을 저장한다.", "selected일 때 필수"),
    ("request_period_selection", "output", "end_date", "resume.value.end_date", "false", "Backend가 정규화한 조회 종료일을 저장한다.", "selected일 때 필수"),
    ("query_transactions", "input", "account_ids", "request.account_ids", "true", "검증된 단일·복수 계좌 ID를 거래내역 조회 요청에 전달한다.", "최소 1개, 최대 20개"),
    ("query_transactions", "input", "start_date", "request.start_date", "true", "정규화된 조회 시작일을 전달한다.", "Backend가 최대 기간을 최종 검증"),
    ("query_transactions", "input", "end_date", "request.end_date", "true", "정규화된 조회 종료일을 전달한다.", "Backend가 날짜 순서를 최종 검증"),
    ("query_transactions", "input", "keyword", "request.keyword", "false", "선택적인 거래 검색어를 전달한다.", "최대 100자"),
    ("query_transactions", "input", "transaction_type", "request.transaction_type", "false", "선택적인 TransactionType 필터를 전달한다.", "요청과 응답에서 같은 Enum 사용"),
    ("query_transactions", "output", "transaction_results", "response.data.transaction_results", "true", "Backend가 반환한 첫 페이지 거래내역을 저장한다.", "빈 배열도 정상 결과"),
    ("query_transactions", "output", "transaction_query_id", "response.data.transaction_query_id", "true", "이후 Frontend 페이지 조회용 Query Context를 저장한다.", "Agent는 다음 페이지를 조회하지 않음"),
    ("query_transactions", "output", "next_cursor", "response.data.next_cursor", "false", "다음 페이지 Cursor를 저장한다.", "다음 페이지가 없으면 null"),
    ("emit_transaction_result", "input", "start_date", "webhook.metadata.ui.payload.period.start_date", "true", "조회 시작일을 결과 UI에 전달한다.", ""),
    ("emit_transaction_result", "input", "end_date", "webhook.metadata.ui.payload.period.end_date", "true", "조회 종료일을 결과 UI에 전달한다.", ""),
    ("emit_transaction_result", "input", "account_ids", "webhook.metadata.ui.payload.account_ids", "true", "조회 계좌 범위를 결과 UI에 전달한다.", "검증된 ID만 사용"),
    ("emit_transaction_result", "input", "keyword", "webhook.metadata.ui.payload.keyword", "false", "적용된 검색어를 결과 UI에 전달한다.", ""),
    ("emit_transaction_result", "input", "transaction_results", "webhook.metadata.ui.payload.transactions", "true", "첫 페이지 거래내역을 결과 UI에 전달한다.", "빈 배열이면 동일 UI가 빈 상태 표시"),
    ("emit_transaction_result", "input", "transaction_query_id", "webhook.metadata.ui.payload.transaction_query_id", "true", "Frontend의 이후 페이지 조회 Context를 전달한다.", ""),
    ("emit_transaction_result", "input", "next_cursor", "webhook.metadata.ui.payload.pagination.next_cursor", "false", "다음 페이지 Cursor를 결과 UI에 전달한다.", "null이면 has_more=false"),
]

TRANSACTION_MAPPING_ROWS = [
    {
        "workflow_id": "wf_transaction_history",
        "step_id": step_id,
        "direction": direction,
        "state_key": state_key,
        "contract_field_path": contract_field_path,
        "required_at_step": required_at_step,
        "mapping_description": mapping_description,
        "notes": notes,
        "validation_result": "OK",
    }
    for (
        step_id,
        direction,
        state_key,
        contract_field_path,
        required_at_step,
        mapping_description,
        notes,
    ) in TRANSACTION_MAPPING_DEFINITIONS
]

SUMMARY_MAPPING_DEFINITIONS = [
    ("extract_amount_summary_slots", "output", "account_hint", "", "false", "사용자 발화에서 선택적인 계좌 검색 힌트를 추출한다.", "힌트가 없으면 null"),
    ("extract_amount_summary_slots", "output", "all_accounts_requested", "", "true", "계좌 힌트가 없으면 전체 계좌 집계 의도를 적용한다.", "합계 조회 기본값 true"),
    ("extract_amount_summary_slots", "output", "start_date", "", "false", "사용자 발화의 합계 시작일을 정규화한다.", "기간 미입력 또는 해석 실패이면 null"),
    ("extract_amount_summary_slots", "output", "end_date", "", "false", "사용자 발화의 합계 종료일을 정규화한다.", "기간 미입력 또는 해석 실패이면 null"),
    ("extract_amount_summary_slots", "output", "summary_type", "", "false", "사용자 발화에서 spending 또는 income 목적을 추출한다.", "불명확하면 null"),
    ("extract_amount_summary_slots", "output", "keyword", "", "false", "선택적인 가맹점·상대방 검색어를 추출한다.", "원문은 일반 로그에서 제외"),
    ("resolve_summary_accounts", "input", "account_hint", "query.account_hint", "false", "계좌 검색 힌트를 Backend 계좌 확인 요청에 전달한다.", "resolve_selection=true, account_capability=inquiry는 Tool 설정값"),
    ("resolve_summary_accounts", "input", "all_accounts_requested", "query.all_accounts_requested", "true", "전체 계좌 집계 여부를 Backend 자동 확정 요청에 전달한다.", "힌트가 없으면 true"),
    ("resolve_summary_accounts", "output", "account_resolution_outcome", "response.data.account_resolution_outcome", "true", "Backend의 계좌 자동 확정 결과를 Route 값으로 저장한다.", "후보 개수로 다시 판단하지 않음"),
    ("resolve_summary_accounts", "output", "accounts", "response.data.accounts", "true", "Backend가 검증한 마스킹 계좌 후보를 저장한다.", "no_accounts이면 빈 배열"),
    ("resolve_summary_accounts", "output", "account_ids", "response.data.account_ids", "false", "Backend가 자동 확정한 검증된 계좌 ID를 저장한다.", "resolved일 때 필수"),
    ("request_summary_account_selection", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "검증된 복수 계좌 후보를 선택 UI Payload로 전달한다.", "selection_mode=multiple은 UI 계약값"),
    ("request_summary_account_selection", "output", "account_selection_outcome", "resume.value.account_selection_outcome", "true", "Backend가 검증한 선택 또는 취소 결과를 저장한다.", "selected, cancelled"),
    ("request_summary_account_selection", "output", "account_ids", "resume.value.account_ids", "false", "선택 완료 시 Backend가 검증한 계좌 ID를 저장한다.", "selected일 때 필수"),
    ("emit_summary_accounts_empty", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "빈 계좌 목록을 동일한 선택 UI의 빈 상태로 전달한다.", "accounts=[]이며 resume을 기다리지 않음"),
    ("check_summary_period", "input", "start_date", "", "false", "추출된 합계 시작일의 존재와 해석 결과를 확인한다.", "업무 제한 검증은 하지 않음"),
    ("check_summary_period", "input", "end_date", "", "false", "추출된 합계 종료일의 존재와 해석 결과를 확인한다.", "업무 제한 검증은 하지 않음"),
    ("check_summary_period", "output", "start_date", "", "false", "기간 미입력 시 최근 1개월 시작일을 적용한다.", "requested_at과 timezone 기준"),
    ("check_summary_period", "output", "end_date", "", "false", "기간 미입력 시 조회 기준일을 종료일로 적용한다.", "requested_at과 timezone 기준"),
    ("request_period_selection", "output", "period_selection_outcome", "resume.value.period_selection_outcome", "true", "Backend가 검증한 기간 선택 또는 취소 결과를 저장한다.", "selected, cancelled"),
    ("request_period_selection", "output", "start_date", "resume.value.start_date", "false", "Backend가 정규화한 합계 시작일을 저장한다.", "selected일 때 필수"),
    ("request_period_selection", "output", "end_date", "resume.value.end_date", "false", "Backend가 정규화한 합계 종료일을 저장한다.", "selected일 때 필수"),
    ("check_summary_type", "input", "summary_type", "", "false", "추출된 합계 유형이 확정되었는지 확인한다.", "원장 거래 분류는 수행하지 않음"),
    ("check_summary_type", "output", "summary_type", "", "false", "확정된 spending 또는 income을 유지한다.", "불명확하면 선택 UI로 이동"),
    ("request_summary_type", "output", "summary_type_selection_outcome", "resume.value.summary_type_selection_outcome", "true", "Backend가 검증한 합계 유형 선택 또는 취소 결과를 저장한다.", "selected, cancelled"),
    ("request_summary_type", "output", "summary_type", "resume.value.summary_type", "false", "Backend가 검증한 spending 또는 income을 저장한다.", "selected일 때 필수"),
    ("query_transaction_summary", "input", "account_ids", "request.account_ids", "true", "검증된 단일·복수 계좌 ID를 합계 조회 요청에 전달한다.", "최소 1개, 최대 20개"),
    ("query_transaction_summary", "input", "start_date", "request.start_date", "true", "정규화된 합계 시작일을 전달한다.", "Backend가 최대 기간을 최종 검증"),
    ("query_transaction_summary", "input", "end_date", "request.end_date", "true", "정규화된 합계 종료일을 전달한다.", "Backend가 날짜 순서를 최종 검증"),
    ("query_transaction_summary", "input", "summary_type", "request.summary_type", "true", "검증된 합계 유형을 전달한다.", "spending, income"),
    ("query_transaction_summary", "input", "keyword", "request.keyword", "false", "선택적인 가맹점·상대방 검색어를 전달한다.", "최대 100자"),
    ("query_transaction_summary", "output", "summary_result", "response.data.summary_result", "true", "Backend가 계산한 합계 결과를 저장한다.", "total_amount=0도 정상 결과"),
    ("emit_amount_summary", "input", "account_ids", "webhook.metadata.ui.payload.account_ids", "true", "집계 계좌 범위를 결과 UI에 전달한다.", "검증된 ID만 사용"),
    ("emit_amount_summary", "input", "keyword", "webhook.metadata.ui.payload.keyword", "false", "적용된 검색어를 결과 UI에 전달한다.", ""),
    ("emit_amount_summary", "input", "summary_result", "webhook.metadata.ui.payload.summary", "true", "Backend 합계 결과를 결과 UI에 전달한다.", "금액 원문은 일반 로그에서 제외"),
]

SUMMARY_MAPPING_ROWS = [
    {
        "workflow_id": "wf_period_amount_summary",
        "step_id": step_id,
        "direction": direction,
        "state_key": state_key,
        "contract_field_path": contract_field_path,
        "required_at_step": required_at_step,
        "mapping_description": mapping_description,
        "notes": notes,
        "validation_result": "OK",
    }
    for (
        step_id,
        direction,
        state_key,
        contract_field_path,
        required_at_step,
        mapping_description,
        notes,
    ) in SUMMARY_MAPPING_DEFINITIONS
]

DEFAULT_ACCOUNT_MAPPING_DEFINITIONS = [
    ("extract_default_account_slots", "output", "account_hint", "", "false", "사용자 발화에서 선택적인 기본계좌 대상 힌트를 추출한다.", "힌트가 없으면 null"),
    ("resolve_default_account", "input", "account_hint", "query.account_hint", "false", "계좌 힌트를 Backend 계좌 확인 요청에 전달한다.", "reset 이후에는 null로 전체 최신 후보를 요청"),
    ("resolve_default_account", "output", "account_resolution_outcome", "response.data.account_resolution_outcome", "true", "Backend의 계좌 자동 확정 결과를 Route 값으로 저장한다.", "resolved, selection_required, no_accounts"),
    ("resolve_default_account", "output", "accounts", "response.data.accounts", "true", "Backend가 검증한 최신 마스킹 계좌 후보를 저장한다.", "no_accounts이면 빈 배열"),
    ("resolve_default_account", "output", "account_id", "response.data.account_ids[0]", "false", "resolved 응답의 단일 계좌 ID를 Workflow State에 저장한다.", "account_ids 길이가 정확히 1일 때만 저장"),
    ("request_default_account_selection", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "검증된 최신 계좌 후보를 단일 선택 UI에 전달한다.", "selection_mode를 Agent API 필드로 추가하지 않음"),
    ("request_default_account_selection", "input", "correction_view", "webhook.metadata.ui.payload.correction_view", "false", "Backend가 반환한 수정 사유를 계좌 선택 UI에 전달한다.", "최초 선택에서는 null"),
    ("request_default_account_selection", "output", "account_selection_outcome", "resume.value.account_selection_outcome", "true", "Backend가 검증한 선택 또는 취소 결과를 저장한다.", "selected, cancelled"),
    ("request_default_account_selection", "output", "account_id", "resume.value.account_ids[0]", "false", "선택 완료 시 단일 계좌 ID를 저장한다.", "account_ids 길이가 정확히 1일 때만 저장"),
    ("emit_default_account_selection_empty", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "빈 계좌 목록을 동일한 선택 UI의 빈 상태로 전달한다.", "accounts=[]이며 resume을 기다리지 않음"),
    ("start_default_account_prepare", "output", "prepare_attempt", "agent.state.prepare_attempt", "true", "새로운 논리 Prepare의 시도 번호를 증가시켜 저장한다.", "통신 재시도에서는 증가시키지 않음"),
    ("start_default_account_prepare", "output", "correction_view", "agent.state.correction_view=null", "false", "새로운 Prepare를 시작하기 전에 이전 수정 사유를 제거한다.", ""),
    ("prepare_default_account_change", "input", "account_id", "request.account_id", "true", "검증된 대상 계좌 ID를 Prepare 요청에 전달한다.", "현재 기본계좌는 Backend가 확인"),
    ("prepare_default_account_change", "input", "prepare_attempt", "header.Idempotency-Key", "true", "실행 Context와 시도 번호로 Prepare 멱등성 키를 생성한다.", "default_account_prepare:{execution_context_id}:{prepare_attempt}"),
    ("prepare_default_account_change", "output", "confirmation_id", "response.data.confirmation_id", "false", "승인 준비 완료 시 Confirmation 식별자를 저장한다.", "ready_for_confirmation일 때 필수"),
    ("prepare_default_account_change", "output", "confirmation_view", "response.data.confirmation_view", "false", "승인 화면용 변경 전후 마스킹 데이터를 임시 저장한다.", "ready_for_confirmation일 때 필수"),
    ("prepare_default_account_change", "output", "correction_view", "response.data.correction_view", "false", "수정 필요 시 허용된 수정 대상을 저장한다.", "correction_required일 때 필수"),
    ("request_default_account_approval", "input", "confirmation_id", "webhook.interaction.confirmation_id", "true", "승인 요청과 Backend resume을 연결할 Confirmation 식별자를 전달한다.", ""),
    ("request_default_account_approval", "input", "confirmation_view", "webhook.metadata.ui.payload", "true", "Backend가 생성한 변경 전후 표시 데이터를 승인 UI에 전달한다.", "Agent가 원장 정보를 추가하지 않음"),
    ("request_default_account_approval", "output", "approval_outcome", "resume.value.approval_outcome", "true", "Backend가 검증하고 상태 변경까지 완료한 승인 결과를 저장한다.", "approved, change_requested, cancelled"),
    ("reset_default_account_target", "output", "account_hint", "agent.state.account_hint=null", "true", "기존 계좌 힌트를 제거하여 최신 전체 후보를 다시 조회한다.", ""),
    ("reset_default_account_target", "output", "account_resolution_outcome", "agent.state.account_resolution_outcome=null", "true", "이전 자동 확정 결과를 제거한다.", ""),
    ("reset_default_account_target", "output", "accounts", "agent.state.accounts=[]", "true", "오래된 계좌 후보를 제거한다.", "다음 API 응답으로 교체"),
    ("reset_default_account_target", "output", "account_id", "agent.state.account_id=null", "true", "기존 변경 대상 계좌를 제거한다.", ""),
    ("reset_default_account_target", "output", "account_selection_outcome", "agent.state.account_selection_outcome=null", "true", "이전 선택 결과를 제거한다.", ""),
    ("reset_default_account_target", "output", "input_request_id", "agent.state.input_request_id=null", "true", "이전 입력 요청 식별자를 제거한다.", ""),
    ("reset_default_account_target", "output", "confirmation_id", "agent.state.confirmation_id=null", "true", "무효화되거나 실행 불가능한 Confirmation을 제거한다.", ""),
    ("reset_default_account_target", "output", "confirmation_view", "agent.state.confirmation_view=null", "true", "이전 승인 화면 데이터를 제거한다.", ""),
    ("reset_default_account_target", "output", "approval_outcome", "agent.state.approval_outcome=null", "true", "처리한 승인 결과를 제거한다.", ""),
    ("execute_default_account_change", "input", "confirmation_id", "request.confirmation_id", "true", "승인된 Confirmation 식별자만 Execute 요청에 전달한다.", "account_id를 다시 전달하지 않음"),
    ("execute_default_account_change", "output", "account_id", "response.data.account_id", "false", "Backend가 실제 반영한 최종 기본계좌 ID를 저장한다.", "completed일 때 필수"),
    ("execute_default_account_change", "output", "completed_at", "response.data.completed_at", "false", "기본계좌 변경 완료 시각을 저장한다.", "completed일 때 필수"),
    ("execute_default_account_change", "output", "correction_view", "response.data.correction_view", "false", "실행 시점 수정 필요 결과를 저장한다.", "correction_required일 때 필수"),
    ("emit_default_account_unchanged", "input", "account_id", "webhook.metadata.ui.payload.account_id", "true", "이미 기본계좌인 대상 ID를 setting_result에 전달한다.", "outcome=unchanged는 UI 계약 상수"),
    ("emit_default_account_result", "input", "account_id", "webhook.metadata.ui.payload.account_id", "true", "실제로 반영된 기본계좌 ID를 결과 UI에 전달한다.", "outcome=completed는 UI 계약 상수"),
    ("emit_default_account_result", "input", "completed_at", "webhook.metadata.ui.payload.completed_at", "true", "변경 완료 시각을 결과 UI에 전달한다.", ""),
]

DEFAULT_ACCOUNT_MAPPING_ROWS = [
    {
        "workflow_id": "wf_set_default_account",
        "step_id": step_id,
        "direction": direction,
        "state_key": state_key,
        "contract_field_path": contract_field_path,
        "required_at_step": required_at_step,
        "mapping_description": mapping_description,
        "notes": notes,
        "validation_result": "OK",
    }
    for (
        step_id,
        direction,
        state_key,
        contract_field_path,
        required_at_step,
        mapping_description,
        notes,
    ) in DEFAULT_ACCOUNT_MAPPING_DEFINITIONS
]

ACCOUNT_ALIAS_MAPPING_DEFINITIONS = [
    ("extract_account_alias_slots", "output", "account_hint", "", "false", "사용자 발화에서 대상 계좌 힌트를 추출한다.", "힌트가 없으면 null"),
    ("extract_account_alias_slots", "output", "alias", "", "false", "사용자 발화에서 선택적 새 별칭을 추출한다.", "없으면 입력 UI로 이동"),
    ("resolve_account_alias_target", "input", "account_hint", "query.account_hint", "false", "계좌 힌트를 Backend 계좌 확인 요청에 전달한다.", "reset 이후에는 null로 최신 전체 후보 요청"),
    ("resolve_account_alias_target", "output", "account_resolution_outcome", "response.data.account_resolution_outcome", "true", "Backend 계좌 자동 확정 결과를 저장한다.", "resolved, selection_required, no_accounts"),
    ("resolve_account_alias_target", "output", "accounts", "response.data.accounts", "true", "Backend가 검증한 최신 마스킹 계좌 후보를 저장한다.", "no_accounts이면 빈 배열"),
    ("resolve_account_alias_target", "output", "account_id", "response.data.account_ids[0]", "false", "resolved 응답의 단일 계좌 ID를 저장한다.", "배열 길이가 정확히 1일 때만 저장"),
    ("request_account_alias_selection", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "검증된 최신 계좌 후보를 선택 UI에 전달한다.", ""),
    ("request_account_alias_selection", "input", "correction_view", "webhook.metadata.ui.payload.correction_view", "false", "계좌 수정 사유를 선택 UI에 전달한다.", "최초 선택에서는 null"),
    ("request_account_alias_selection", "output", "account_selection_outcome", "resume.value.account_selection_outcome", "true", "Backend가 검증한 선택 또는 취소 결과를 저장한다.", "selected, cancelled"),
    ("request_account_alias_selection", "output", "account_id", "resume.value.account_ids[0]", "false", "선택 완료 시 단일 계좌 ID를 저장한다.", "배열 길이가 정확히 1일 때만 저장"),
    ("emit_account_alias_selection_empty", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "빈 계좌 목록을 동일한 선택 UI의 빈 상태로 전달한다.", "accounts=[]"),
    ("check_account_alias_value", "input", "alias", "", "false", "새 별칭이 State에 존재하는지 확인한다.", "정책 검증은 수행하지 않음"),
    ("request_account_alias_input", "input", "correction_view", "webhook.metadata.ui.payload.correction_view", "false", "별칭 수정 사유를 입력 UI에 전달한다.", "최초 입력에서는 null"),
    ("request_account_alias_input", "output", "alias_input_outcome", "resume.value.alias_input_outcome", "true", "Backend가 검증한 제출 또는 취소 결과를 저장한다.", "submitted, cancelled"),
    ("request_account_alias_input", "output", "alias", "resume.value.alias", "false", "Backend가 정규화한 별칭을 저장한다.", "submitted일 때 필수"),
    ("start_account_alias_prepare", "output", "prepare_attempt", "agent.state.prepare_attempt", "true", "새로운 논리 Prepare 시도 번호를 증가시킨다.", "통신 재시도에서는 증가시키지 않음"),
    ("start_account_alias_prepare", "output", "correction_view", "agent.state.correction_view=null", "false", "새 Prepare 전에 이전 수정 사유를 제거한다.", ""),
    ("prepare_account_alias_change", "input", "account_id", "request.account_id", "true", "검증된 대상 계좌 ID를 Prepare에 전달한다.", ""),
    ("prepare_account_alias_change", "input", "alias", "request.alias", "true", "검증·정규화된 새 별칭을 Prepare에 전달한다.", ""),
    ("prepare_account_alias_change", "input", "prepare_attempt", "header.Idempotency-Key", "true", "실행 Context와 시도 번호로 멱등성 키를 생성한다.", "account_alias_prepare:{execution_context_id}:{prepare_attempt}"),
    ("prepare_account_alias_change", "output", "confirmation_id", "response.data.confirmation_id", "false", "승인 준비 완료 시 Confirmation ID를 저장한다.", "ready_for_confirmation일 때 필수"),
    ("prepare_account_alias_change", "output", "confirmation_view", "response.data.confirmation_view", "false", "승인 화면용 계좌와 최종 별칭을 임시 저장한다.", "ready_for_confirmation일 때 필수"),
    ("prepare_account_alias_change", "output", "correction_view", "response.data.correction_view", "false", "수정 필요 시 단일 허용 대상을 저장한다.", "correction_required일 때 필수"),
    ("request_account_alias_approval", "input", "confirmation_id", "webhook.interaction.confirmation_id", "true", "승인 요청과 resume을 연결할 ID를 전달한다.", ""),
    ("request_account_alias_approval", "input", "confirmation_view", "webhook.metadata.ui.payload", "true", "Backend가 생성한 승인 표시 데이터를 전달한다.", "account 객체와 alias만 포함"),
    ("request_account_alias_approval", "output", "approval_outcome", "resume.value.approval_outcome", "true", "Backend가 검증한 승인 결과를 저장한다.", "approved, change_requested, cancelled"),
    ("request_account_alias_approval", "output", "change_target", "resume.value.change_target", "false", "수정 요청의 대상을 저장한다.", "change_requested일 때 account 또는 alias"),
    ("reset_account_alias_target", "output", "account_hint", "agent.state.account_hint=null", "true", "기존 계좌 힌트를 제거한다.", ""),
    ("reset_account_alias_target", "output", "account_resolution_outcome", "agent.state.account_resolution_outcome=null", "true", "이전 자동 확정 결과를 제거한다.", ""),
    ("reset_account_alias_target", "output", "accounts", "agent.state.accounts=[]", "true", "오래된 계좌 후보를 제거한다.", ""),
    ("reset_account_alias_target", "output", "account_id", "agent.state.account_id=null", "true", "기존 대상 계좌 ID를 제거한다.", ""),
    ("reset_account_alias_target", "output", "account_selection_outcome", "agent.state.account_selection_outcome=null", "true", "이전 계좌 선택 결과를 제거한다.", ""),
    ("reset_account_alias_target", "output", "alias_input_outcome", "agent.state.alias_input_outcome=null", "true", "이전 별칭 입력 결과를 제거한다.", "alias 값은 유지"),
    ("reset_account_alias_target", "output", "input_request_id", "agent.state.input_request_id=null", "true", "이전 입력 요청 ID를 제거한다.", ""),
    ("reset_account_alias_target", "output", "confirmation_id", "agent.state.confirmation_id=null", "true", "무효화된 Confirmation을 제거한다.", ""),
    ("reset_account_alias_target", "output", "confirmation_view", "agent.state.confirmation_view=null", "true", "이전 승인 표시 데이터를 제거한다.", ""),
    ("reset_account_alias_target", "output", "approval_outcome", "agent.state.approval_outcome=null", "true", "처리한 승인 결과를 제거한다.", ""),
    ("reset_account_alias_target", "output", "change_target", "agent.state.change_target=null", "true", "처리한 수정 대상을 제거한다.", "correction_view는 유지"),
    ("reset_account_alias_value", "output", "alias", "agent.state.alias=null", "true", "기존 새 별칭 값을 제거한다.", ""),
    ("reset_account_alias_value", "output", "alias_input_outcome", "agent.state.alias_input_outcome=null", "true", "이전 별칭 입력 결과를 제거한다.", ""),
    ("reset_account_alias_value", "output", "input_request_id", "agent.state.input_request_id=null", "true", "이전 입력 요청 ID를 제거한다.", ""),
    ("reset_account_alias_value", "output", "confirmation_id", "agent.state.confirmation_id=null", "true", "무효화된 Confirmation을 제거한다.", ""),
    ("reset_account_alias_value", "output", "confirmation_view", "agent.state.confirmation_view=null", "true", "이전 승인 표시 데이터를 제거한다.", ""),
    ("reset_account_alias_value", "output", "approval_outcome", "agent.state.approval_outcome=null", "true", "처리한 승인 결과를 제거한다.", ""),
    ("reset_account_alias_value", "output", "change_target", "agent.state.change_target=null", "true", "처리한 수정 대상을 제거한다.", "correction_view는 유지"),
    ("execute_account_alias_change", "input", "confirmation_id", "request.confirmation_id", "true", "승인된 Confirmation ID만 Execute에 전달한다.", "account_id와 alias를 다시 전달하지 않음"),
    ("execute_account_alias_change", "output", "account_id", "response.data.account_id", "false", "Backend가 실제 반영한 계좌 ID를 저장한다.", "completed일 때 필수"),
    ("execute_account_alias_change", "output", "alias", "response.data.alias", "false", "Backend가 실제 반영한 별칭을 저장한다.", "completed일 때 필수"),
    ("execute_account_alias_change", "output", "completed_at", "response.data.completed_at", "false", "별칭 변경 완료 시각을 저장한다.", "completed일 때 필수"),
    ("execute_account_alias_change", "output", "correction_view", "response.data.correction_view", "false", "실행 시점 단일 수정 대상을 저장한다.", "correction_required일 때 필수"),
    ("emit_account_alias_unchanged", "input", "account_id", "webhook.metadata.ui.payload.account_id", "true", "이미 같은 별칭인 계좌 ID를 결과 UI에 전달한다.", "outcome=unchanged"),
    ("emit_account_alias_unchanged", "input", "alias", "webhook.metadata.ui.payload.alias", "true", "이미 적용된 별칭을 결과 UI에 전달한다.", ""),
    ("emit_account_alias_result", "input", "account_id", "webhook.metadata.ui.payload.account_id", "true", "실제로 반영된 계좌 ID를 결과 UI에 전달한다.", "outcome=completed"),
    ("emit_account_alias_result", "input", "alias", "webhook.metadata.ui.payload.alias", "true", "실제로 반영된 별칭을 결과 UI에 전달한다.", ""),
    ("emit_account_alias_result", "input", "completed_at", "webhook.metadata.ui.payload.completed_at", "true", "별칭 변경 완료 시각을 결과 UI에 전달한다.", ""),
]

ACCOUNT_ALIAS_MAPPING_ROWS = [
    {
        "workflow_id": "wf_set_account_alias",
        "step_id": step_id,
        "direction": direction,
        "state_key": state_key,
        "contract_field_path": contract_field_path,
        "required_at_step": required_at_step,
        "mapping_description": mapping_description,
        "notes": notes,
        "validation_result": "OK",
    }
    for (
        step_id,
        direction,
        state_key,
        contract_field_path,
        required_at_step,
        mapping_description,
        notes,
    ) in ACCOUNT_ALIAS_MAPPING_DEFINITIONS
]

INTERNAL_TRANSFER_MAPPING_DEFINITIONS = [
    ("extract_internal_transfer_slots", "output", "from_account_hint", "", "false", "사용자 발화에서 선택적 출금 계좌 힌트를 추출한다.", "없으면 null"),
    ("extract_internal_transfer_slots", "output", "to_account_hint", "", "false", "사용자 발화에서 선택적 입금 계좌 힌트를 추출한다.", "없으면 null"),
    ("extract_internal_transfer_slots", "output", "amount", "", "false", "사용자 발화에서 선택적 이체 금액을 추출한다.", "없으면 금액 입력 UI로 이동"),
    ("resolve_internal_from_account", "input", "from_account_hint", "query.account_hint", "false", "출금 계좌 힌트를 Backend 계좌 확인 요청에 전달한다.", ""),
    ("resolve_internal_from_account", "input", "to_account_id", "query.exclude_account_ids[0]", "false", "유지한 입금 계좌를 출금 후보에서 제외한다.", "출금 계좌 재선택에서 사용"),
    ("resolve_internal_from_account", "output", "account_resolution_outcome", "response.data.account_resolution_outcome", "true", "Backend 출금 계좌 자동 확정 결과를 저장한다.", ""),
    ("resolve_internal_from_account", "output", "accounts", "response.data.accounts", "true", "Backend가 검증한 최신 출금 계좌 후보를 저장한다.", ""),
    ("resolve_internal_from_account", "output", "from_account_id", "response.data.account_ids[0]", "false", "자동 확정된 단일 출금 계좌 ID를 저장한다.", "배열 길이가 정확히 1일 때만 저장"),
    ("request_from_account_selection", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "출금 계좌 후보를 선택 UI에 전달한다.", ""),
    ("request_from_account_selection", "input", "correction_view", "webhook.metadata.ui.payload.correction_view", "false", "출금 계좌 수정 사유를 선택 UI에 전달한다.", "최초 선택에서는 null"),
    ("request_from_account_selection", "output", "account_selection_outcome", "resume.value.account_selection_outcome", "true", "Backend가 검증한 선택 또는 취소 결과를 저장한다.", ""),
    ("request_from_account_selection", "output", "from_account_id", "resume.value.account_ids[0]", "false", "선택한 단일 출금 계좌 ID를 저장한다.", ""),
    ("emit_internal_from_accounts_empty", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "빈 출금 계좌 목록을 동일한 선택 UI로 전달한다.", "accounts=[]"),
    ("resolve_internal_to_account", "input", "to_account_hint", "query.account_hint", "false", "입금 계좌 힌트를 Backend 계좌 확인 요청에 전달한다.", ""),
    ("resolve_internal_to_account", "input", "from_account_id", "query.exclude_account_ids[0]", "true", "확정된 출금 계좌를 입금 후보에서 제외한다.", ""),
    ("resolve_internal_to_account", "output", "account_resolution_outcome", "response.data.account_resolution_outcome", "true", "Backend 입금 계좌 자동 확정 결과를 저장한다.", ""),
    ("resolve_internal_to_account", "output", "accounts", "response.data.accounts", "true", "Backend가 검증한 최신 입금 계좌 후보를 저장한다.", ""),
    ("resolve_internal_to_account", "output", "to_account_id", "response.data.account_ids[0]", "false", "자동 확정된 단일 입금 계좌 ID를 저장한다.", "배열 길이가 정확히 1일 때만 저장"),
    ("request_to_account_selection", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "입금 계좌 후보를 선택 UI에 전달한다.", ""),
    ("request_to_account_selection", "input", "correction_view", "webhook.metadata.ui.payload.correction_view", "false", "입금 계좌 수정 사유를 선택 UI에 전달한다.", "최초 선택에서는 null"),
    ("request_to_account_selection", "output", "account_selection_outcome", "resume.value.account_selection_outcome", "true", "Backend가 검증한 선택 또는 취소 결과를 저장한다.", ""),
    ("request_to_account_selection", "output", "to_account_id", "resume.value.account_ids[0]", "false", "선택한 단일 입금 계좌 ID를 저장한다.", ""),
    ("emit_internal_to_accounts_empty", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "빈 입금 계좌 목록을 동일한 선택 UI로 전달한다.", "accounts=[]"),
    ("check_internal_transfer_amount", "input", "amount", "agent.state.amount", "false", "금액의 존재, 정수 여부와 양수 여부만 확인한다.", "금융 검증은 수행하지 않음"),
    ("request_internal_transfer_amount", "input", "correction_view", "webhook.metadata.ui.payload.correction_view", "false", "금액 수정 사유를 입력 UI에 전달한다.", "최초 입력에서는 null"),
    ("request_internal_transfer_amount", "input", "currency", "webhook.metadata.ui.payload.currency", "true", "금액 입력 통화를 UI에 전달한다.", "KRW"),
    ("request_internal_transfer_amount", "output", "amount_input_outcome", "resume.value.amount_input_outcome", "true", "Backend가 검증한 제출 또는 취소 결과를 저장한다.", "submitted, cancelled"),
    ("request_internal_transfer_amount", "output", "amount", "resume.value.amount", "false", "Backend가 검증한 이체 금액을 저장한다.", "submitted일 때 필수"),
    ("start_internal_transfer_prepare", "output", "prepare_attempt", "agent.state.prepare_attempt", "true", "새로운 논리 Prepare 시도 번호를 증가시킨다.", "통신 재시도에서는 유지"),
    ("start_internal_transfer_prepare", "output", "correction_view", "agent.state.correction_view=null", "false", "새 Prepare 전에 이전 수정 사유를 제거한다.", ""),
    ("prepare_internal_transfer", "input", "from_account_id", "request.from_account_id", "true", "검증된 출금 계좌 ID를 Prepare에 전달한다.", ""),
    ("prepare_internal_transfer", "input", "to_account_id", "request.to_account_id", "true", "검증된 입금 계좌 ID를 Prepare에 전달한다.", ""),
    ("prepare_internal_transfer", "input", "amount", "request.amount", "true", "이체 금액을 Prepare에 전달한다.", ""),
    ("prepare_internal_transfer", "input", "currency", "request.currency", "true", "이체 통화를 Prepare에 전달한다.", ""),
    ("prepare_internal_transfer", "input", "prepare_attempt", "header.Idempotency-Key", "true", "실행 Context와 시도 번호로 Prepare 멱등성 키를 생성한다.", "internal_transfer_prepare:{execution_context_id}:{prepare_attempt}"),
    ("prepare_internal_transfer", "output", "confirmation_id", "response.data.confirmation_id", "false", "승인 준비 완료 시 Confirmation ID를 저장한다.", ""),
    ("prepare_internal_transfer", "output", "confirmation_view", "response.data.confirmation_view", "false", "Backend 승인 표시 데이터를 저장한다.", "완료 결과까지 재사용"),
    ("prepare_internal_transfer", "output", "correction_view", "response.data.correction_view", "false", "수정 가능한 대상과 표시 사유를 저장한다.", ""),
    ("prepare_internal_transfer", "output", "blocked_view", "response.data.blocked_view", "false", "사용자 표시용 차단 안내를 저장한다.", ""),
    ("request_internal_transfer_approval", "input", "confirmation_id", "webhook.interaction.confirmation_id", "true", "승인 요청과 재개를 연결할 Confirmation ID를 전달한다.", ""),
    ("request_internal_transfer_approval", "input", "confirmation_view", "webhook.metadata.ui.payload", "true", "Backend가 생성한 승인 표시 데이터를 그대로 전달한다.", ""),
    ("request_internal_transfer_approval", "output", "approval_outcome", "resume.value.approval_outcome", "true", "Backend가 검증한 승인 결과를 저장한다.", ""),
    ("request_internal_transfer_approval", "output", "change_target", "resume.value.change_target", "false", "수정 요청 대상을 저장한다.", "change_requested일 때 필수"),
    ("route_internal_transfer_correction", "input", "correction_view", "agent.state.correction_view", "true", "허용 수정 대상의 개수와 값을 구조적으로 확인한다.", ""),
    ("request_internal_transfer_correction", "input", "correction_view", "webhook.metadata.ui.payload", "true", "Backend 수정 표시 데이터를 선택 UI에 전달한다.", ""),
    ("request_internal_transfer_correction", "output", "correction_selection_outcome", "resume.value.correction_selection_outcome", "true", "수정 대상 선택 또는 취소 결과를 저장한다.", ""),
    ("request_internal_transfer_correction", "output", "change_target", "resume.value.change_target", "false", "선택한 수정 대상을 저장한다.", "selected일 때 필수"),
    ("start_internal_auth", "output", "auth_context_id", "agent.state.auth_context_id=null", "false", "기존 Auth Context ID를 제거한다.", ""),
    ("start_internal_auth", "output", "auth_request_view", "agent.state.auth_request_view=null", "false", "기존 인증 표시 데이터를 제거한다.", ""),
    ("start_internal_auth", "output", "auth_status", "agent.state.auth_status=null", "false", "기존 인증 결과를 제거한다.", ""),
    ("start_internal_auth", "output", "auth_retry_outcome", "agent.state.auth_retry_outcome=null", "false", "기존 재인증 선택 결과를 제거한다.", ""),
    ("start_internal_auth", "output", "auth_attempt", "agent.state.auth_attempt", "true", "새로운 인증 시도 번호를 증가시킨다.", "통신 재시도에서는 유지"),
    ("create_internal_auth_context", "input", "confirmation_id", "request.confirmation_id", "true", "승인된 Confirmation ID를 인증 Context 생성 요청에 전달한다.", ""),
    ("create_internal_auth_context", "input", "auth_attempt", "header.Idempotency-Key", "true", "Confirmation과 인증 시도로 멱등성 키를 생성한다.", "internal_transfer_auth:{confirmation_id}:{auth_attempt}"),
    ("create_internal_auth_context", "output", "auth_context_id", "response.data.auth_context_id", "false", "Backend가 발급한 Auth Context ID를 저장한다.", ""),
    ("create_internal_auth_context", "output", "auth_request_view", "response.data.auth_request_view", "false", "Backend 인증 표시 데이터를 저장한다.", ""),
    ("create_internal_auth_context", "output", "blocked_view", "response.data.blocked_view", "false", "인증 진행 차단 안내를 저장한다.", ""),
    ("request_internal_authentication", "input", "auth_context_id", "webhook.interaction.auth_context_id", "true", "인증 요청과 재개를 연결할 Auth Context ID를 전달한다.", ""),
    ("request_internal_authentication", "input", "auth_request_view", "webhook.metadata.ui.payload", "true", "Backend 인증 표시 데이터를 그대로 전달한다.", ""),
    ("request_internal_authentication", "output", "auth_status", "resume.value.auth_status", "true", "Backend가 검증한 인증 결과를 저장한다.", ""),
    ("request_internal_auth_retry", "input", "auth_request_view", "webhook.metadata.ui.payload.auth_request_view", "false", "만료·실패한 인증 안내를 재인증 선택 UI에 전달한다.", ""),
    ("request_internal_auth_retry", "output", "auth_retry_outcome", "resume.value.auth_retry_outcome", "true", "재인증 또는 취소 선택 결과를 저장한다.", "retry, cancelled"),
    ("execute_internal_transfer", "input", "confirmation_id", "request.confirmation_id", "true", "승인된 Confirmation ID를 Execute에 전달한다.", "송금 조건을 다시 보내지 않음"),
    ("execute_internal_transfer", "input", "auth_context_id", "request.auth_context_id", "true", "검증된 Auth Context ID를 Execute에 전달한다.", ""),
    ("execute_internal_transfer", "input", "auth_attempt", "header.Idempotency-Key", "true", "Confirmation과 인증 시도로 Execute 멱등성 키를 생성한다.", "internal_transfer_execute:{confirmation_id}:{auth_attempt}"),
    ("execute_internal_transfer", "output", "transaction_id", "response.data.transaction_id", "false", "완료된 거래 ID를 저장한다.", "completed일 때 필수"),
    ("execute_internal_transfer", "output", "completed_at", "response.data.completed_at", "false", "송금 완료 시각을 저장한다.", "completed일 때 필수"),
    ("execute_internal_transfer", "output", "correction_view", "response.data.correction_view", "false", "실행 시점 수정 가능 대상과 사유를 저장한다.", ""),
    ("execute_internal_transfer", "output", "blocked_view", "response.data.blocked_view", "false", "실행 차단 안내를 저장한다.", ""),
    ("emit_internal_transfer_result", "input", "transaction_id", "webhook.metadata.ui.payload.transaction_id", "true", "거래 ID를 완료 UI에 전달한다.", ""),
    ("emit_internal_transfer_result", "input", "completed_at", "webhook.metadata.ui.payload.completed_at", "true", "완료 시각을 완료 UI에 전달한다.", ""),
    ("emit_internal_transfer_result", "input", "confirmation_view", "webhook.metadata.ui.payload.from_account", "true", "Prepare의 출금 계좌 표시정보를 완료 UI에 재사용한다.", "confirmation_view.from_account 투영"),
    ("emit_internal_transfer_result", "input", "confirmation_view", "webhook.metadata.ui.payload.to_account", "true", "Prepare의 입금 계좌 표시정보를 완료 UI에 재사용한다.", "confirmation_view.to_account 투영"),
    ("emit_internal_transfer_result", "input", "confirmation_view", "webhook.metadata.ui.payload.amount", "true", "Prepare의 금액을 완료 UI에 재사용한다.", "confirmation_view.amount 투영"),
    ("emit_internal_transfer_result", "input", "confirmation_view", "webhook.metadata.ui.payload.currency", "true", "Prepare의 통화를 완료 UI에 재사용한다.", "confirmation_view.currency 투영"),
    ("emit_internal_transfer_blocked", "input", "blocked_view", "webhook.metadata.ui.payload", "true", "Backend 사용자 표시용 차단 안내를 그대로 전달한다.", ""),
]

INTERNAL_TRANSFER_RESET_FIELDS = {
    "reset_internal_from_account": [
        "from_account_hint", "from_account_id", "account_resolution_outcome", "accounts",
        "account_selection_outcome", "amount_input_outcome", "input_request_id", "confirmation_id",
        "confirmation_view", "approval_outcome", "change_target", "correction_selection_outcome",
        "blocked_view", "auth_context_id", "auth_request_view", "auth_status", "auth_retry_outcome", "auth_attempt",
    ],
    "reset_internal_to_account": [
        "to_account_hint", "to_account_id", "account_resolution_outcome", "accounts",
        "account_selection_outcome", "amount_input_outcome", "input_request_id", "confirmation_id",
        "confirmation_view", "approval_outcome", "change_target", "correction_selection_outcome",
        "blocked_view", "auth_context_id", "auth_request_view", "auth_status", "auth_retry_outcome", "auth_attempt",
    ],
    "reset_internal_transfer_amount": [
        "amount", "amount_input_outcome", "input_request_id", "confirmation_id", "confirmation_view",
        "approval_outcome", "change_target", "correction_selection_outcome", "blocked_view",
        "auth_context_id", "auth_request_view", "auth_status", "auth_retry_outcome", "auth_attempt",
    ],
}


def internal_reset_mapping_definitions() -> list[tuple[str, str, str, str, str, str, str]]:
    definitions: list[tuple[str, str, str, str, str, str, str]] = []
    for step_id, state_keys in INTERNAL_TRANSFER_RESET_FIELDS.items():
        for state_key in state_keys:
            cleared_value = "[]" if state_key == "accounts" else "0" if state_key == "auth_attempt" else "null"
            definitions.append(
                (
                    step_id,
                    "output",
                    state_key,
                    f"agent.state.{state_key}={cleared_value}",
                    "true",
                    "수정 전에 오래된 입력·승인·인증 임시 State를 제거한다.",
                    "correction_view와 prepare_attempt는 유지",
                )
            )
    return definitions


INTERNAL_TRANSFER_MAPPING_DEFINITIONS.extend(internal_reset_mapping_definitions())

INTERNAL_TRANSFER_MAPPING_ROWS = [
    {
        "workflow_id": "wf_internal_transfer",
        "step_id": step_id,
        "direction": direction,
        "state_key": state_key,
        "contract_field_path": contract_field_path,
        "required_at_step": required_at_step,
        "mapping_description": mapping_description,
        "notes": notes,
        "validation_result": "OK",
    }
    for (
        step_id,
        direction,
        state_key,
        contract_field_path,
        required_at_step,
        mapping_description,
        notes,
    ) in INTERNAL_TRANSFER_MAPPING_DEFINITIONS
]

EXTERNAL_TRANSFER_MAPPING_DEFINITIONS = [
    ("extract_external_transfer_slots", "output", "from_account_hint", "", "false", "선택적 출금 계좌 힌트를 추출한다.", ""),
    ("extract_external_transfer_slots", "output", "recipient_name_hint", "", "false", "선택적 수취인 이름 힌트를 추출한다.", ""),
    ("extract_external_transfer_slots", "output", "amount", "", "false", "선택적 송금 금액을 추출한다.", ""),
    ("resolve_recipient_hint", "input", "recipient_name_hint", "request.recipient_name_hint", "true", "수취인 이름 힌트를 기존 거래 자동 확정 API에 전달한다.", ""),
    ("resolve_recipient_hint", "output", "recipient_resolution_outcome", "response.data.outcome", "true", "수취인 자동 확정 결과를 저장한다.", "resolved, selection_required"),
    ("resolve_recipient_hint", "output", "recipient_selection_reason", "response.data.selection_reason", "false", "선택이 필요한 사유를 저장한다.", "multiple_matches, no_match"),
    ("resolve_recipient_hint", "output", "to_recipient_id", "response.data.to_recipient_id", "false", "자동 확정된 기존 수취인 ID를 저장한다.", "resolved일 때 필수"),
    ("request_recipient_selection", "input", "recipient_name_hint", "webhook.metadata.ui.payload.recipient_name_hint", "false", "최초 발화의 이름 힌트를 수취인 UI 요청에 전달한다.", ""),
    ("request_recipient_selection", "input", "recipient_selection_reason", "webhook.metadata.ui.payload.recipient_selection_reason", "false", "이름 후보 또는 초기 화면 구성을 위한 사유를 전달한다.", ""),
    ("request_recipient_selection", "input", "correction_view", "webhook.metadata.ui.payload.correction_view", "false", "수취인 수정 사유를 전달한다.", ""),
    ("request_recipient_selection", "output", "recipient_selection_outcome", "resume.value.recipient_selection_outcome", "true", "Backend가 검증한 선택 또는 취소 결과를 저장한다.", ""),
    ("request_recipient_selection", "output", "to_recipient_id", "resume.value.to_recipient_id", "false", "선택된 기존 수취인 ID를 저장한다.", "selected에서 후보 ID와 정확히 하나"),
    ("request_recipient_selection", "output", "to_recipient_candidate_id", "resume.value.to_recipient_candidate_id", "false", "검증된 신규 수취인 후보 ID를 저장한다.", "selected에서 기존 ID와 정확히 하나"),
    ("resolve_external_from_account", "input", "from_account_hint", "query.account_hint", "false", "출금 계좌 힌트를 Backend 계좌 확인에 전달한다.", ""),
    ("resolve_external_from_account", "output", "account_resolution_outcome", "response.data.account_resolution_outcome", "true", "출금 계좌 자동 확정 결과를 저장한다.", ""),
    ("resolve_external_from_account", "output", "accounts", "response.data.accounts", "true", "Backend가 검증한 최신 출금 계좌 후보를 저장한다.", ""),
    ("resolve_external_from_account", "output", "from_account_id", "response.data.account_ids[0]", "false", "자동 확정된 단일 출금 계좌 ID를 저장한다.", ""),
    ("request_external_from_account_selection", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "출금 계좌 후보를 선택 UI에 전달한다.", ""),
    ("request_external_from_account_selection", "input", "correction_view", "webhook.metadata.ui.payload.correction_view", "false", "출금 계좌 수정 사유를 전달한다.", ""),
    ("request_external_from_account_selection", "output", "account_selection_outcome", "resume.value.account_selection_outcome", "true", "계좌 선택 또는 취소 결과를 저장한다.", ""),
    ("request_external_from_account_selection", "output", "from_account_id", "resume.value.account_ids[0]", "false", "선택된 단일 출금 계좌 ID를 저장한다.", ""),
    ("emit_external_from_accounts_empty", "input", "accounts", "webhook.metadata.ui.payload.accounts", "true", "빈 출금 계좌 목록을 동일 선택 UI에 전달한다.", "accounts=[]"),
    ("check_external_transfer_amount", "input", "amount", "agent.state.amount", "false", "금액의 존재, 정수 여부와 양수 여부만 확인한다.", ""),
    ("request_external_transfer_amount", "input", "correction_view", "webhook.metadata.ui.payload.correction_view", "false", "금액 수정 사유를 전달한다.", ""),
    ("request_external_transfer_amount", "input", "currency", "webhook.metadata.ui.payload.currency", "true", "금액 입력 통화를 전달한다.", "KRW"),
    ("request_external_transfer_amount", "output", "amount_input_outcome", "resume.value.amount_input_outcome", "true", "금액 제출 또는 취소 결과를 저장한다.", ""),
    ("request_external_transfer_amount", "output", "amount", "resume.value.amount", "false", "Backend가 검증한 송금 금액을 저장한다.", "submitted일 때 필수"),
    ("start_external_transfer_prepare", "output", "prepare_attempt", "agent.state.prepare_attempt", "true", "새 논리 Prepare 시도 번호를 증가시킨다.", ""),
    ("start_external_transfer_prepare", "output", "correction_view", "agent.state.correction_view=null", "false", "새 Prepare 전에 이전 수정 사유를 제거한다.", ""),
    ("prepare_external_transfer", "input", "from_account_id", "request.from_account_id", "true", "검증된 출금 계좌 ID를 전달한다.", ""),
    ("prepare_external_transfer", "input", "to_recipient_id", "request.to_recipient_id", "false", "기존 수취인 ID를 조건부 전달한다.", "후보 ID와 정확히 하나"),
    ("prepare_external_transfer", "input", "to_recipient_candidate_id", "request.to_recipient_candidate_id", "false", "신규 검증 수취인 후보 ID를 조건부 전달한다.", "기존 ID와 정확히 하나"),
    ("prepare_external_transfer", "input", "amount", "request.amount", "true", "송금 금액을 전달한다.", ""),
    ("prepare_external_transfer", "input", "currency", "request.currency", "true", "송금 통화를 전달한다.", ""),
    ("prepare_external_transfer", "input", "prepare_attempt", "header.Idempotency-Key", "true", "실행 Context와 시도로 Prepare 멱등성 키를 생성한다.", "external_transfer_prepare:{execution_context_id}:{prepare_attempt}"),
    ("prepare_external_transfer", "output", "confirmation_id", "response.data.confirmation_id", "false", "승인 준비 완료 시 Confirmation ID를 저장한다.", ""),
    ("prepare_external_transfer", "output", "confirmation_view", "response.data.confirmation_view", "false", "Backend 승인 표시 데이터를 저장한다.", "완료 결과까지 재사용"),
    ("prepare_external_transfer", "output", "correction_view", "response.data.correction_view", "false", "수정 가능 대상과 사유를 저장한다.", ""),
    ("prepare_external_transfer", "output", "blocked_view", "response.data.blocked_view", "false", "사용자 표시용 차단 안내를 저장한다.", ""),
    ("request_external_transfer_approval", "input", "confirmation_id", "webhook.interaction.confirmation_id", "true", "승인 요청과 재개를 연결할 ID를 전달한다.", ""),
    ("request_external_transfer_approval", "input", "confirmation_view", "webhook.metadata.ui.payload", "true", "Backend 승인 표시 데이터를 그대로 전달한다.", ""),
    ("request_external_transfer_approval", "output", "approval_outcome", "resume.value.approval_outcome", "true", "Backend가 검증한 승인 결과를 저장한다.", ""),
    ("request_external_transfer_approval", "output", "change_target", "resume.value.change_target", "false", "수정 요청 대상을 저장한다.", ""),
    ("route_external_transfer_correction", "input", "correction_view", "agent.state.correction_view", "true", "허용 수정 대상의 개수와 값을 확인한다.", ""),
    ("request_external_transfer_correction", "input", "correction_view", "webhook.metadata.ui.payload", "true", "수정 표시 데이터를 선택 UI에 전달한다.", ""),
    ("request_external_transfer_correction", "output", "correction_selection_outcome", "resume.value.correction_selection_outcome", "true", "수정 선택 또는 취소 결과를 저장한다.", ""),
    ("request_external_transfer_correction", "output", "change_target", "resume.value.change_target", "false", "선택한 수정 대상을 저장한다.", ""),
    ("start_external_auth", "output", "auth_context_id", "agent.state.auth_context_id=null", "false", "기존 Auth Context ID를 제거한다.", ""),
    ("start_external_auth", "output", "auth_request_view", "agent.state.auth_request_view=null", "false", "기존 인증 표시 데이터를 제거한다.", ""),
    ("start_external_auth", "output", "auth_status", "agent.state.auth_status=null", "false", "기존 인증 결과를 제거한다.", ""),
    ("start_external_auth", "output", "auth_retry_outcome", "agent.state.auth_retry_outcome=null", "false", "기존 재인증 선택 결과를 제거한다.", ""),
    ("start_external_auth", "output", "auth_attempt", "agent.state.auth_attempt", "true", "새 인증 시도 번호를 증가시킨다.", ""),
    ("create_external_auth_context", "input", "confirmation_id", "request.confirmation_id", "true", "승인된 Confirmation ID를 전달한다.", ""),
    ("create_external_auth_context", "input", "auth_attempt", "header.Idempotency-Key", "true", "Confirmation과 인증 시도로 멱등성 키를 생성한다.", "external_transfer_auth:{confirmation_id}:{auth_attempt}"),
    ("create_external_auth_context", "output", "auth_context_id", "response.data.auth_context_id", "false", "Backend Auth Context ID를 저장한다.", ""),
    ("create_external_auth_context", "output", "auth_request_view", "response.data.auth_request_view", "false", "Backend 인증 표시 데이터를 저장한다.", ""),
    ("create_external_auth_context", "output", "blocked_view", "response.data.blocked_view", "false", "인증 진행 차단 안내를 저장한다.", ""),
    ("request_external_authentication", "input", "auth_context_id", "webhook.interaction.auth_context_id", "true", "인증 요청과 재개를 연결할 ID를 전달한다.", ""),
    ("request_external_authentication", "input", "auth_request_view", "webhook.metadata.ui.payload", "true", "Backend 인증 표시 데이터를 그대로 전달한다.", ""),
    ("request_external_authentication", "output", "auth_status", "resume.value.auth_status", "true", "Backend가 검증한 인증 결과를 저장한다.", ""),
    ("request_external_auth_retry", "input", "auth_request_view", "webhook.metadata.ui.payload.auth_request_view", "false", "실패·만료 인증 안내를 재인증 UI에 전달한다.", ""),
    ("request_external_auth_retry", "output", "auth_retry_outcome", "resume.value.auth_retry_outcome", "true", "재인증 또는 취소 결과를 저장한다.", ""),
    ("execute_external_transfer", "input", "confirmation_id", "request.confirmation_id", "true", "승인된 Confirmation ID를 Execute에 전달한다.", ""),
    ("execute_external_transfer", "input", "auth_context_id", "request.auth_context_id", "true", "검증된 Auth Context ID를 전달한다.", ""),
    ("execute_external_transfer", "input", "auth_attempt", "header.Idempotency-Key", "true", "Confirmation과 인증 시도로 Execute 멱등성 키를 생성한다.", "external_transfer_execute:{confirmation_id}:{auth_attempt}"),
    ("execute_external_transfer", "output", "transaction_id", "response.data.transaction_id", "false", "완료된 거래 ID를 저장한다.", ""),
    ("execute_external_transfer", "output", "completed_at", "response.data.completed_at", "false", "송금 완료 시각을 저장한다.", ""),
    ("execute_external_transfer", "output", "correction_view", "response.data.correction_view", "false", "실행 시점 수정 대상과 사유를 저장한다.", ""),
    ("execute_external_transfer", "output", "blocked_view", "response.data.blocked_view", "false", "실행 차단 안내를 저장한다.", ""),
    ("emit_external_transfer_result", "input", "transaction_id", "webhook.metadata.ui.payload.transaction_id", "true", "거래 ID를 완료 UI에 전달한다.", ""),
    ("emit_external_transfer_result", "input", "completed_at", "webhook.metadata.ui.payload.completed_at", "true", "완료 시각을 완료 UI에 전달한다.", ""),
    ("emit_external_transfer_result", "input", "confirmation_view", "webhook.metadata.ui.payload.from_account", "true", "Prepare 출금 계좌 표시정보를 재사용한다.", "confirmation_view.from_account 투영"),
    ("emit_external_transfer_result", "input", "confirmation_view", "webhook.metadata.ui.payload.recipient", "true", "Prepare 수취인 표시정보를 재사용한다.", "confirmation_view.recipient 투영"),
    ("emit_external_transfer_result", "input", "confirmation_view", "webhook.metadata.ui.payload.amount", "true", "Prepare 금액을 재사용한다.", "confirmation_view.amount 투영"),
    ("emit_external_transfer_result", "input", "confirmation_view", "webhook.metadata.ui.payload.currency", "true", "Prepare 통화를 재사용한다.", "confirmation_view.currency 투영"),
    ("emit_external_transfer_blocked", "input", "blocked_view", "webhook.metadata.ui.payload", "true", "Backend 차단 안내를 그대로 전달한다.", ""),
]

EXTERNAL_TRANSFER_RESET_FIELDS = {
    "reset_external_from_account": ["from_account_hint", "account_resolution_outcome", "accounts", "account_selection_outcome", "from_account_id", "amount_input_outcome", "input_request_id", "confirmation_id", "confirmation_view", "approval_outcome", "change_target", "correction_selection_outcome", "blocked_view", "auth_context_id", "auth_request_view", "auth_status", "auth_retry_outcome", "auth_attempt"],
    "reset_external_recipient": ["recipient_name_hint", "recipient_resolution_outcome", "recipient_selection_reason", "recipient_selection_outcome", "to_recipient_id", "to_recipient_candidate_id", "input_request_id", "confirmation_id", "confirmation_view", "approval_outcome", "change_target", "correction_selection_outcome", "blocked_view", "auth_context_id", "auth_request_view", "auth_status", "auth_retry_outcome", "auth_attempt"],
    "reset_external_transfer_amount": ["amount", "amount_input_outcome", "input_request_id", "confirmation_id", "confirmation_view", "approval_outcome", "change_target", "correction_selection_outcome", "blocked_view", "auth_context_id", "auth_request_view", "auth_status", "auth_retry_outcome", "auth_attempt"],
}


def external_reset_mapping_definitions() -> list[tuple[str, str, str, str, str, str, str]]:
    definitions: list[tuple[str, str, str, str, str, str, str]] = []
    for step_id, state_keys in EXTERNAL_TRANSFER_RESET_FIELDS.items():
        for state_key in state_keys:
            cleared_value = "[]" if state_key == "accounts" else "0" if state_key == "auth_attempt" else "null"
            definitions.append((step_id, "output", state_key, f"agent.state.{state_key}={cleared_value}", "true", "수정 전에 오래된 입력·승인·인증 임시 State를 제거한다.", "correction_view와 prepare_attempt는 유지"))
    return definitions


EXTERNAL_TRANSFER_MAPPING_DEFINITIONS.extend(external_reset_mapping_definitions())

EXTERNAL_TRANSFER_MAPPING_ROWS = [
    {"workflow_id": "wf_external_transfer", "step_id": step_id, "direction": direction,
     "state_key": state_key, "contract_field_path": contract_field_path,
     "required_at_step": required_at_step, "mapping_description": mapping_description,
     "notes": notes, "validation_result": "OK"}
    for (step_id, direction, state_key, contract_field_path, required_at_step, mapping_description, notes)
    in EXTERNAL_TRANSFER_MAPPING_DEFINITIONS
]

GENERAL_INPUT_REQUEST_STEPS = [
    ("wf_balance_inquiry", "request_balance_account_selection"),
    ("wf_transaction_history", "request_transaction_account_selection"),
    ("wf_transaction_history", "request_period_selection"),
    ("wf_period_amount_summary", "request_summary_account_selection"),
    ("wf_period_amount_summary", "request_period_selection"),
    ("wf_period_amount_summary", "request_summary_type"),
    ("wf_set_default_account", "request_default_account_selection"),
    ("wf_set_account_alias", "request_account_alias_selection"),
    ("wf_set_account_alias", "request_account_alias_input"),
    ("wf_internal_transfer", "request_from_account_selection"),
    ("wf_internal_transfer", "request_to_account_selection"),
    ("wf_internal_transfer", "request_internal_transfer_amount"),
    ("wf_internal_transfer", "request_internal_transfer_correction"),
    ("wf_internal_transfer", "request_internal_auth_retry"),
    ("wf_external_transfer", "request_recipient_selection"),
    ("wf_external_transfer", "request_external_from_account_selection"),
    ("wf_external_transfer", "request_external_transfer_amount"),
    ("wf_external_transfer", "request_external_transfer_correction"),
    ("wf_external_transfer", "request_external_auth_retry"),
]

STEP_ROWS_BY_WORKFLOW = {
    "wf_balance_inquiry": BALANCE_STEP_ROWS,
    "wf_transaction_history": TRANSACTION_STEP_ROWS,
    "wf_period_amount_summary": SUMMARY_STEP_ROWS,
    "wf_set_default_account": DEFAULT_ACCOUNT_STEP_ROWS,
    "wf_set_account_alias": ACCOUNT_ALIAS_STEP_ROWS,
    "wf_internal_transfer": INTERNAL_TRANSFER_STEP_ROWS,
    "wf_external_transfer": EXTERNAL_TRANSFER_STEP_ROWS,
}

MAPPING_ROWS_BY_WORKFLOW = {
    "wf_balance_inquiry": BALANCE_MAPPING_ROWS,
    "wf_transaction_history": TRANSACTION_MAPPING_ROWS,
    "wf_period_amount_summary": SUMMARY_MAPPING_ROWS,
    "wf_set_default_account": DEFAULT_ACCOUNT_MAPPING_ROWS,
    "wf_set_account_alias": ACCOUNT_ALIAS_MAPPING_ROWS,
    "wf_internal_transfer": INTERNAL_TRANSFER_MAPPING_ROWS,
    "wf_external_transfer": EXTERNAL_TRANSFER_MAPPING_ROWS,
}


def add_general_input_request_mappings() -> None:
    for workflow_id, step_id in GENERAL_INPUT_REQUEST_STEPS:
        matching_steps = [
            row for row in STEP_ROWS_BY_WORKFLOW[workflow_id]
            if row["step_id"] == step_id
        ]
        if len(matching_steps) != 1:
            raise ValueError(f"일반 입력 Step을 하나로 찾을 수 없습니다: {workflow_id}/{step_id}")

        step = matching_steps[0]
        output_keys = [key.strip() for key in step["output_state_keys"].split(",") if key.strip()]
        if "input_request_id" not in output_keys:
            output_keys.append("input_request_id")
            step["output_state_keys"] = ", ".join(output_keys)

        MAPPING_ROWS_BY_WORKFLOW[workflow_id].append(
            {
                "workflow_id": workflow_id,
                "step_id": step_id,
                "direction": "output",
                "state_key": "input_request_id",
                "contract_field_path": "webhook.metadata.input_request_id",
                "required_at_step": "true",
                "mapping_description": "Agent가 새 일반 입력 요청 ID를 생성하여 State와 Webhook에 동일하게 저장한다.",
                "notes": "Backend는 같은 ID에 Execution Context, Agent Thread와 ui_contract_id를 연결하고 Resume에서 검증",
                "validation_result": "OK",
            }
        )


add_general_input_request_mappings()


def enum_rows() -> list[dict[str, Any]]:
    groups = {
        "workflow_type": [
            ("global", "글로벌"),
            ("inquiry", "조회"),
            ("setting_change", "설정 변경"),
            ("transfer", "송금"),
        ],
        "interaction_mode": [
            ("agent_internal", "Agent 내부 실행"),
            ("backend_tool_api", "Backend Tool API"),
            ("webhook", "결과 Webhook"),
            ("webhook_then_resume", "Webhook 후 재개 대기"),
        ],
        "risk_level": [(f"R{index}", f"위험등급 R{index}") for index in range(6)],
        "approval_policy": [("none", "승인 없음"), ("required", "승인 필수")],
        "auth_policy": [("none", "추가 인증 없음"), ("required", "추가 인증 필수")],
        "status": [
            ("draft", "초안"),
            ("review", "검토 중"),
            ("active", "사용 중"),
            ("deprecated", "폐기 예정"),
            ("removed", "제거"),
        ],
        "schema_scope": [("common", "공통 State"), ("workflow", "Workflow State")],
        "retention_scope": [
            ("interaction", "상호작용 중 보존"),
            ("workflow", "Workflow 중 보존"),
            ("result", "결과로 보존"),
        ],
        "log_policy": [("allow", "기록 허용"), ("masked", "마스킹 기록"), ("exclude", "기록 제외")],
        "mapping_direction": [("input", "Step 입력"), ("output", "Step 출력")],
        "contract_type": [("agent_tool_api", "Agent Tool API"), ("ui_hitl", "UI·HITL")],
        "guardrail_outcome": [("allowed", "정책 허용"), ("blocked", "정책 차단")],
        "workflow_match_outcome": [("matched", "지원 Workflow 매칭"), ("no_match", "일치 Workflow 없음")],
        "dispatch_outcome": [("completed", "하위 Workflow 정상 종료"), ("failed", "하위 Workflow 실행 실패")],
        "account_resolution_outcome": [
            ("resolved", "계좌 자동 확정"),
            ("selection_required", "사용자 선택 필요"),
            ("no_accounts", "조회 가능 계좌 없음"),
        ],
        "account_selection_outcome": [("selected", "계좌 선택 완료"), ("cancelled", "계좌 선택 취소")],
        "approval_outcome": [
            ("approved", "승인 완료"),
            ("change_requested", "수정 요청"),
            ("cancelled", "승인 취소"),
        ],
        "alias_input_outcome": [("submitted", "별칭 입력 완료"), ("cancelled", "별칭 입력 취소")],
        "account_alias_change_target": [("account", "대상 계좌"), ("alias", "새 별칭")],
        "account_alias_prepare_outcome": [
            ("ready_for_confirmation", "승인 준비 완료"),
            ("unchanged", "변경 없음"),
            ("correction_required", "수정 필요"),
            ("blocked", "변경 차단"),
        ],
        "account_alias_execute_outcome": [
            ("completed", "변경 완료"),
            ("correction_required", "수정 필요"),
            ("blocked", "변경 차단"),
        ],
        "amount_input_outcome": [("submitted", "금액 입력 완료"), ("cancelled", "금액 입력 취소")],
        "correction_selection_outcome": [("selected", "수정 대상 선택"), ("cancelled", "수정 취소")],
        "internal_transfer_change_target": [
            ("from_account", "출금 계좌"),
            ("to_account", "입금 계좌"),
            ("amount", "송금 금액"),
        ],
        "auth_status": [
            ("verified", "인증 완료"),
            ("failed", "인증 실패"),
            ("cancelled", "인증 취소"),
            ("expired", "인증 만료"),
        ],
        "auth_retry_outcome": [("retry", "재인증"), ("cancelled", "재인증 취소")],
        "internal_transfer_prepare_outcome": [
            ("ready_for_confirmation", "승인 준비 완료"),
            ("correction_required", "입력 수정 필요"),
            ("blocked", "송금 차단"),
        ],
        "internal_transfer_execute_outcome": [
            ("completed", "송금 완료"),
            ("correction_required", "입력 수정 필요"),
            ("reauthentication_required", "재인증 필요"),
            ("blocked", "송금 차단"),
        ],
        "recipient_resolution_outcome": [("resolved", "수취인 자동 확정"), ("selection_required", "수취인 선택 필요")],
        "recipient_selection_reason": [("multiple_matches", "동명이인 복수"), ("no_match", "기존 거래 일치 없음")],
        "recipient_selection_outcome": [("selected", "수취인 선택 완료"), ("cancelled", "수취인 선택 취소")],
        "external_transfer_change_target": [("from_account", "출금 계좌"), ("recipient", "수취인"), ("amount", "송금 금액")],
        "external_transfer_prepare_outcome": [
            ("ready_for_confirmation", "승인 준비 완료"),
            ("correction_required", "입력 수정 필요"),
            ("blocked", "송금 차단"),
        ],
        "external_transfer_execute_outcome": [
            ("completed", "송금 완료"),
            ("correction_required", "입력 수정 필요"),
            ("reauthentication_required", "재인증 필요"),
            ("blocked", "송금 차단"),
        ],
        "default_account_prepare_outcome": [
            ("ready_for_confirmation", "승인 준비 완료"),
            ("unchanged", "변경 없음"),
            ("correction_required", "계좌 수정 필요"),
            ("blocked", "변경 차단"),
        ],
        "default_account_execute_outcome": [
            ("completed", "변경 완료"),
            ("correction_required", "계좌 수정 필요"),
            ("blocked", "변경 차단"),
        ],
        "period_selection_outcome": [("selected", "기간 선택 완료"), ("cancelled", "기간 선택 취소")],
        "summary_type": [("spending", "지출 합계"), ("income", "수입 합계")],
        "summary_type_selection_outcome": [
            ("selected", "합계 유형 선택 완료"),
            ("cancelled", "합계 유형 선택 취소"),
        ],
        "transaction_type": [
            ("deposit", "입금"),
            ("withdrawal", "일반 출금"),
            ("transfer", "계좌이체"),
            ("card_payment", "카드 결제"),
            ("atm_withdrawal", "ATM 출금"),
            ("fee", "수수료"),
            ("interest", "이자"),
        ],
    }
    rows: list[dict[str, Any]] = []
    for group, values in groups.items():
        for index, (value, display_name) in enumerate(values, start=1):
            rows.append(
                {
                    "enum_group": group,
                    "enum_value": value,
                    "display_name": display_name,
                    "description": "관리시트 드롭다운과 검증에 사용하는 값",
                    "source_type": "sheet_rule",
                    "source_document": "agent-team-integration-implementation-roadmap.md",
                    "status": "active",
                    "sort_order": index,
                }
            )
    return rows


HEADER_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EDITABLE_FILL = PatternFill("solid", fgColor="FFFFFF")
GENERATED_FILL = PatternFill("solid", fgColor="E7E6E6")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)

GENERATED_COLUMNS = {
    "Workflow Steps": {
        "external_action",
        "input_state_keys",
        "output_state_keys",
        "route_summary",
        "validation_result",
    },
    "Step Data Mapping": {"validation_result"},
}


def append_rows(sheet: Any, headers: list[str], rows: Iterable[dict[str, Any]]) -> None:
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def style_sheet(sheet: Any, headers: list[str]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(row=max(sheet.max_row, 2), column=len(headers)).coordinate}"
    sheet.row_dimensions[1].height = 30
    generated = GENERATED_COLUMNS.get(sheet.title, set())

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row in sheet.iter_rows(min_row=2, max_row=max(sheet.max_row, 2), max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            header = headers[cell.column - 1]
            cell.fill = GENERATED_FILL if header in generated else EDITABLE_FILL

    widths = {
        "description": 42,
        "step_purpose": 42,
        "example_utterances": 42,
        "condition_description": 48,
        "mapping_description": 44,
        "contract_summary": 38,
        "transport_target": 48,
        "notes": 36,
        "route_summary": 42,
        "input_state_keys": 36,
        "output_state_keys": 36,
        "clear_when": 36,
        "source_document": 42,
    }
    for index, header in enumerate(headers, start=1):
        width = widths.get(header, max(14, min(len(header) + 6, 28)))
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width

    if "validation_result" in headers:
        column = sheet.cell(row=1, column=headers.index("validation_result") + 1).column_letter
        sheet.conditional_formatting.add(
            f"{column}2:{column}2000",
            FormulaRule(formula=[f'AND({column}2<>"",{column}2<>"OK")'], fill=ERROR_FILL),
        )


def add_list_validation(sheet: Any, header: str, values: list[str]) -> None:
    headers = [cell.value for cell in sheet[1]]
    if header not in headers:
        return
    column = sheet.cell(row=1, column=headers.index(header) + 1).column_letter
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "허용된 목록에서 값을 선택해 주세요."
    validation.errorTitle = "허용되지 않은 값"
    validation.prompt = "드롭다운 목록에서 값을 선택할 수 있습니다."
    validation.promptTitle = header
    sheet.add_data_validation(validation)
    validation.add(f"{column}2:{column}2000")


def add_validations(workbook: Workbook) -> None:
    catalog = workbook["Workflow Catalog"]
    add_list_validation(catalog, "workflow_type", ["global", "inquiry", "setting_change", "transfer"])
    add_list_validation(catalog, "max_risk_level", [f"R{index}" for index in range(6)])
    add_list_validation(catalog, "approval_policy", ["none", "required"])
    add_list_validation(catalog, "auth_policy", ["none", "required"])
    add_list_validation(catalog, "status", ["draft", "review", "active", "deprecated", "removed"])

    steps = workbook["Workflow Steps"]
    add_list_validation(steps, "interaction_mode", ["agent_internal", "backend_tool_api", "webhook", "webhook_then_resume"])
    add_list_validation(steps, "step_risk_level", [f"R{index}" for index in range(6)])
    add_list_validation(steps, "status", ["draft", "review", "active", "deprecated", "removed"])

    routes = workbook["Workflow Routes"]
    add_list_validation(routes, "status", ["draft", "review", "active", "deprecated", "removed"])

    schema = workbook["Workflow Data Schema"]
    add_list_validation(schema, "schema_scope", ["common", "workflow"])
    add_list_validation(schema, "nullable", ["true", "false"])
    add_list_validation(schema, "retention_scope", ["interaction", "workflow", "result"])
    add_list_validation(schema, "sensitive", ["true", "false"])
    add_list_validation(schema, "log_policy", ["allow", "masked", "exclude"])

    mapping = workbook["Step Data Mapping"]
    add_list_validation(mapping, "direction", ["input", "output"])
    add_list_validation(mapping, "required_at_step", ["true", "false"])

    registry = workbook["Contract Registry"]
    add_list_validation(registry, "contract_type", ["agent_tool_api", "ui_hitl"])
    add_list_validation(registry, "status", ["draft", "review", "active", "deprecated", "removed"])

    enums = workbook["Enum Registry"]
    add_list_validation(enums, "source_type", ["sheet_rule", "api_contract", "ui_contract"])
    add_list_validation(enums, "status", ["active", "deprecated", "removed"])


def contract_mapping_rows() -> list[dict[str, Any]]:
    registry = {row["contract_id"]: row for row in CONTRACT_ROWS}
    rows: list[dict[str, Any]] = []
    for step in (
        GLOBAL_STEP_ROWS
        + ACCOUNT_LIST_STEP_ROWS
        + BALANCE_STEP_ROWS
        + TRANSACTION_STEP_ROWS
        + SUMMARY_STEP_ROWS
        + DEFAULT_ACCOUNT_STEP_ROWS
        + ACCOUNT_ALIAS_STEP_ROWS
        + INTERNAL_TRANSFER_STEP_ROWS
        + EXTERNAL_TRANSFER_STEP_ROWS
    ):
        contract_id = step["contract_id"]
        if not contract_id:
            continue
        contract = registry[contract_id]
        rows.append(
            {
                "workflow_id": step["workflow_id"],
                "step_id": step["step_id"],
                "interaction_mode": step["interaction_mode"],
                "contract_id": contract_id,
                "transport_target": contract["transport_target"],
                "contract_version": contract["contract_version"],
            }
        )
    return rows


def build_workbook(output_path: Path, force: bool = False) -> Path:
    if output_path.exists() and not force:
        raise FileExistsError(f"이미 파일이 있습니다: {output_path}. 다시 만들려면 --force를 사용하세요.")

    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, headers in SHEET_COLUMNS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        if sheet_name == "Workflow Catalog":
            append_rows(sheet, headers, WORKFLOW_ROWS)
        elif sheet_name == "Workflow Steps":
            append_rows(
                sheet,
                headers,
                GLOBAL_STEP_ROWS
                + ACCOUNT_LIST_STEP_ROWS
                + BALANCE_STEP_ROWS
                + TRANSACTION_STEP_ROWS
                + SUMMARY_STEP_ROWS
                + DEFAULT_ACCOUNT_STEP_ROWS
                + ACCOUNT_ALIAS_STEP_ROWS
                + INTERNAL_TRANSFER_STEP_ROWS
                + EXTERNAL_TRANSFER_STEP_ROWS,
            )
        elif sheet_name == "Workflow Routes":
            append_rows(
                sheet,
                headers,
                GLOBAL_ROUTE_ROWS
                + ACCOUNT_LIST_ROUTE_ROWS
                + BALANCE_ROUTE_ROWS
                + TRANSACTION_ROUTE_ROWS
                + SUMMARY_ROUTE_ROWS
                + DEFAULT_ACCOUNT_ROUTE_ROWS
                + ACCOUNT_ALIAS_ROUTE_ROWS
                + INTERNAL_TRANSFER_ROUTE_ROWS
                + EXTERNAL_TRANSFER_ROUTE_ROWS,
            )
        elif sheet_name == "Workflow Data Schema":
            append_rows(sheet, headers, SCHEMA_ROWS)
        elif sheet_name == "Step Data Mapping":
            append_rows(
                sheet,
                headers,
                GLOBAL_MAPPING_ROWS
                + ACCOUNT_LIST_MAPPING_ROWS
                + BALANCE_MAPPING_ROWS
                + TRANSACTION_MAPPING_ROWS
                + SUMMARY_MAPPING_ROWS
                + DEFAULT_ACCOUNT_MAPPING_ROWS
                + ACCOUNT_ALIAS_MAPPING_ROWS
                + INTERNAL_TRANSFER_MAPPING_ROWS
                + EXTERNAL_TRANSFER_MAPPING_ROWS,
            )
        elif sheet_name == "Contract Registry":
            append_rows(sheet, headers, CONTRACT_ROWS)
        elif sheet_name == "Contract Mapping":
            append_rows(sheet, headers, contract_mapping_rows())
        elif sheet_name == "Enum Registry":
            append_rows(sheet, headers, enum_rows())
        style_sheet(sheet, headers)

    add_validations(workbook)
    workbook.properties.title = "Agent Workflow 관리시트 v3"
    workbook.properties.subject = "Agent 내부 실행, Backend Tool API, Webhook과 HITL 계약 매핑"
    workbook.properties.creator = "Agent Team"
    workbook.properties.description = "사람 중심 Workflow 설계 정본과 계약 연결 관리시트"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def validate_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != list(SHEET_COLUMNS):
        raise ValueError(f"탭 순서가 다릅니다: {workbook.sheetnames}")

    for sheet_name, expected_headers in SHEET_COLUMNS.items():
        sheet = workbook[sheet_name]
        actual_headers = [cell.value for cell in sheet[1]]
        if actual_headers != expected_headers:
            raise ValueError(f"{sheet_name} 컬럼이 다릅니다: {actual_headers}")

    catalog = workbook["Workflow Catalog"]
    workflow_ids = [catalog.cell(row=row, column=1).value for row in range(2, catalog.max_row + 1)]
    if len(workflow_ids) != 9 or len(set(workflow_ids)) != 9:
        raise ValueError(f"Workflow Catalog는 고유한 9개 Workflow여야 합니다: {workflow_ids}")

    registry = workbook["Contract Registry"]
    registry_rows = [
        {
            "contract_id": registry.cell(row=row, column=1).value,
            "contract_type": registry.cell(row=row, column=2).value,
        }
        for row in range(2, registry.max_row + 1)
    ]
    contract_ids = [row["contract_id"] for row in registry_rows]
    api_contract_ids = [row["contract_id"] for row in registry_rows if row["contract_type"] == "agent_tool_api"]
    if len(contract_ids) != len(set(contract_ids)):
        raise ValueError(f"Contract Registry의 contract_id가 중복됩니다: {contract_ids}")
    if len(api_contract_ids) != 14:
        raise ValueError(f"Contract Registry는 14개 API 계약을 가져야 합니다: {api_contract_ids}")

    steps = workbook["Workflow Steps"]
    step_rows = [
        {
            "workflow_id": steps.cell(row=row, column=1).value,
            "step_id": steps.cell(row=row, column=3).value,
        }
        for row in range(2, steps.max_row + 1)
    ]
    expected_step_counts = {
        "wf_global_agent_entry": 6,
        "wf_account_list": 4,
        "wf_balance_inquiry": 7,
        "wf_transaction_history": 9,
        "wf_period_amount_summary": 11,
        "wf_set_default_account": 13,
        "wf_set_account_alias": 16,
        "wf_internal_transfer": 25,
        "wf_external_transfer": 24,
    }
    for workflow_id, expected_count in expected_step_counts.items():
        workflow_step_ids = [row["step_id"] for row in step_rows if row["workflow_id"] == workflow_id]
        if len(workflow_step_ids) != expected_count or len(set(workflow_step_ids)) != expected_count:
            raise ValueError(
                f"{workflow_id}는 고유한 {expected_count}개 Step을 가져야 합니다: {workflow_step_ids}"
            )
    routes = workbook["Workflow Routes"]
    step_ids_by_workflow = {
        workflow_id: {row["step_id"] for row in step_rows if row["workflow_id"] == workflow_id}
        for workflow_id in expected_step_counts
    }
    invalid_routes = []
    for row in range(2, routes.max_row + 1):
        workflow_id = routes.cell(row=row, column=1).value
        from_step_id = routes.cell(row=row, column=2).value
        to_step_id = routes.cell(row=row, column=5).value
        valid_step_ids = step_ids_by_workflow.get(workflow_id, set())
        if from_step_id not in valid_step_ids or (to_step_id != "END" and to_step_id not in valid_step_ids):
            invalid_routes.append((workflow_id, from_step_id, to_step_id))
    if invalid_routes:
        raise ValueError(f"같은 Workflow에 존재하지 않는 Step을 참조하는 Route가 있습니다: {invalid_routes}")

    schema = workbook["Workflow Data Schema"]
    mapping = workbook["Step Data Mapping"]
    for workflow_id in expected_step_counts:
        workflow_state_keys = {
            schema.cell(row=row, column=3).value
            for row in range(2, schema.max_row + 1)
            if schema.cell(row=row, column=2).value == workflow_id
        }
        mapped_state_keys = {
            mapping.cell(row=row, column=4).value
            for row in range(2, mapping.max_row + 1)
            if mapping.cell(row=row, column=1).value == workflow_id
        }
        unknown_state_keys = sorted(mapped_state_keys - workflow_state_keys)
        if unknown_state_keys:
            raise ValueError(f"{workflow_id} Data Schema에 없는 State Mapping이 있습니다: {unknown_state_keys}")

    invalid_mappings = []
    for row in range(2, mapping.max_row + 1):
        workflow_id = mapping.cell(row=row, column=1).value
        step_id = mapping.cell(row=row, column=2).value
        if step_id not in step_ids_by_workflow.get(workflow_id, set()):
            invalid_mappings.append((workflow_id, step_id))
    if invalid_mappings:
        raise ValueError(f"같은 Workflow에 존재하지 않는 Step을 참조하는 Mapping이 있습니다: {invalid_mappings}")

    contract_mapping = workbook["Contract Mapping"]
    invalid_contract_mappings = []
    for row in range(2, contract_mapping.max_row + 1):
        workflow_id = contract_mapping.cell(row=row, column=1).value
        step_id = contract_mapping.cell(row=row, column=2).value
        contract_id = contract_mapping.cell(row=row, column=4).value
        if step_id not in step_ids_by_workflow.get(workflow_id, set()) or contract_id not in contract_ids:
            invalid_contract_mappings.append((workflow_id, step_id, contract_id))
    if invalid_contract_mappings:
        raise ValueError(f"유효하지 않은 Contract Mapping이 있습니다: {invalid_contract_mappings}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=OUTPUT_PATH)
    parser.add_argument("--force", action="store_true")
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_args()
    created = build_workbook(arguments.output, force=arguments.force)
    validate_workbook(created)
    print(created)
