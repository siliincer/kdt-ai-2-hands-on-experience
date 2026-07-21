"""관리시트 V3를 개발 에이전트용 JSON 계약으로 변환하고 검증한다."""

from __future__ import annotations

import argparse
import hashlib
import json
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

AGENT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = AGENT_DIR / "docs" / "agent-management-sheet-v3.xlsx"
DEFAULT_OUTPUT = AGENT_DIR / "contracts" / "workflow-contracts.json"

SHEETS = {
    "catalog": "Workflow Catalog",
    "steps": "Workflow Steps",
    "routes": "Workflow Routes",
    "data_schema": "Workflow Data Schema",
    "step_mappings": "Step Data Mapping",
    "contract_registry": "Contract Registry",
    "contract_mappings": "Contract Mapping",
    "enums": "Enum Registry",
}

LIST_FIELDS = {"input_state_keys", "output_state_keys"}
BOOLEAN_FIELDS = {"nullable", "sensitive", "required_at_step"}
INTERACTION_MODES = {
    "agent_internal",
    "backend_tool_api",
    "webhook",
    "webhook_then_resume",
}


class ContractValidationError(ValueError):
    """관리시트 계약이 구현에 사용할 수 없는 경우 발생한다."""


def _normalize_value(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _split_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in str(value).split(",") if item.strip()]


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() == "true"


def _read_rows(sheet: Worksheet) -> list[dict[str, Any]]:
    values = sheet.iter_rows(values_only=True)
    try:
        header_row = next(values)
    except StopIteration:
        return []

    headers = [str(value).strip() if value is not None else "" for value in header_row]
    rows: list[dict[str, Any]] = []
    for raw_row in values:
        row = {
            header: _normalize_value(raw_row[index] if index < len(raw_row) else None)
            for index, header in enumerate(headers)
            if header
        }
        if not any(value is not None for value in row.values()):
            continue
        for field in LIST_FIELDS:
            if field in row:
                row[field] = _split_list(row[field])
        for field in BOOLEAN_FIELDS:
            if field in row:
                row[field] = _to_bool(row[field])
        rows.append(row)
    return rows


def _unique_index(
    rows: list[dict[str, Any]], key: str, label: str
) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        raw_value = row.get(key)
        if not raw_value:
            raise ContractValidationError(f"{label}에 {key}가 없는 행이 있습니다.")
        value = str(raw_value)
        if value in result:
            raise ContractValidationError(f"{label}의 {key}가 중복입니다: {value}")
        result[value] = row
    return result


def build_manifest(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        missing_sheets = [
            name for name in SHEETS.values() if name not in workbook.sheetnames
        ]
        if missing_sheets:
            missing = ", ".join(missing_sheets)
            raise ContractValidationError(f"관리시트 탭이 없습니다: {missing}")
        rows = {
            key: _read_rows(workbook[sheet_name]) for key, sheet_name in SHEETS.items()
        }
    finally:
        workbook.close()

    catalog = _unique_index(rows["catalog"], "workflow_id", "Workflow Catalog")
    contracts = _unique_index(
        rows["contract_registry"], "contract_id", "Contract Registry"
    )

    common_state_schema: list[dict[str, Any]] = []
    workflow_state_schema: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows["data_schema"]:
        if row.get("schema_scope") == "common":
            common_state_schema.append(row)
            continue
        workflow_id = row.get("workflow_id")
        if not workflow_id:
            raise ContractValidationError(
                "Workflow Data Schema에 workflow_id가 없습니다."
            )
        workflow_state_schema[str(workflow_id)].append(row)

    grouped: dict[str, dict[str, list[dict[str, Any]]]] = {
        workflow_id: {
            "steps": [],
            "routes": [],
            "state_schema": workflow_state_schema.get(workflow_id, []),
            "step_data_mappings": [],
            "contract_mappings": [],
        }
        for workflow_id in catalog
    }

    grouping = {
        "steps": "steps",
        "routes": "routes",
        "step_mappings": "step_data_mappings",
        "contract_mappings": "contract_mappings",
    }
    for source_key, target_key in grouping.items():
        for row in rows[source_key]:
            workflow_id = row.get("workflow_id")
            if workflow_id not in grouped:
                raise ContractValidationError(
                    f"{SHEETS[source_key]}에 등록되지 않은 workflow_id가 있습니다: "
                    f"{workflow_id}"
                )
            grouped[str(workflow_id)][target_key].append(row)

    workflows = {
        workflow_id: {
            "catalog": catalog_row,
            **grouped[workflow_id],
        }
        for workflow_id, catalog_row in catalog.items()
    }

    enums: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows["enums"]:
        enum_group = row.get("enum_group")
        if not enum_group:
            raise ContractValidationError("Enum Registry에 enum_group이 없습니다.")
        enums[str(enum_group)].append(row)

    manifest = {
        "manifest_version": "1.0.0",
        "source": {
            "workbook": "docs/agent-management-sheet-v3.xlsx",
            "sha256": hashlib.sha256(workbook_path.read_bytes()).hexdigest(),
        },
        "common_state_schema": common_state_schema,
        "contracts": contracts,
        "enums": dict(enums),
        "workflows": workflows,
    }
    validate_manifest(manifest)
    return manifest


def validate_manifest(manifest: dict[str, Any]) -> None:
    contracts: dict[str, dict[str, Any]] = manifest["contracts"]
    common_keys = {
        row["state_key"]
        for row in manifest["common_state_schema"]
        if row.get("state_key")
    }

    for workflow_id, workflow in manifest["workflows"].items():
        steps: list[dict[str, Any]] = workflow["steps"]
        step_ids: set[str] = set()
        for step in steps:
            step_id = str(step.get("step_id") or "")
            if not step_id:
                raise ContractValidationError(
                    f"[{workflow_id}] step_id가 없는 Step입니다."
                )
            if step_id in step_ids:
                raise ContractValidationError(
                    f"[{workflow_id}] step_id가 중복입니다: {step_id}"
                )
            step_ids.add(step_id)

        entry_step_id = workflow["catalog"].get("entry_step_id")
        if entry_step_id not in step_ids:
            raise ContractValidationError(
                f"[{workflow_id}] entry_step_id가 Step에 없습니다: {entry_step_id}"
            )

        schema_keys = common_keys | {
            row["state_key"] for row in workflow["state_schema"] if row.get("state_key")
        }
        for step in steps:
            step_id = str(step["step_id"])
            interaction_mode = step.get("interaction_mode")
            if interaction_mode not in INTERACTION_MODES:
                raise ContractValidationError(
                    f"[{workflow_id}/{step_id}] interaction_mode가 잘못되었습니다: "
                    f"{interaction_mode}"
                )
            contract_id = step.get("contract_id")
            if contract_id and contract_id not in contracts:
                raise ContractValidationError(
                    f"[{workflow_id}/{step_id}] 등록되지 않은 contract_id입니다: "
                    f"{contract_id}"
                )
            for state_key in step.get("input_state_keys", []) + step.get(
                "output_state_keys", []
            ):
                if state_key not in schema_keys:
                    raise ContractValidationError(
                        f"[{workflow_id}/{step_id}] 선언되지 않은 State입니다: "
                        f"{state_key}"
                    )

        for route in workflow["routes"]:
            from_step_id = route.get("from_step_id")
            to_step_id = route.get("to_step_id")
            if from_step_id not in step_ids:
                raise ContractValidationError(
                    f"[{workflow_id}] Route 출발 Step이 없습니다: {from_step_id}"
                )
            if to_step_id != "END" and to_step_id not in step_ids:
                raise ContractValidationError(
                    f"[{workflow_id}] Route 도착 Step이 없습니다: {to_step_id}"
                )

        for mapping in workflow["step_data_mappings"]:
            step_id = mapping.get("step_id")
            state_key = mapping.get("state_key")
            if step_id not in step_ids:
                raise ContractValidationError(
                    f"[{workflow_id}] Mapping 대상 Step이 없습니다: {step_id}"
                )
            if state_key not in schema_keys:
                raise ContractValidationError(
                    f"[{workflow_id}/{step_id}] Mapping State가 선언되지 않았습니다: "
                    f"{state_key}"
                )

        for mapping in workflow["contract_mappings"]:
            step_id = mapping.get("step_id")
            contract_id = mapping.get("contract_id")
            if step_id not in step_ids:
                raise ContractValidationError(
                    f"[{workflow_id}] 계약 대상 Step이 없습니다: {step_id}"
                )
            if contract_id not in contracts:
                raise ContractValidationError(
                    f"[{workflow_id}/{step_id}] 등록되지 않은 계약입니다: {contract_id}"
                )


def render_manifest(manifest: dict[str, Any]) -> str:
    return json.dumps(manifest, ensure_ascii=False, indent=2) + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument(
        "--check",
        action="store_true",
        help="생성 파일이 현재 관리시트와 일치하는지만 검사합니다.",
    )
    mode.add_argument(
        "--workflow",
        help="지정한 workflow_id의 계약만 JSON으로 출력합니다.",
    )
    mode.add_argument(
        "--contract",
        help="지정한 contract_id의 계약만 JSON으로 출력합니다.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = build_manifest(args.workbook)
    if args.workflow:
        workflow = manifest["workflows"].get(args.workflow)
        if workflow is None:
            print(f"등록되지 않은 workflow_id입니다: {args.workflow}")
            return 1
        print(json.dumps(workflow, ensure_ascii=False, indent=2))
        return 0
    if args.contract:
        contract = manifest["contracts"].get(args.contract)
        if contract is None:
            print(f"등록되지 않은 contract_id입니다: {args.contract}")
            return 1
        print(json.dumps(contract, ensure_ascii=False, indent=2))
        return 0
    rendered = render_manifest(manifest)
    if args.check:
        if not args.output.exists():
            print(f"계약 생성 파일이 없습니다: {args.output}")
            return 1
        if args.output.read_text(encoding="utf-8") != rendered:
            print("계약 생성 파일이 관리시트와 일치하지 않습니다.")
            return 1
        print(
            "Workflow 계약 확인 완료: "
            f"{len(manifest['workflows'])}개 Workflow, "
            f"{len(manifest['contracts'])}개 계약"
        )
        return 0

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(rendered, encoding="utf-8")
    print(f"Workflow 계약 생성 완료: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
