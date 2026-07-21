"""Google 스프레드시트 → agent config/*.yaml 생성 스크립트 (시트 v2 대응).

스프레드시트가 YAML의 source of truth다. 각 탭을 공개 CSV(gviz)로 내려받아
엔진이 읽는 형식으로 config YAML 5개를 생성한다.

fin-ai 원본 스크립트와의 주요 차이:
  - gid 대신 탭 이름으로 페치(gviz endpoint) — 탭 추가/재정렬에 안전
  - tools는 Tool_v2 탭 사용 (input_state_keys/write_state_keys 계약)
  - Step의 flat output_data_key를 네임스페이스 키(balance.* 등)로 매핑
  - 라우트 하드닝: 빈 to_step_id는 END 합성, 깨진 참조 라우트는 드롭(경고)

탭 구성 → YAML:
    Workflow + Workflow Step + Workflow Routing + Workflow Data Schema
                                                       → workflows.yaml
    Task        → tasks.yaml
    Tool_v2     → tools.yaml
    Risk Level  → risk_levels.yaml
    Guardrail Rule → guardrail_rules.yaml

사용법 (레포 루트에서):
    uv run python agent/scripts/sync_config_from_sheets.py --dry-run
    uv run python agent/scripts/sync_config_from_sheets.py
    uv run python agent/scripts/sync_config_from_sheets.py --xlsx sheet.xlsx

의존성: 표준 라이브러리 + pyyaml (오프라인 --xlsx 모드만 openpyxl 필요).
경고는 전부 advisory다 — 시트 정리 요청 목록으로 활용한다.
"""

from __future__ import annotations

import argparse
import csv
import io
import os
import re
import shutil
import sys
import urllib.parse
import urllib.request

import yaml

SPREADSHEET_ID = "18gNcQfyC4EhYZricaSHLXbCmjkT5_VDE1c6jmlgChao"

# 논리 이름 -> 시트 탭 이름
SHEETS = {
    "workflows": "Workflow",
    "steps": "Workflow Step",
    "routes": "Workflow Routing",
    "data_schema": "Workflow Data Schema",
    "tasks": "Task",
    "tools": "Tool_v2",
    "risk_levels": "Risk Level",
    "guardrail_rules": "Guardrail Rule",
}

# 생성 위치 기본값: agent/src/agent/config
_AGENT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_DEFAULT_OUT = os.path.join(_AGENT_DIR, "src", "agent", "config")

# output_data_key 중 state 시스템 필드로 유지되는 키(네임스페이스 매핑 제외)
SYSTEM_OUTPUT_KEYS = {
    "final_response",
    "log_id",
    "guardrail_result",
    "matched_workflow_id",
    "workflow_result",
}

# 워크플로우별 네임스페이스 (매핑에 없는 flat 키의 자동 접두용)
WORKFLOW_NAMESPACE = {
    "wf_balance_inquiry": "balance",
    "wf_external_transfer": "transfer",
}

# Step 시트 flat 키 → 네임스페이스 키 매핑 (Tool_v2/Data Schema 기준)
OUTPUT_KEY_MAP = {
    "wf_balance_inquiry": {
        "account_hint": "balance.account_hint",
        "account_candidates": "balance.account_candidates",
        "selected_account": "balance.selected_accounts",
        "selected_accounts": "balance.selected_accounts",
        "account_selection_input": "balance.account_selection_input",
        "balance_result": "balance.balance_results",
        "balance_results": "balance.balance_results",
    },
    "wf_external_transfer": {
        "recipient": "transfer.recipient",
        "recipient_name": "transfer.recipient",
        "amount": "transfer.amount",
        "from_account": "transfer.from_account",
    },
}

_warnings: list[str] = []


def warn(message: str) -> None:
    _warnings.append(message)
    print(f"  ! {message}")


# ---------------------------------------------------------------------------
# Fetch
# ---------------------------------------------------------------------------


def _dedupe_header(header: list[str]) -> list[str]:
    """중복 헤더를 접미사(_2, _3...)로 분리해 보존한다."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for raw in header:
        name = (raw or "").strip()
        if name in seen:
            seen[name] += 1
            out.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 1
            out.append(name)
    return out


def _rows_from_matrix(matrix: list[list]) -> list[dict]:
    """(헤더 포함) 2차원 값 배열을 행(dict) 리스트로 변환한다."""
    if not matrix:
        return []
    header = _dedupe_header([str(c) if c is not None else "" for c in matrix[0]])
    out: list[dict] = []
    for cells in matrix[1:]:
        values = ["" if c is None else str(c) for c in cells]
        if not any(v.strip() for v in values):
            continue
        row = {
            header[i]: (values[i] if i < len(values) else "")
            for i in range(len(header))
        }
        out.append(row)
    return out


def fetch_csv(sheet_name: str) -> list[dict]:
    """탭을 gviz CSV로 받아 행(dict) 리스트로 반환한다 (gid 불필요)."""
    url = (
        f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}"
        f"/gviz/tq?tqx=out:csv&sheet={urllib.parse.quote(sheet_name)}"
    )
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:  # noqa: S310 (공개 시트)
        text = resp.read().decode("utf-8")

    if text.lstrip().startswith("<"):
        raise RuntimeError(
            f"'{sheet_name}' 탭 응답이 CSV가 아닙니다(HTML). "
            "시트가 비공개이거나 탭 이름이 바뀌었는지 확인하세요."
        )
    matrix = list(csv.reader(io.StringIO(text)))
    return _rows_from_matrix(matrix)


def fetch_all(xlsx_path: str | None) -> dict[str, list[dict]]:
    """온라인(gviz) 또는 로컬 xlsx에서 모든 탭을 읽는다."""
    if xlsx_path is None:
        print("스프레드시트(gviz)에서 탭을 내려받는 중...")
        return {name: fetch_csv(tab) for name, tab in SHEETS.items()}

    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover - 환경 의존
        raise RuntimeError(
            "--xlsx 모드에는 openpyxl이 필요합니다: "
            "uv run --with openpyxl python agent/scripts/sync_config_from_sheets.py"
        ) from e

    print(f"로컬 xlsx에서 탭을 읽는 중: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    raw = {}
    for name, tab in SHEETS.items():
        if tab not in wb.sheetnames:
            raise RuntimeError(f"xlsx에 '{tab}' 탭이 없습니다.")
        matrix = [list(row) for row in wb[tab].iter_rows(values_only=True)]
        raw[name] = _rows_from_matrix(matrix)
    return raw


# ---------------------------------------------------------------------------
# 변환 유틸
# ---------------------------------------------------------------------------


def to_bool(value: str) -> bool:
    return str(value).strip().upper() == "TRUE"


def to_list(value: str) -> list[str]:
    """콤마 / | / 개행으로 구분된 문자열을 리스트로. 빈 항목 제외."""
    if not value:
        return []
    parts = re.split(r"[,|\n]", str(value))
    return [p.strip() for p in parts if p.strip()]


def clean(value):
    """문자열 양끝 공백 제거. None은 그대로."""
    return value.strip() if isinstance(value, str) else value


_BOOL_COLS = {
    "risk_levels": [
        "approval_required",
        "additional_auth_required",
        "audit_required",
        "allow_tool_execution",
    ],
    "guardrail_rules": ["enabled"],
}
_LIST_COLS = {
    "tools": ["input_state_keys", "write_state_keys"],
    "guardrail_rules": ["applies_to_ids"],
}

# 시트에서 흘러들어오는 잡동사니 컬럼은 YAML에서 제외한다
_JUNK_COL = re.compile(r"^(notes( \d+)?|notes_\d+|\d+열(_\d+)?)$")

# Workflow Step 탭에서 사람의 연동 검토를 위해 사용하는 컬럼이다.
# Workflow Step XLSX에는 유지하지만 런타임 workflows.yaml에는 포함하지 않는다.
_STEP_MANAGEMENT_COLS = {
    "interaction_mode",
    "execution_owner",
    "backend_api_method",
    "backend_api_path",
    "request_fields",
    "response_fields",
    "webhook_event_type",
    "webhook_method",
    "webhook_path",
    "prompt_for",
    "frontend_endpoint",
    "resume_type",
    "wait_for_resume",
    "direct_frontend_call",
    "direct_ledger_call",
    "integration_notes",
}


def _convert_row(row: dict, sheet: str) -> dict:
    """한 행에 bool/list 변환을 적용한 dict를 만든다(컬럼 순서 보존)."""
    bool_cols = set(_BOOL_COLS.get(sheet, []))
    list_cols = set(_LIST_COLS.get(sheet, []))
    out = {}
    for key, value in row.items():
        if not key or _JUNK_COL.match(key):
            continue
        if sheet == "steps" and key in _STEP_MANAGEMENT_COLS:
            continue
        if key in bool_cols:
            out[key] = to_bool(value)
        elif key in list_cols:
            out[key] = to_list(value)
        else:
            out[key] = clean(value)
    return out


def _keyed(rows: list[dict], id_col: str, sheet: str) -> dict:
    """행 리스트를 id_col 기준 매핑으로. 중복 id는 첫 행 유지 + 경고."""
    result = {}
    for row in rows:
        rid = clean(row.get(id_col))
        if not rid:
            continue
        body = _convert_row(row, sheet)
        body.pop(id_col, None)
        if rid in result:
            warn(
                f"{sheet} 중복 {id_col} '{rid}' — 첫 행 유지, "
                f"변형 행 폐기 (write_state_keys={body.get('write_state_keys')})"
            )
            continue
        result[rid] = body
    return result


def to_int(value, default=None):
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


# ---------------------------------------------------------------------------
# workflows.yaml 빌드
# ---------------------------------------------------------------------------


def _map_output_key(wf_id: str, key: str) -> str:
    """Step 시트의 flat output_data_key를 네임스페이스 키로 매핑한다."""
    if not key:
        return key
    if "." in key or key in SYSTEM_OUTPUT_KEYS:
        return key
    mapped = OUTPUT_KEY_MAP.get(wf_id, {}).get(key)
    if mapped:
        return mapped
    namespace = WORKFLOW_NAMESPACE.get(wf_id)
    if namespace:
        auto = f"{namespace}.{key}"
        warn(f"[{wf_id}] 매핑에 없는 output_data_key '{key}' → '{auto}' 자동 접두")
        return auto
    warn(f"[{wf_id}] 네임스페이스 미지정 워크플로우의 output_data_key '{key}' 유지")
    return key


def _group_by_workflow(rows: list[dict], sheet: str) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = {}
    for raw in rows:
        wf_id = clean(raw.get("workflow_id"))
        if not wf_id:
            continue
        item = _convert_row(raw, sheet)
        item.pop("workflow_id", None)
        grouped.setdefault(wf_id, []).append(item)
    return grouped


def _harden_routes(wf_id: str, steps: list[dict], routes: list[dict]) -> list[dict]:
    """라우트 정합성 보정: 빈 to는 END 합성, 깨진 참조는 드롭, 중복은 첫 행."""
    step_ids = {s.get("step_id") for s in steps}
    seen: set[tuple] = set()
    out = []
    for r in routes:
        from_id = r.get("from_step_id")
        route_key = r.get("route_key")
        to_id = r.get("to_step_id")

        if from_id not in step_ids:
            warn(f"[{wf_id}] 라우트 from_step_id '{from_id}'가 steps에 없음 → 드롭")
            continue
        if not to_id:
            warn(f"[{wf_id}] {from_id}[{route_key}]의 to_step_id 빈 값 → END 합성")
            r = {**r, "to_step_id": "END"}
            to_id = "END"
        if to_id != "END" and to_id not in step_ids:
            warn(f"[{wf_id}] {from_id}[{route_key}] → '{to_id}' 목적지 없음 → 드롭")
            continue
        dedupe_key = (from_id, route_key)
        if dedupe_key in seen:
            warn(f"[{wf_id}] 라우트 중복 ({from_id}, {route_key}) → 첫 행 유지")
            continue
        seen.add(dedupe_key)
        out.append(r)
    return out


def build_workflows(raw: dict[str, list[dict]], tools: dict) -> dict:
    """Workflow 4개 탭을 workflow_id로 조인해 workflows.yaml 데이터를 만든다."""
    workflows = _keyed(raw["workflows"], "workflow_id", "workflows")

    steps_by_wf = _group_by_workflow(raw["steps"], "steps")
    for wf_id, steps in steps_by_wf.items():
        for step in steps:
            step["step_order"] = to_int(step.get("step_order"), step.get("step_order"))
            step["output_data_key"] = _map_output_key(
                wf_id, step.get("output_data_key") or ""
            )
            # input 스텝의 output_data_key가 비어 있으면 Tool_v2 동명 항목의
            # write_state_keys로 백필한다 (예: ask_recipient -> transfer.recipient).
            # 시트에서는 input UI가 Tool_v2 탭에 tool로 기술되어 있기 때문.
            if step.get("step_type") == "input" and not step.get("output_data_key"):
                write_keys = (
                    tools.get(step.get("step_id"), {}).get("write_state_keys") or []
                )
                if write_keys:
                    step["output_data_key"] = write_keys[0]
                    warn(
                        f"[{wf_id}] input 스텝 '{step.get('step_id')}' "
                        f"output_data_key 빈 값 → Tool_v2 백필: '{write_keys[0]}'"
                    )
        steps.sort(key=lambda s: s.get("step_order") or 0)
        workflows.setdefault(wf_id, {})["steps"] = steps

    schema_by_wf = _group_by_workflow(raw["data_schema"], "data_schema")
    for wf_id, schema in schema_by_wf.items():
        workflows.setdefault(wf_id, {})["data_schema"] = schema

    routes_by_wf = _group_by_workflow(raw["routes"], "routes")
    for wf_id, routes in routes_by_wf.items():
        steps = workflows.get(wf_id, {}).get("steps", [])
        workflows.setdefault(wf_id, {})["routes"] = _harden_routes(wf_id, steps, routes)

    return workflows


# ---------------------------------------------------------------------------
# guardrail_rules.yaml 빌드
# ---------------------------------------------------------------------------


def _parse_contains(condition: str) -> list[str]:
    """contains('X') / contains(\"X\") 토큰에서 리터럴만 추출한다."""
    literals = []
    for m in re.finditer(r"contains\(\s*(['\"])(.*?)\1\s*\)", condition):
        literals.append(m.group(2))
    return literals


def build_guardrail_rules(rows: list[dict]) -> dict:
    """contains(...) 조건은 엔진 호환 contains_any로, 나머지는 expression으로."""
    rules = _keyed(rows, "guardrail_rule_id", "guardrail_rules")
    for rule in rules.values():
        raw = rule.get("condition") or ""
        if "contains(" in raw:
            literals = _parse_contains(raw)
            rule["condition"] = {"contains_any": literals} if literals else {"raw": raw}
        else:
            rule["condition"] = {"expression": raw}
        if not (rule.get("risk_level_override") or "").strip():
            rule["risk_level_override"] = None
    return rules


# ---------------------------------------------------------------------------
# 검증
# ---------------------------------------------------------------------------


def validate(generated: dict) -> None:
    """생성 데이터의 상호 참조를 점검한다 (advisory 경고만)."""
    workflows = generated["workflows.yaml"]
    tasks = generated["tasks.yaml"]
    tools = generated["tools.yaml"]
    risk_levels = generated["risk_levels.yaml"]

    try:
        from agent.tools.registry import TOOL_REGISTRY

        registered = set(TOOL_REGISTRY.keys())
    except Exception:  # noqa: BLE001 - 레지스트리 없이도 생성은 가능
        registered = set()
        warn("TOOL_REGISTRY를 import하지 못해 레지스트리 대조를 건너뜀")

    for wf_id, wf in workflows.items():
        step_ids = {s.get("step_id") for s in wf.get("steps", [])}
        for step in wf.get("steps", []):
            tool_id = step.get("tool_id")
            task_id = step.get("task_id")
            if tool_id and tool_id not in tools:
                warn(
                    f"[{wf_id}] step '{step.get('step_id')}'의 tool_id "
                    f"'{tool_id}'가 Tool_v2에 없음"
                )
            if tool_id and registered and tool_id not in registered:
                warn(
                    f"[{wf_id}] step '{step.get('step_id')}'의 tool_id "
                    f"'{tool_id}'가 TOOL_REGISTRY에 미등록(런타임 error 라우팅)"
                )
            if task_id and task_id not in tasks:
                warn(
                    f"[{wf_id}] step '{step.get('step_id')}'의 task_id "
                    f"'{task_id}'가 Task 시트에 없음"
                )
        # Data Schema 정합성: source_step_id 존재 + flat/네임스페이스 중복
        keys_seen: dict[str, str] = {}
        for ds in wf.get("data_schema", []):
            src = ds.get("source_step_id")
            if src and step_ids and src not in step_ids:
                warn(
                    f"[{wf_id}] data_schema '{ds.get('data_key')}'의 "
                    f"source_step_id '{src}'가 steps에 없음"
                )
            key = ds.get("data_key") or ""
            base = key.split(".", 1)[1] if "." in key else key
            if base in keys_seen and keys_seen[base] != key:
                warn(
                    f"[{wf_id}] data_schema에 flat/네임스페이스 중복: "
                    f"'{keys_seen[base]}' vs '{key}'"
                )
            keys_seen.setdefault(base, key)

    for tid, task in tasks.items():
        rl = task.get("risk_level")
        if rl and rl not in risk_levels:
            warn(f"task '{tid}'의 risk_level '{rl}'가 risk_levels에 없음")
    for tid, tool in tools.items():
        rl = tool.get("risk_level")
        if rl and rl not in risk_levels:
            warn(f"tool '{tid}'의 risk_level '{rl}'가 risk_levels에 없음")


# ---------------------------------------------------------------------------
# 쓰기 / main
# ---------------------------------------------------------------------------


def dump_yaml(data: dict) -> str:
    return yaml.safe_dump(
        data, allow_unicode=True, sort_keys=False, default_flow_style=False
    )


def backup_existing(out_dir: str, filenames: list[str]) -> None:
    backup_dir = os.path.join(out_dir, "backup")
    os.makedirs(backup_dir, exist_ok=True)
    for name in filenames:
        src = os.path.join(out_dir, name)
        if os.path.exists(src):
            shutil.copy2(src, os.path.join(backup_dir, name))


def generate(xlsx_path: str | None) -> dict:
    raw = fetch_all(xlsx_path)
    for name, rows in raw.items():
        print(f"  - {name}: {len(rows)} 행")

    # tools를 먼저 빌드한다 — input 스텝 output_data_key 백필에 필요
    tools = _keyed(raw["tools"], "tool_id", "tools")

    return {
        "workflows.yaml": build_workflows(raw, tools),
        "tasks.yaml": _keyed(raw["tasks"], "task_id", "tasks"),
        "tools.yaml": tools,
        "risk_levels.yaml": _keyed(raw["risk_levels"], "risk_level", "risk_levels"),
        "guardrail_rules.yaml": build_guardrail_rules(raw["guardrail_rules"]),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Google Sheets → config YAML 생성")
    parser.add_argument("--dry-run", action="store_true", help="쓰지 않고 요약만")
    parser.add_argument("--out", default=_DEFAULT_OUT, help="출력 디렉토리")
    parser.add_argument("--xlsx", default=None, help="로컬 xlsx 경로(오프라인 모드)")
    args = parser.parse_args()

    generated = generate(args.xlsx)

    print("\n=== 생성 요약 ===")
    for name, data in generated.items():
        print(f"  {name}: {len(data)} 항목")

    print("\n=== 검증 ===")
    validate(generated)
    print(f"경고 총 {len(_warnings)}건 (advisory — 시트 정리 요청 목록)")

    if args.dry_run:
        print("\n[dry-run] 파일을 쓰지 않았습니다.")
        return

    os.makedirs(args.out, exist_ok=True)
    backup_existing(args.out, list(generated.keys()))
    for name, data in generated.items():
        path = os.path.join(args.out, name)
        with open(path, "w", encoding="utf-8") as f:
            f.write(dump_yaml(data))
    print(f"\n{len(generated)}개 YAML을 '{args.out}'에 기록 (기존은 backup/에 백업)")


if __name__ == "__main__":
    sys.exit(main())
